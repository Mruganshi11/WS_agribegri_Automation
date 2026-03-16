import re
import os
import math
import sys
import json
import time
from openpyxl import load_workbook
from datetime import datetime, timedelta
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException, TimeoutException, StaleElementReferenceException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.action_chains import ActionChains

sys.path.append(os.path.dirname(os.path.abspath(__file__)))
import email_helper

# PLACEHOLDERS
SENDER_EMAIL = "dispatch.agribegri@gmail.com"
SENDER_PASSWORD = "nxxn wqse vhdn rcwr"
ROW_LIMIT = 0
download_dir = os.path.abspath(r"Downloads")
if not os.path.exists(download_dir):
    os.makedirs(download_dir)

# Constants from requirements
WEIGHT_LIMIT_GRAM = 7000
AMOUNT_LIMIT = 8000
COST_PER_ORDER = 0.75

BALANCE_FILE = "balance.txt"
ORDERS_COUNT_FILE = "orders_count.txt"
HISTORY_FILE = "processed_orders.json"
SELLER_EXCEL = "seller_pickup.xlsx"

def load_seller_pickup_map():
    if not os.path.exists(SELLER_EXCEL):
        print(f"⚠ {SELLER_EXCEL} not found. Using empty mapping.")
        return {}
    try:
        # User specified: get data from 'Seller Name', 'Pickup Address' columns
        wb = load_workbook(SELLER_EXCEL, data_only=True)
        ws = wb.active
        
        # Find column indices
        headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]
        try:
            seller_idx = headers.index("Seller Name")
            pickup_idx = headers.index("Pickup Address")
        except ValueError:
            print(f"⚠ Column 'Seller Name' or 'Pickup Address' not found in {SELLER_EXCEL}")
            return {}

        mapping = {}
        # Iterate rows starting from the second row
        for row in ws.iter_rows(min_row=2, values_only=True):
            seller = str(row[seller_idx]).strip().lower() if row[seller_idx] else ""
            pickup = str(row[pickup_idx]).strip() if row[pickup_idx] else ""
            if seller and pickup:
                mapping[seller] = pickup
        
        print(f" Loaded {len(mapping)} seller mappings from Excel.")
        return mapping
    except Exception as e:
        print(f" Error loading seller mappings: {e}")
        return {}

SELLER_PICKUP_MAP = load_seller_pickup_map()

# Chrome Setup
current_dir = os.path.dirname(os.path.abspath(__file__))
chrome_driver_path = os.path.join(current_dir, "chromedriver.exe")
chrome_options = webdriver.ChromeOptions()
prefs = {
    "download.default_directory": download_dir,
    "download.prompt_for_download": False,
    "download.directory_upgrade": True,
    "plugins.always_open_pdf_externally": True,
    "safebrowsing.enabled": True
}
chrome_options.add_experimental_option("prefs", prefs)
service = Service(chrome_driver_path)
driver = webdriver.Chrome(service=service, options=chrome_options)

def login_to_agribegri(username, password, otp):
    print(f"DEBUG: Attempting login with:")
    print(f"DEBUG: Username: {username}")
    print(f"DEBUG: Password: {password}")
    print(f"DEBUG: OTP: {otp}")
    
    print(f"🔐 Logging in...")
    driver.get('https://agribegri.com/admin/')
    driver.maximize_window()
    wait = WebDriverWait(driver, 10)

    wait.until(EC.presence_of_element_located((By.ID, "username"))).send_keys(username)
    driver.find_element(By.ID, "btnSubmit").click()
    time.sleep(2)

    pass_el = wait.until(EC.presence_of_element_located((By.ID, "password")))
    pass_el.clear()
    pass_el.send_keys(password)
    driver.find_element(By.ID, "btnSubmit").click()
    time.sleep(2)

    wait.until(EC.presence_of_element_located((By.ID, "otp"))).send_keys(otp)
    driver.find_element(By.ID, "btnSubmit").click()
    
    print("⏳ Waiting 10 seconds to verify login state...")
    time.sleep(10) 
    print("✅ Login process completed.")

def apply_filter(target_order_id=None):
    driver.get('https://agribegri.com/admin/manage_orders.php')
    wait = WebDriverWait(driver, 30)
    print("📅 Applying primary filters...")
    today = datetime.now()
    thirty_days_ago = today - timedelta(days=30)
    from_date_str = thirty_days_ago.strftime("%Y-%m-%d")
    to_date_str = today.strftime("%Y-%m-%d")

    # 1. Select dates
    driver.execute_script(f"document.getElementById('from_date').value = '{from_date_str}';")
    driver.execute_script(f"document.getElementById('to_date').value = '{to_date_str}';")

    # 2. Select 'Confirm'
    print("📋 Selecting 'Confirm' status via UI...")
    try:
        # Click the multiselect button to open it
        dropdown_btn = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[@title='Select Order Status']")))
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", dropdown_btn)
        dropdown_btn.click()
        time.sleep(1)

        # Click the 'Confirm' checkbox/option
        confirm_option = wait.until(EC.element_to_be_clickable((By.XPATH, "//input[@value='Confirm']")))
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", confirm_option)
        confirm_option.click()
        print("✅ 'Confirm' status selected in UI.")
        
        # Close the dropdown by clicking the button again or clicking away
        dropdown_btn.click() 
        time.sleep(1)
    except Exception as e:
        print(f"⚠ UI selection failed, falling back to JS: {e}")
        driver.execute_script("""
            const select = document.getElementById('search_status');
            if (select) {
                for (let i = 0; i < select.options.length; i++) {
                    if (select.options[i].value === 'Confirm') select.options[i].selected = true;
                    else select.options[i].selected = false;
                }
                if (typeof $ !== 'undefined' && $.fn.multiselect) $(select).multiselect('refresh');
                select.dispatchEvent(new Event('change', { bubbles: true }));
            }
        """)

    # 3. Wait for 2 seconds for the registration
    print("⏳ Waiting for selection to register...")
    time.sleep(2)

    # 4. Click on search button
    print("🔘 Triggering Search button...")
    btn = driver.find_element(By.ID, 'srchSubmit')
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btn)
    driver.execute_script("arguments[0].click();", btn)

    # Wait for the table to load before further filtering
    wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "table tbody tr")))
    print("✅ Primary filters applied.")

    if target_order_id:
        print(f"🔍 Searching for single order: {target_order_id} inside results...")
        search_box = wait.until(EC.presence_of_element_located((By.XPATH, "//div[@id='dyntable_filter']//input")))
        search_box.clear()
        search_box.send_keys(target_order_id)
        time.sleep(5)  # Give time for the table to filter
        
        # Verify if the order is actually in the table after search
        try:
            order_found = driver.find_element(By.XPATH, f"//table[@id='dyntable']/tbody/tr[contains(., '{target_order_id}')]")
            print(f"✅ Single order {target_order_id} found.")
        except NoSuchElementException:
            print(f"⚠ Order {target_order_id} NOT found in filtered results. Stopping.")
            return "NOT_FOUND"

    print("✅ Loading results process completed.")

    # Sort Oldest First
    try:
        print("⏳ Sorting by Date (Oldest First)...")
        date_header = wait.until(EC.element_to_be_clickable((By.XPATH, "//table[@id='dyntable']//th[text()='Date']")))
        current_class = date_header.get_attribute("class")
        if "sorting_asc" not in current_class:
            driver.execute_script("arguments[0].click();", date_header)
            time.sleep(3)
            if "sorting_asc" not in date_header.get_attribute("class"):
                driver.execute_script("arguments[0].click();", date_header)
                time.sleep(3)
        print(f"✅ Final Sort: {date_header.get_attribute('class')}")
    except Exception as e:
        print(f"⚠ Failed to sort: {e}")

def click_truck_icons_one_by_one(row_index):
    wait = WebDriverWait(driver, 30)
    order_id_in_row = "Unknown"
    seller_in_row = "Unknown"
    customer_name = "-"
    net_amount = 0

    try:
        rows = wait.until(EC.presence_of_all_elements_located((By.XPATH, "//table[@id='dyntable']/tbody/tr")))
        if row_index >= len(rows): return "END"
        row = rows[row_index]
        row_text = row.text.lower()
        tds = row.find_elements(By.TAG_NAME, "td")
        order_id_in_row = tds[4].text.strip()
        seller_in_row = " ".join(tds[7].text.split()).lower()
    except:
        return "END"

    if "shipping through transport" in row_text:
        return {"status": "Skipped", "order_id": order_id_in_row, "reason": "Shipping through transport"}

    print(f"🚀 Processing Row {row_index}: Order {order_id_in_row}")

    try:
        is_partial_payment = "partial" in row_text and "payment" in row_text
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", row)
        time.sleep(0.5)

        # Surface Modal
        truck_icon = row.find_element(By.CSS_SELECTOR, "a.get_order_id")
        driver.execute_script("arguments[0].click();", truck_icon)
        wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, 'div[data-remodal-id="surface_modal"]')))
        
        # Select Surface
        surf_dropdown = Select(wait.until(EC.element_to_be_clickable((By.ID, "serviceSurface"))))
        surf_dropdown.select_by_visible_text("Agribegri Surface")
        driver.execute_script("document.getElementById('serviceSurface').dispatchEvent(new Event('change', { bubbles: true }));")
        
        wait.until(lambda d: d.execute_script("return document.getElementById('li_pickup_address').style.display !== 'none';"))

        # Pickup Mapping
        pickup_text = SELLER_PICKUP_MAP.get(seller_in_row, seller_in_row)
        
        # --- STRATEGY: STRIP STATES ---
        # The user noted that including states like 'maharashtra' causes search failures.
        states_to_strip = ["maharashtra", "gujarat", "madhya pradesh", "karnataka", "delhi", "punjab", "haryana", "rajasthan", "uttar pradesh", "tamil nadu", "telangana", "andhra pradesh", "west bengal"]
        cleaned_pickup_text = pickup_text.lower()
        for state in states_to_strip:
            if cleaned_pickup_text.endswith(state):
                cleaned_pickup_text = cleaned_pickup_text.replace(state, "").strip()
        
        print(f"DEBUG: Original Target: {pickup_text} | Cleaned Target: {cleaned_pickup_text}")
        print(f"Normalized seller name: {pickup_text}")
        
        pickup_dropdown = Select(driver.find_element(By.ID, "servicePickupAddress"))
        selected_val = None
        
        # Strategy 1: Look for the cleaned name in dropdown options
        for opt in pickup_dropdown.options:
            if cleaned_pickup_text.lower() in opt.text.lower():
                selected_val = opt.get_attribute("value")
                print(f"✅ Strategy 1 Match: '{opt.text}'")
                break
        
        # Strategy 2: If failed, try matching the first two words (often enough for seller identity)
        if not selected_val:
            words = cleaned_pickup_text.split()
            if len(words) >= 2:
                short_text = " ".join(words[:2])
                for opt in pickup_dropdown.options:
                    if short_text.lower() in opt.text.lower():
                        selected_val = opt.get_attribute("value")
                        print(f"✅ Strategy 2 Match (Shortened): '{opt.text}'")
                        break
        
        if not selected_val:
            raise Exception(f"Pickup not found for: {seller_in_row} (Searched: {cleaned_pickup_text})")
        
        pickup_dropdown.select_by_value(selected_val)
        driver.execute_script("document.getElementById('servicePickupAddress').dispatchEvent(new Event('change', { bubbles: true }));")
        time.sleep(1)

        driver.find_element(By.NAME, "submit_surface").click()
        wait.until(EC.element_to_be_clickable((By.ID, "popup_ok"))).click()
        time.sleep(2)

        # Order Details Tab
        view_icon = row.find_element(By.CSS_SELECTOR, "img[title='View']")
        driver.execute_script("arguments[0].click();", view_icon)
        driver.switch_to.window(driver.window_handles[-1])
        agribegri_order_tab = driver.current_window_handle

        # Read Dimensions, Weight, Amount
        dim_text = wait.until(EC.visibility_of_element_located((By.XPATH, "//span[contains(@class,'dimension_lbl')]"))).text.strip()
        dims = re.findall(r"\d+(?:\.\d+)?", dim_text)
        length, breadth, height = dims if len(dims) == 3 else (10, 10, 10)

        weight_per_unit = float(wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "span.weight_lbl"))).text.replace(",", "").strip())
        qty = float(driver.find_element(By.XPATH, "//table[contains(@class,'table')]/tbody/tr").find_elements(By.TAG_NAME, "td")[7].text.strip())
        total_weight = weight_per_unit * qty
        net_amount = float(driver.find_element(By.ID, "span_abo_net_amt").text.replace(",", "").strip())
        customer_name = driver.find_element(By.CSS_SELECTOR, "span.name_lbl").text.strip()
        order_number_raw = driver.find_element(By.XPATH, "//span[contains(text(),'Order Number')]/strong").text.strip()

        print(f"Order Number copied: {order_number_raw}")
        print(f"Customer Name: {customer_name}")
        print(f"Net Amount from Agribegri: {net_amount}")

        # SPLIT LOGIC (7kg / 8000 INR)
        weight_splits = math.ceil(total_weight / WEIGHT_LIMIT_GRAM)
        amount_splits = math.ceil(net_amount / AMOUNT_LIMIT)
        split_count = max(weight_splits, amount_splits)
        per_split_weight = total_weight / split_count
        per_split_amount = round(net_amount / split_count, 2)

        print(f"📦 Order {order_number_raw} | Splits: {split_count} | Weight: {total_weight}g | Amount: ₹{net_amount}")

        # Delhivery Tab
        driver.execute_script("window.open('https://one.delhivery.com/v2/login','_blank')")
        driver.switch_to.window(driver.window_handles[-1])
        
        # Delhivery Login (Skip if dashboard)
        try:
            WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH, "//div[contains(text(),'Domestic')]")))
        except:
            wait.until(EC.visibility_of_element_located((By.NAME, "email"))).send_keys("complain@agribegri.com")
            driver.find_element(By.XPATH, "//button[contains(text(),'Continue')]").click()
            time.sleep(2)
            wait.until(EC.visibility_of_element_located((By.XPATH, "//input[@type='password']"))).send_keys("AGRIBEGRI!@#26SURFACE")
            driver.find_element(By.ID, "kc-login").click()
        
        wait.until(EC.presence_of_element_located((By.XPATH, "//div[contains(text(),'Domestic')]")))
        
        # --- HANDLE POPUPS ---
        time.sleep(3) # Wait for potential popups
        try:
            # Check for common popup close buttons or overlay IDs
            popups = driver.find_elements(By.XPATH, "//*[contains(@id, 'campaign') or contains(@id, 'popup')]//button[contains(@class, 'close') or contains(@aria-label, 'close')]")
            for p in popups:
                if p.is_displayed():
                    driver.execute_script("arguments[0].click();", p)
                    print("✅ Dismissed a popup.")
        except:
            pass

        # Select Surface Service in Delhivery
        trigger = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[contains(@class,'ap-menu-trigger-root')][.//i[contains(@class,'fa-truck')]]")))
        driver.execute_script("arguments[0].click();", trigger)
        
        # Using JS click for service selection to bypass potential overlays
        surface_opt = wait.until(EC.presence_of_element_located((By.XPATH, "//button[.//div[text()='AGRIBEGRI SURFACE']]")))
        driver.execute_script("arguments[0].click();", surface_opt)
        print("✅ AGRIBEGRI SURFACE selected.")
        time.sleep(3)

        # Global Search order
        awb_btn = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[.//span[text()='AWB']]")))
        driver.execute_script("arguments[0].click();", awb_btn)
        
        id_opt = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[.//span[text()='Order ID']]")))
        driver.execute_script("arguments[0].click();", id_opt)
        
        search_in = wait.until(EC.presence_of_element_located((By.XPATH, "//input[contains(@placeholder,'ORDER ID')]")))
        search_in.clear()
        search_in.send_keys(order_number_raw)
        
        result_xpath = f"//div[contains(@class,'cursor-pointer')][.//span[text()='{order_number_raw}']]"
        result_item = wait.until(EC.element_to_be_clickable((By.XPATH, result_xpath)))
        driver.execute_script("arguments[0].click();", result_item)
        print(f"✅ Order {order_number_raw} selected in Delhivery.")
        time.sleep(3)

        # Print Label
        print_btn = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[normalize-space()='Print Shipping Label']")))
        driver.execute_script("arguments[0].click();", print_btn)
        
        # Rename File
        time.sleep(5)
        latest_file = max([os.path.join(download_dir, f) for f in os.listdir(download_dir)], key=os.path.getmtime)
        new_pdf_name = f"{order_number_raw}.pdf"
        new_pdf_path = os.path.join(download_dir, new_pdf_name)
        if os.path.exists(new_pdf_path): os.remove(new_pdf_path)
        os.rename(latest_file, new_pdf_path)

        # Back to Agribegri to update status
        driver.switch_to.window(agribegri_order_tab)
        driver.refresh()
        time.sleep(2)
        
        Select(wait.until(EC.presence_of_element_located((By.ID, "abo_status")))).select_by_visible_text("Packed")
        Select(driver.find_element(By.ID, "abo_packed_reason")).select_by_visible_text("CL Surface")
        if is_partial_payment:
            Select(driver.find_element(By.ID, "abo_payment_type")).select_by_visible_text("Bank Partial Payment Recieved")
        
        driver.find_element(By.NAME, "update_order_status").click()
        wait.until(EC.element_to_be_clickable((By.ID, "popup_ok"))).click()
        time.sleep(2)

        # Upload PDF
        driver.find_element(By.NAME, "label_file").send_keys(new_pdf_path)
        driver.find_element(By.NAME, "update_manifest_file").click()
        wait.until(EC.element_to_be_clickable((By.ID, "popup_ok"))).click()
        
        return {"status": "Processed", "order_id": order_number_raw, "seller": seller_in_row, "amount": net_amount, "customer": customer_name}

    except Exception as e:
        print(f"❌ Error: {e}")
        return {"status": "Error", "order_id": order_id_in_row, "error": str(e)}

def get_balance():
    if not os.path.exists(BALANCE_FILE): return 0.0
    with open(BALANCE_FILE, "r") as f: return float(f.read().strip() or 0)

def update_balance(val):
    with open(BALANCE_FILE, "w") as f: f.write(str(round(val, 2)))

def update_history(entry):
    history = []
    if os.path.exists(HISTORY_FILE):
        with open(HISTORY_FILE, "r") as f:
            try: history = json.load(f)
            except: pass
    entry["timestamp"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    history.append(entry)
    with open(HISTORY_FILE, "w") as f: json.dump(history[-1000:], f, indent=4)
    with open(ORDERS_COUNT_FILE, "w") as f: f.write(str(len([e for e in history if e.get("status") == "Processed"])))

def main():
    if len(sys.argv) < 4:
        print("Usage: python script.py <username> <password> <otp>")
        return
    
    username, password, otp = sys.argv[1], sys.argv[2], sys.argv[3]
    target_order_id = sys.argv[4] if len(sys.argv) > 4 else None
    
    try:
        login_to_agribegri(username, password, otp)
        filter_status = apply_filter(target_order_id)
        if filter_status == "NOT_FOUND":
            return
        
        curr_idx = 0
        while True:
            if get_balance() < COST_PER_ORDER:
                print("🛑 Low Balance!")
                break
            
            result = click_truck_icons_one_by_one(curr_idx)
            if result == "END": break
            
            if isinstance(result, dict):
                if result.get("status") == "Processed":
                    update_balance(get_balance() - COST_PER_ORDER)
                update_history(result)
            
            curr_idx += 1
            # Cleanup tabs
            while len(driver.window_handles) > 1:
                driver.switch_to.window(driver.window_handles[-1])
                driver.close()
            driver.switch_to.window(driver.window_handles[0])
            time.sleep(2)

    finally:
        driver.quit()

if __name__ == "__main__":
    main()
