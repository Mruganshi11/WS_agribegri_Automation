from selenium import webdriver
import time
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException, TimeoutException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
import os
import openpyxl
from openpyxl import Workbook, load_workbook
from selenium.webdriver.chrome.service import Service
import datetime
import math

# Initialize EXCEL LOGGING
EXCEL_FILE = "processing_report.xlsx"
WEIGHT_LIMIT_GRAM = 7000
AMOUNT_LIMIT = 8000



def init_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Report"
        ws.append(["Timestamp", "Row_Index", "Order_ID", "Status", "Reason"])
        wb.save(EXCEL_FILE)

init_excel()

def log_to_excel(row_index, order_id, status, reason):
    try:
        if os.path.exists(EXCEL_FILE):
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Report"
            ws.append(["Timestamp", "Row_Index", "Order_ID", "Status", "Reason"])
        
        ws.append([
            datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            row_index,
            order_id,
            status,
            reason
        ])
        wb.save(EXCEL_FILE)
    except Exception as e:
        print(f"⚠ Failed to write to Excel: {e}")
from selenium.webdriver.common.action_chains import ActionChains
import sys
sys.path.append(os.path.dirname(os.path.abspath(__file__)))
import email_helper

# PLACEHOLDERS (User to fill)
SENDER_EMAIL = "dispatch.agribegri@gmail.com"
SENDER_PASSWORD = "nxxn wqse vhdn rcwr"
ROW_LIMIT = 0  # 0 for ALL rows, or set a number (e.g. 5)
download_dir = os.path.abspath(r"Downloads")

SELLER_PICKUP_MAP = {
    "rk chemicals gujarat": "dobariya sunil Of RK chemicals",
    "geolife agritech india pvt. ltd maharashtra": "Geolife Agritech India Pvt. Ltd",
    "urja agriculture company delhi": "1Urja Agriculture Company",
    "sickle innovations private limited gujarat": "sickle",
    "agribegri trade link pvt. ltd. gujarat": "Godawon",
    "atpl": "Neptune",
    "sagar biotech pvt ltd": "Sagar Biotech Pvt Ltd",
    "essential biosciences": "Essential Biosciences",
    "piyush kataria rajasthan": "PIYUSH KATARIA",
    "agribegri trade link pvt. ltd.":"Godawon",
    "Rahul gurg PUNJAB":"Rahul Gurg"
}


# Set up Chrome options
chrome_options = webdriver.ChromeOptions()
prefs = {
    "download.default_directory": download_dir,
    "download.prompt_for_download": False,
    "download.directory_upgrade": True,
    "plugins.always_open_pdf_externally": True,

    "safebrowsing.enabled": True
}
chrome_options.add_experimental_option("prefs", prefs)

service = Service(r"C:\Users\ghara\Downloads\agribegri\chromedriver.exe")
# service = Service(r"C:\inetpub\wwwroot\AI Python\AgriBegri CRON\chromedriver.exe")
driver = webdriver.Chrome(service=service, options=chrome_options)

def login_to_agribegri():
    driver.get('https://agribegri.com/admin/')
    driver.maximize_window()

    username_field = driver.find_element(By.ID, "username").send_keys('Namrata')
    next_button = driver.find_element(By.ID, "btnSubmit")
    next_button.click()
    time.sleep(3)

    password_field = driver.find_element(By.ID, "password").send_keys('Websmith$123456')
    pass_next_button = driver.find_element(By.ID, "btnSubmit")
    pass_next_button.click()
    time.sleep(3)

    enter_otp = driver.find_element(By.ID, "otp").send_keys('123456')
    sign_in = driver.find_element(By.ID, "btnSubmit")
    sign_in.click()
    time.sleep(2)

def apply_filter():
    driver.get('https://agribegri.com/admin/manage_orders.php')

    # Wait for page to load completely
    time.sleep(3)

    wait = WebDriverWait(driver, 30)

    print("📋 Waiting for orders page to load...")

        # ================== SET DATE FILTER (FROM = 30 DAYS AGO, TO = TODAY) ==================

    from datetime import datetime, timedelta

    today_date = datetime.now()
    from_date_value = (today_date - timedelta(days=30)).strftime("%Y-%m-%d")
    to_date_value = today_date.strftime("%Y-%m-%d")

    print(f"📅 Setting From Date: {from_date_value}")
    print(f"📅 Setting To Date: {to_date_value}")

    # Set From Date using JS (because input is readonly)
    driver.execute_script("""
        let fromInput = document.getElementById('from_date');
        fromInput.removeAttribute('readonly');
        fromInput.value = arguments[0];
        fromInput.dispatchEvent(new Event('change', { bubbles: true }));
    """, from_date_value)

    time.sleep(0.5)

    # Set To Date using JS
    driver.execute_script("""
        let toInput = document.getElementById('to_date');
        toInput.removeAttribute('readonly');
        toInput.value = arguments[0];
        toInput.dispatchEvent(new Event('change', { bubbles: true }));
    """, to_date_value)

    time.sleep(1)

    print("✅ Date filter applied successfully")


    # phone_input = wait.until(
    #     EC.presence_of_element_located((By.ID, 'srchby_phone'))
    # )
    # phone_input.clear()
    # phone_input.send_keys('6354058079')

    # ================== FILTER BY STATUS: CONFIRM ==================
    print("📋 Selecting Order Status: Confirm...")

    # 1. Click Dropdown Button
    status_dropdown_btn = wait.until(EC.element_to_be_clickable(
        (By.XPATH, "//button[@title='Select Order Status']")
    ))
    status_dropdown_btn.click()
    time.sleep(1)

    # 2. Select 'Confirm'
    confirm_option = wait.until(EC.element_to_be_clickable(
        (By.XPATH, "//input[@value='Confirm']")
    ))
    confirm_option.click()
    print("✅ 'Confirm' status selected")
    time.sleep(0.5)

    # 3. Click Search
    search_button = driver.find_element(By.ID, 'srchSubmit')
    search_button.click()

    # wait for table load
    wait.until(
        EC.presence_of_element_located((By.CSS_SELECTOR, "table tbody tr"))
    )

    print("Search completed, table is stable")

    time.sleep(1)
   

def click_truck_icons_one_by_one(row_index):
    try:

        wait = WebDriverWait(driver, 30)

        # Re-fetch rows every time to avoid Stale Elements
        rows = wait.until(
            EC.presence_of_all_elements_located(
                (By.XPATH, "//table[@id='dyntable']/tbody/tr")
            )
        )

        print(f"Total filtered rows in table: {len(rows)}")

        # Check bounds
        if row_index >= len(rows):
            print(f"⚠ Index {row_index} out of bounds (End of list)")
            return {"status": "END"}

        # Select the target row
        try:
            row = rows[row_index]
            row_text = row.text.lower()
        except IndexError:
            return {"status": "END"}


        # CHECK REMARK
        if "shipping through transport" in row_text:
            print(f" Row {row_index} skipped (Shipping Through Transport)")
            return {"status": "SKIPPED", "reason": "Shipping Through Transport", "order_id": "N/A"}
        else:
            print(f"✅ Row {row_index} selected (Valid for processing)")

        # ================== PAYMENT STATUS CHECK ==================

        is_partial_payment = False

        if "partial" in row_text and "payment" in row_text:
            is_partial_payment = True
            print("💰 Partial payment detected for this order")
        else:
            print("💳 Full / Non-partial payment detected")


        # scroll to row
        driver.execute_script(
            "arguments[0].scrollIntoView({block:'center'});", row
        )
        time.sleep(0.5)

        # click truck icon
        truck_icon = row.find_element(By.CSS_SELECTOR, "a.get_order_id")
        driver.execute_script("arguments[0].click();", truck_icon)

        print(" Truck icon clicked for first row")

        # wait for modal
        wait.until(
            EC.visibility_of_element_located(
            (By.CSS_SELECTOR, 'div[data-remodal-id="surface_modal"]')
        )

        )
        print(" Surface modal opened")


        # click dropdown
        surface_dropdown = wait.until(
            EC.element_to_be_clickable((By.ID, "serviceSurface"))
        )
        
        # 1. Try selecting via Selenium Select class
        print("Trying to select 'Agribegri Surface'...")
        try:
            select_surf = Select(surface_dropdown)
            select_surf.select_by_visible_text("Agribegri Surface")
            print("✅ Selected via Select class")
        except Exception as e:
            print(f"⚠ Select class error: {e}")
        
        # 2. FORCE update value via JS (Backup)
        driver.execute_script("""
            var select = document.getElementById('serviceSurface');
            for(var i=0; i<select.options.length; i++){
                if(select.options[i].text.trim() == 'Agribegri Surface'){
                    select.options[i].selected = true;
                    select.value = select.options[i].value;
                    break;
                }
            }
            select.dispatchEvent(new Event('change', { bubbles: true }));
            select.dispatchEvent(new Event('input', { bubbles: true }));
        """)
        
        time.sleep(1)

        # 3. VERIFY Selection
        selected_option = Select(surface_dropdown).first_selected_option
        print(f"🔍 Current Selection: '{selected_option.text}'")

        if "Select One" in selected_option.text:
            print("❌ Selection failed! Retrying with explicit index...")
            # fallback by index if text fails (assuming it's usually 2nd or 3rd option)
            try:
                Select(surface_dropdown).select_by_index(1) 
                print("✅ Selected by index 1")
            except:
                pass

        #  force onchange (CRITICAL)
        driver.execute_script("""
            const select = document.getElementById('serviceSurface');
            select.dispatchEvent(new Event('change', { bubbles: true }));
        """)

        print(" Agribegri Surface selected")


        
        #  correct pickup visibility wait
        wait.until(
            lambda d: d.execute_script(
                "return document.getElementById('li_pickup_address').style.display !== 'none';"
            )
        )


        # ================== PICKUP ADDRESS SELECTION START ==================

        # get seller name from table row (7th column)
        seller_raw = row.find_elements(By.TAG_NAME, "td")[7].text
        seller_name = " ".join(seller_raw.split()).lower()


        print(" Normalized seller name:", seller_name)

        # decide pickup address (mapping + fallback)
        pickup_text = SELLER_PICKUP_MAP.get(seller_name, seller_name)
        print(" Target Pickup Name:", pickup_text)


        # select pickup address from dropdown
        pickup_dropdown = wait.until(
            EC.element_to_be_clickable((By.ID, "servicePickupAddress"))
        )

        options = pickup_dropdown.find_elements(By.TAG_NAME, "option")

        selected_value = None
        matched_text = ""

        # --- STRATEGY 1: DIRECT MAPPING / EXACT CONTAINMENT ---
        for opt in options:
            if pickup_text.lower() in opt.text.lower():
                selected_value = opt.get_attribute("value")
                matched_text = opt.text
                print(f" ✅ Strategy 1 Match: '{matched_text}' (value={selected_value})")
                break

        # --- STRATEGY 2: REMOVE STATE SUFFIXES (common issue) ---
        if not selected_value:
            print(" ⚠ Strategy 1 failed. Trying Strategy 2 (Remove State Suffix)...")
            # List of common states to strip from the end
            states = ["maharashtra", "gujarat", "madhya pradesh", "karnataka", "delhi", "punjab", "haryana", "rajasthan", "uttar pradesh", "tamil nadu", "telangana", "andhra pradesh", "west bengal"]
            
            cleaned_name = seller_name
            for state in states:
                if cleaned_name.endswith(state):
                    cleaned_name = cleaned_name.replace(state, "").strip()
                    break # Remove only the last matching state
            
            print(f" Cleaned Name: '{cleaned_name}'")

            if len(cleaned_name) > 3: # Ensure we don't match empty string
                for opt in options:
                    if cleaned_name in opt.text.lower():
                        selected_value = opt.get_attribute("value")
                        matched_text = opt.text
                        pickup_text = cleaned_name # Update pickup_text for Delhivery search
                        print(f" ✅ Strategy 2 Match: '{matched_text}' (value={selected_value})")
                        break

        # --- STRATEGY 3: FIRST WORD MATCH (Last Resort) ---
        if not selected_value:
            print(" ⚠ Strategy 2 failed. Trying Strategy 3 (First Word Match)...")
            first_word = seller_name.split()[0]
            if len(first_word) > 3:
                for opt in options:
                    if first_word in opt.text.lower():
                        selected_value = opt.get_attribute("value")
                        matched_text = opt.text
                        pickup_text = first_word # Update pickup_text for Delhivery search
                        print(f" ✅ Strategy 3 Match: '{matched_text}' (value={selected_value})")
                        break

        if not selected_value:
            # debug: print available options
            print("❌ FAILED. Available Options were:")
            for opt in options:
                if opt.text.strip():
                    print(f" - {opt.text}")
            
            raise Exception(f" Pickup address not found for seller: {seller_name} (Target: {pickup_text})")

        #  SET VALUE DIRECTLY (Using Select Class now for better reliability)
        try:
            # User Select class (Global import)
            
            # 1. Click the dropdown first to ensure focus
            driver.execute_script("arguments[0].click();", pickup_dropdown)
            time.sleep(0.5)

            # 2. Use Select class
            select_elem = Select(pickup_dropdown)
            select_elem.select_by_value(selected_value)
            print(f"✅ Selected value {selected_value} using Select class")
            
            # 3. Explicitly trigger events just in case
            driver.execute_script("""
                const select = arguments[0];
                select.dispatchEvent(new Event('change', { bubbles: true }));
                select.dispatchEvent(new Event('input', { bubbles: true }));
                select.dispatchEvent(new Event('blur', { bubbles: true }));
            """, pickup_dropdown)

        except Exception as e:
            print(f"⚠ Select class failed, trying JS fallback: {e}")
            driver.execute_script("""
                const select = document.getElementById('servicePickupAddress');
                select.value = arguments[0];
                select.dispatchEvent(new Event('change', { bubbles: true }));
                select.dispatchEvent(new Event('input', { bubbles: true }));
                select.dispatchEvent(new Event('blur', { bubbles: true }));
            """, selected_value)

        print("✅ Pickup address selected successfully")
        time.sleep(1)

        # REMOVED "RE-SELECT SURFACE" BLOCK causing reset issues


        # ================== SUBMIT SURFACE FORM ==================

        submit_btn = wait.until(
            EC.element_to_be_clickable((By.NAME, "submit_surface"))
        )

        driver.execute_script("arguments[0].click();", submit_btn)

        print(" Submit button clicked successfully")

        # ================== SUCCESS POPUP HANDLING ==================

        # wait for success popup
        ok_button = wait.until(
            EC.element_to_be_clickable((By.ID, "popup_ok"))
        )

        driver.execute_script("arguments[0].click();", ok_button)

        print(" Success popup OK clicked")

        # optional: small wait for backend processing
        time.sleep(2)

        # ================== OPEN ORDER DETAIL (ONCE) ==================

        view_icon = row.find_element(By.CSS_SELECTOR, "img[title='View']")
        driver.execute_script("arguments[0].click();", view_icon)
        driver.switch_to.window(driver.window_handles[-1])

        agribegri_order_tab = driver.current_window_handle
        print("📌 Order detail tab opened & stored")

        # ================== READ DIMENSION (CORRECT PLACE) ==================

        import re   # make sure this is at TOP of file ideally

        dimension_text = WebDriverWait(driver, 20).until(
            EC.visibility_of_element_located((
                By.XPATH, "//span[contains(@class,'dimension_lbl')]"
            ))
        ).text.strip()

        print(f"📐 Dimension found: {dimension_text}")

        values = re.findall(r"\d+(?:\.\d+)?", dimension_text)

        if len(values) != 3:
            raise Exception(f"❌ Invalid dimension format: {dimension_text}")

        length, breadth, height = values

        print(f"➡ Parsed: L={length}, B={breadth}, H={height}")



        # ================== POST-SURFACE WEIGHT CHECK ==================

        # if not is_partial_payment:
        #     print("⚖ Performing weight check (non-partial payment)")

        #     WEIGHT_LIMIT_GRAM = 13000

        #     split_count = math.ceil(total_weight / WEIGHT_LIMIT_GRAM)
        #     per_split_weight = total_weight / split_count

        #     print(f"🔀 Total shipments required: {split_count}")
        #     print(f"📦 Per shipment weight: {per_split_weight:.2f} g")

        #     # open order detail page
        #     view_icon = row.find_element(By.CSS_SELECTOR, "img[title='View']")
        #     driver.execute_script("arguments[0].click();", view_icon)
        #     driver.switch_to.window(driver.window_handles[-1])

        #     wait = WebDriverWait(driver, 20)

        #     weight_elem = wait.until(
        #         EC.presence_of_element_located((By.CSS_SELECTOR, "span.weight_lbl"))
        #     )
        #     weight_per_unit = float(weight_elem.text.replace(",", "").strip())

        #     product_row = wait.until(
        #         EC.presence_of_element_located((
        #             By.XPATH,
        #             "//table[contains(@class,'table')]/tbody/tr"
        #         ))
        #     )

        #     quantity = float(product_row.find_elements(By.TAG_NAME, "td")[7].text.strip())

        #     total_weight = weight_per_unit * quantity

        #     print(f"📦 Total weight: {total_weight} g")

        #     print("✅ Weight within limit")

        #     driver.close()
        #     driver.switch_to.window(driver.window_handles[0])
        # else:
        #     print("ℹ Partial payment order — skipping weight check")

        if not is_partial_payment:
            print("⚖ Performing weight & amount split check")

            # ---------- READ WEIGHT ----------
            weight_elem = wait.until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "span.weight_lbl"))
            )
            weight_per_unit = float(weight_elem.text.replace(",", "").strip())

            product_row = wait.until(
                EC.presence_of_element_located((
                    By.XPATH,
                    "//table[contains(@class,'table')]/tbody/tr"
                ))
            )

            quantity = float(product_row.find_elements(By.TAG_NAME, "td")[7].text.strip())

            total_weight = weight_per_unit * quantity
            print(f"📦 Total weight: {total_weight} g")

            # ---------- READ AMOUNT ----------
            net_amount_elem = wait.until(
                EC.presence_of_element_located((By.ID, "span_abo_net_amt"))
            )
            net_amount = float(net_amount_elem.text.replace(",", "").strip())

            
            print(f"💰 Net Amount: {net_amount}")

            # ---------- CALCULATE SPLITS ----------
            weight_split = math.ceil(total_weight / WEIGHT_LIMIT_GRAM)
            amount_split = math.ceil(net_amount / AMOUNT_LIMIT)

            split_count = max(weight_split, amount_split)

            per_split_weight = total_weight / split_count
            per_split_amount = net_amount / split_count

            print(f"🔀 Final Split Count: {split_count}")
            print(f"📦 Per shipment weight: {per_split_weight:.2f} g")
            print(f"💰 Per shipment amount: {per_split_amount:.2f}")

        else:
            print("ℹ Partial payment order — skipping weight check")
            split_count = 1
            per_split_weight = None


        # ================== CLICK VIEW ICON ==================

        # view_icon = row.find_element(By.CSS_SELECTOR, "img[title='View']")
        # driver.execute_script("arguments[0].click();", view_icon)

        # print(" View icon clicked (new tab opened)")

        # ================== SWITCH TO VIEW TAB ==================

        driver.switch_to.window(driver.window_handles[-1])
        print(" Switched to order detail tab")

        agribegri_order_tab = driver.current_window_handle
        print("✅ Agribegri order tab stored")


        # ================== COPY ORDER NUMBER (FIXED) ==================
        order_number_elem = WebDriverWait(driver, 20).until(
            EC.presence_of_element_located(
                (By.XPATH, "//span[contains(text(),'Order Number')]/strong")
            )
        )



        order_number = order_number_elem.text.strip()
        print(" Order Number copied:", order_number)

        # ================== READ NET AMOUNT FROM AGRIBEGRI ==================

        driver.switch_to.window(agribegri_order_tab)

        net_amount_elem = WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.ID, "span_abo_net_amt"))
        )

        net_amount = float(net_amount_elem.text.replace(",", "").strip())

        print(f"💰 Net Amount from Agribegri: {net_amount}")

        # calculate per-split amount
        per_split_amount = round(net_amount / split_count, 2)
        print(f"💰 Per shipment amount: {per_split_amount}")

        # switch back to Delhivery tab
        driver.switch_to.window(driver.window_handles[-1])


        agribegri_edit_url = driver.current_url
        print("✅ Stored Agribegri edit URL:", agribegri_edit_url)
        

        # ================== EXTRACT SHIPPING ADDRESS DETAILS ==================

        print("📋 Extracting shipping address details...")

        # Name
        customer_name = wait.until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "span.name_lbl"))
        ).text.strip()

        # Email
        try:
            # Email is in a <td> element within the shipping address table
            email_elem = driver.find_element(
                By.XPATH,
                "//td[text()='Email :']/following-sibling::td"
            )
            customer_email = email_elem.text.strip()
        except:
            customer_email = ""  # Optional field

        # Address
        customer_address = driver.find_element(By.CSS_SELECTOR, "span.address_lbl").text.strip()

        # Landmark (optional - might be empty)
        try:
            customer_landmark = driver.find_element(By.CSS_SELECTOR, "span.landmark_lbl").text.strip()
        except:
            customer_landmark = ""

        # Pincode
        customer_pincode = driver.find_element(By.CSS_SELECTOR, "span.zipcode_lbl").text.strip()

        # Phone
        customer_phone = driver.find_element(By.CSS_SELECTOR, "span.phone_lbl").text.strip()

        # City
        customer_city = driver.find_element(By.CSS_SELECTOR, "span.city_lbl").text.strip()

        # Taluka (optional)
        try:
            customer_taluka = driver.find_element(By.CSS_SELECTOR, "span.taluka_lbl").text.strip()
        except:
            customer_taluka = ""

        # District (optional)
        try:
            customer_district = driver.find_element(By.CSS_SELECTOR, "span.district_lbl").text.strip()
        except:
            customer_district = ""

        # State
        customer_state = driver.find_element(By.CSS_SELECTOR, "span.state_lbl").text.strip()

        print(f"📦 Customer Name: {customer_name}")
        print(f"📦 Customer Email: {customer_email}")
        print(f"📦 Customer Address: {customer_address}")
        print(f"📦 Customer Pincode: {customer_pincode}")
        print(f"📦 Customer Phone: {customer_phone}")
        print(f"📦 Customer City: {customer_city}")
        print(f"📦 Customer Taluka: {customer_taluka}")
        print(f"📦 Customer District: {customer_district}")
        print(f"📦 Customer State: {customer_state}")

        # ================== EXTRACT PRODUCT DESCRIPTION ==================

        print("📋 Extracting product description...")

        # Product description from the table (column 2 - index 1)
        product_row = wait.until(
            EC.presence_of_element_located((
                By.XPATH,
                "//table[contains(@class,'table')]/tbody/tr"
            ))
        )

        # Product name is in the 2nd column (index 1)
        product_description = product_row.find_elements(By.TAG_NAME, "td")[1].text.strip()

        print(f"📦 Product Description: {product_description}")

        # ================== EXTRACT SELLER EMAIL (New) ==================
        print("📋 Extracting Seller Email...")
        try:
            seller_email_elem = driver.find_element(
                By.XPATH,
                "//h5[text()='Seller Details']/following-sibling::table//td[text()='Email :']/following-sibling::td"
            )
            seller_email = seller_email_elem.text.strip()
            print(f"📧 Seller Email found: {seller_email}")
        except:
            seller_email = ""
            print("⚠ Seller Email not found")

        # ================== OPEN DELHIVERY ==================

        driver.execute_script("window.open('https://one.delhivery.com/v2/login','_blank')")
        driver.switch_to.window(driver.window_handles[-1])
        print("🚚 Delhivery tab opened")

        # ================== DELHIVERY LOGIN (WITH AUTO-SKIP) ==================

        print("Checking if already logged into Delhivery...")
        try:
            # Check if we are already on the dashboard (look for Domestic dropdown or similar)
            WebDriverWait(driver, 5).until(
                EC.presence_of_element_located((By.XPATH, "//div[contains(text(),'Domestic')]"))
            )
            print("✅ Already logged in - Skipping login steps")
        
        except TimeoutException:
            print("ℹ Not logged in. Proceeding with login...")

            # 1️ Enter email
            email_input = wait.until(
                EC.visibility_of_element_located((By.NAME, "email"))
            )

            email_input.clear()
            email_input.send_keys("complain@agribegri.com")

            print(" Email entered")

            # 2️ Click Continue
            time.sleep(1.5)

            continue_btn = wait.until(
                EC.presence_of_element_located(
                    (By.XPATH, "//button[contains(text(),'Continue')]")
                )
            )

            driver.execute_script("arguments[0].click();", continue_btn)

            print(" Continue clicked")

            # 3️ Wait for password field to appear
            time.sleep(1.5)

            password_input = wait.until(
                EC.visibility_of_element_located((By.XPATH, "//input[@type='password']"))
            )

            password_input.send_keys("AGRIBEGRI!@#26SURFACE")

            print(" Password entered")

            # 5️ Click Login button (Keycloak page)
            login_btn = wait.until(
                EC.presence_of_element_located((By.ID, "kc-login"))
            )

            driver.execute_script("arguments[0].click();", login_btn)

            print(" Login button clicked")


        # ================== SELECT AGRIBEGRI SURFACE (TOP-RIGHT DROPDOWN) ==================

        # wait for Delhivery dashboard to load
        WebDriverWait(driver, 30).until(
            EC.presence_of_element_located(
                (By.XPATH, "//div[contains(text(),'Domestic')]")
            )
        )


    # ================== CLICK DOMESTIC / AGRIBEGRI SURFACE DROPDOWN ==================

        domestic_dropdown = WebDriverWait(driver, 30).until(
            EC.element_to_be_clickable((
                By.XPATH,
                "//button[contains(@class,'ap-menu-trigger-root')]"
                "[.//i[contains(@class,'fa-truck')]]"
            ))
        )

        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", domestic_dropdown)
        time.sleep(0.5)

        driver.execute_script("arguments[0].click();", domestic_dropdown)
        print("✅ Domestic dropdown clicked")



            # ================== SELECT AGRIBEGRI SURFACE ==================

        agribegri_surface = WebDriverWait(driver, 30).until(
            EC.element_to_be_clickable((
                By.XPATH,
                "//button[contains(@class,'ap-menu-item')]"
                "[.//div[text()='AGRIBEGRI SURFACE']]"
            ))
        )

        driver.execute_script("arguments[0].click();", agribegri_surface)
        print("✅ AGRIBEGRI SURFACE selected")

        time.sleep(1.5)
        
        # ================== WEIGHT > 13 KG SPECIAL FLOW ==================

        if split_count > 1:


            print("🛑 Weight > 13 KG detected — moving to Forward Shipments")

            # -------- EXPAND SIDEBAR FIRST --------
            print("🔓 Expanding sidebar...")
            try:
                sidebar_toggle = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((
                        By.XPATH,
                        "//div[contains(@class, 'ap-sidebar-parent--left__toggler')]//i[contains(@class, 'fa-angles-right')]"
                    ))
                )
                driver.execute_script("arguments[0].click();", sidebar_toggle)
                print("✅ Sidebar expanded")
                time.sleep(1)
            except:
                print("ℹ Sidebar already expanded or toggle not found")

            # -------- Click Shipments & Pickups (sidebar) --------
            shipments_menu = WebDriverWait(driver, 30).until(
                EC.element_to_be_clickable((
                    By.XPATH,
                    "//button[.//i[contains(@class,'fa-box-circle-check')] and contains(., 'Shipments')]"
                ))
            )
            
            # Scroll menu into view
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", shipments_menu)
            time.sleep(0.5)
            
            driver.execute_script("arguments[0].click();", shipments_menu)
            print("📦 Shipments & Pickups menu clicked")


            time.sleep(2)  # Give menu time to expand

            # -------- Find and Click Forward Shipments --------
            forward_shipments = None
            attempts = 0
            max_attempts = 3

            while not forward_shipments and attempts < max_attempts:
                attempts += 1
                print(f"🔍 Attempt {attempts} to find Forward Shipments...")

                try:
                    # Try finding the <a> tag specifically that contains "Forward Shipments"
                    forward_shipments = WebDriverWait(driver, 5).until(
                        EC.presence_of_element_located((
                            By.XPATH,
                            "//a[contains(@class, 'ap-sidebar__item') and .//div[contains(text(), 'Forward Shipments')]]"
                        ))
                    )
                    
                    print("✅ Forward Shipments link (anchor tag) found!")
                    
                    # Check if element is visible
                    if not forward_shipments.is_displayed():
                        print("⚠ Element found but not visible, scrolling...")
                        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", forward_shipments)
                        time.sleep(1)
                    
                    break

                except TimeoutException:
                    print(f"⚠ Attempt {attempts} failed, trying alternative selector...")
                    
                    # Try case-insensitive fallback
                    try:
                        forward_shipments = WebDriverWait(driver, 5).until(
                            EC.presence_of_element_located((
                                By.XPATH,
                                "//a[contains(@class, 'sidebar__item')]//div[contains(translate(., 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), 'forward')]/ancestor::a"
                            ))
                        )
                        print("✅ Forward Shipments found via fallback selector!")
                        break
                    except:
                        pass
                    
                    if attempts < max_attempts:
                        # Try clicking menu again
                        print("🔄 Clicking Shipments & Pickups again...")
                        driver.execute_script("arguments[0].click();", shipments_menu)
                        time.sleep(2)
                    else:
                        # Last attempt - detailed debugging
                        print("\n❌ Final attempt failed. Running diagnostics...")
                        
                        # Check page source for "Forward"
                        page_source = driver.page_source.lower()
                        if 'forward' in page_source:
                            print("✅ 'forward' text found in page source")
                            if 'forward shipments' in page_source:
                                print("✅ 'forward shipments' text found in page source")
                        else:
                            print("❌ 'forward' text NOT found in page source")
                        
                        # Print all visible menu items
                        print("\nDEBUG: Searching for menu items...")
                        menu_items = driver.find_elements(By.XPATH, "//div[contains(@class, 'menu') or contains(@class, 'item')]")
                        visible_menu_items = [item for item in menu_items if item.is_displayed() and item.text.strip()]
                        print(f"Found {len(visible_menu_items)} visible menu items:")
                        for item in visible_menu_items[:15]:
                            print(f" - {item.text.strip()[:60]}")
                        
                        raise Exception("Failed to navigate to Forward Shipments after all attempts")

            # Always click Forward Shipments to navigate to that page
            print("🖱 Clicking 'Forward Shipments' link...")
            
            # Scroll into view first
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", forward_shipments)
            time.sleep(0.5)
            
            # Click it
            driver.execute_script("arguments[0].click();", forward_shipments)
            print("✅ Forward Shipments clicked")
            
            # Wait for page to load
            print("⏳ Waiting for Forward Shipments page to load...")
            time.sleep(3)


            # ================== CLICK CREATE FORWARD SHIPMENT ==================

            print("🔍 Looking for 'Create Forward Shipment' button...")
            create_forward_btn = None
            
            # Try multiple selectors
            selectors = [
                ("specific class match", "//button[contains(@class, 'ap-button') and contains(@class, 'blue') and contains(., 'Create Forward Shipment')]"),
                ("data-action", "//button[@data-action='create-shipment']"),
                ("text contains", "//button[contains(text(), 'Create Forward Shipment')]"),
                ("text contains dot", "//button[contains(., 'Create Forward Shipment')]"),
                ("any create button", "//button[contains(@class, 'blue') and contains(., 'Create')]")
            ]
            
            for desc, xpath in selectors:
                try:
                    print(f"  Trying selector: {desc}...")
                    create_forward_btn = WebDriverWait(driver, 10).until(
                        EC.element_to_be_clickable((By.XPATH, xpath))
                    )
                    print(f"  ✅ Found button using: {desc}")
                    break
                except TimeoutException:
                    print(f"  ❌ {desc} failed")
                    continue
            
            if not create_forward_btn:
                print("\n❌ CRITICAL: Could not find 'Create Forward Shipment' button")
                print("DEBUG: Searching for all buttons on page...")
                buttons = driver.find_elements(By.TAG_NAME, "button")
                print(f"Found {len(buttons)} buttons:")
                for idx, btn in enumerate(buttons[:15]):
                    try:
                        btn_text = btn.text.strip()[:60]
                        data_action = btn.get_attribute("data-action")
                        visible = btn.is_displayed()
                        print(f"  Button {idx+1}: text='{btn_text}', data-action='{data_action}', visible={visible}")
                    except:
                        pass
                raise Exception("Create Forward Shipment button not found")
            
            # Click the button
            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", create_forward_btn)
            time.sleep(0.5)
            driver.execute_script("arguments[0].click();", create_forward_btn)
            print("🚀 Create Forward Shipment button clicked")

            time.sleep(3)

            print("✅ Navigation to Create Forward Shipment form successful.")
            
            # ================== FILL DELHIVERY CREATE FORWARD SHIPMENT FORM ==================

            print("📝 Filling Delhivery shipment form...")

            wait = WebDriverWait(driver, 20)

            # -------- ORDER ID --------
            # ================== FILL CREATE FORWARD SHIPMENT FORM ==================

            for i in range(1, split_count + 1):

                print(f"🧩 Creating shipment part {i}/{split_count}")

                # -------- Wait for page to fully load --------
                print("⏳ Waiting for form to load...")
                time.sleep(3)  # Give page time to render

                # -------- Order ID / Reference Number --------
                print("🔍 Looking for Order ID input field...")
                
                try:
                    order_id_input = WebDriverWait(driver, 30).until(
                        EC.presence_of_element_located((
                            By.XPATH,
                            "//input[@placeholder='Enter Order ID / Reference Number']"
                        ))
                    )
                    print("✅ Order ID input field found!")
                    
                except TimeoutException:
                    print("❌ Order ID input field not found within 30 seconds")
                    print("DEBUG: Searching for all input fields on the page...")
                    
                    # Find all input fields
                    all_inputs = driver.find_elements(By.TAG_NAME, "input")
                    print(f"Found {len(all_inputs)} input fields:")
                    for idx, inp in enumerate(all_inputs[:10]):
                        try:
                            placeholder = inp.get_attribute("placeholder")
                            inp_type = inp.get_attribute("type")
                            visible = inp.is_displayed()
                            print(f"  Input {idx+1}: type='{inp_type}', placeholder='{placeholder}', visible={visible}")
                        except:
                            pass
                    
                    # Try alternative selectors
                    print("\nTrying alternative selectors...")
                    
                    # Try by partial placeholder match
                    try:
                        order_id_input = driver.find_element(
                            By.XPATH,
                            "//input[contains(@placeholder, 'Order ID') or contains(@placeholder, 'Reference')]"
                        )
                        print("✅ Found input via partial match!")
                    except:
                        print("❌ Partial match also failed")
                        raise Exception("Could not find Order ID input field")

                order_id_input.clear()
                split_order_id = f"{order_number} - {i}"
                order_id_input.send_keys(split_order_id)

                print(f"✍ Order ID filled: {split_order_id}")

                time.sleep(0.5)

                # 🔴 STOP HERE FOR NOW (as you asked)
                # Later we will:
                # - Fill item count
                # - Fill weight
                # - Submit
                # - Click "Create Another Shipment"

                # -------- SHIPMENT DESCRIPTION --------
                shipment_desc_input = wait.until(
                    EC.presence_of_element_located((
                        By.XPATH,
                        "//input[@placeholder='Enter a description of the item']"
                    ))
                )
                shipment_desc_input.clear()
                shipment_desc_input.send_keys(product_description)
                print(f"✅ Shipment Description filled: {product_description}")

                time.sleep(0.5)

                # -------- ITEM COUNT (set to 1) --------
                item_count_input = wait.until(
                    EC.presence_of_element_located((
                        By.XPATH,
                        "//input[@type='number' and @placeholder='']"
                    ))
                )
                item_count_input.clear()
                item_count_input.send_keys("1")
                print("✅ Item Count set to: 1")

                time.sleep(0.5)

                # -------- PRODUCT CATEGORY --------
                product_category_input = wait.until(
                    EC.presence_of_element_located((
                        By.XPATH,
                        "//input[@placeholder='Select Product Category']"
                    ))
                )
                product_category_input.click()
                product_category_input.send_keys("Agriculture")
                time.sleep(1)

                # Select the "Agriculture" option from dropdown
                try:
                    agri_option = wait.until(
                        EC.element_to_be_clickable((
                            By.XPATH,
                            "//*[contains(text(), 'Agriculture')]"
                        ))
                    )
                    driver.execute_script("arguments[0].click();", agri_option)
                    print("✅ Product Category selected: Agriculture")
                except:
                    print("⚠ Agriculture option not found in dropdown, typed text should remain")

                time.sleep(0.5)

                #shiping value

                shipment_value_input = WebDriverWait(driver, 20).until(
                    EC.presence_of_element_located((
                        By.XPATH,
                        "//input[@placeholder='Enter Item Value']"
                    ))
                )

                shipment_value_input.clear()
                shipment_value_input.send_keys(str(per_split_amount))
                print(f"💵 Shipment Value entered: {per_split_amount}")

                tax_value_input = WebDriverWait(driver, 20).until(
                    EC.presence_of_element_located((
                        By.XPATH,
                        "//input[@placeholder='Enter Tax Value']"
                    ))
                )

                tax_value_input.clear()
                tax_value_input.send_keys("0")
                print("🧾 Tax Value entered: 0")

                total_value_input = WebDriverWait(driver, 20).until(
                    EC.presence_of_element_located((
                        By.XPATH,
                        "//input[@placeholder='Enter Total Value']"
                    ))
                )

                total_value_input.clear()
                total_value_input.send_keys(str(per_split_amount))
                print(f"🧮 Total Value set: {per_split_amount}")




                # -------- PAYMENT MODE (SELECT COD) - WITH RETRY --------
                print("💳 Setting Payment Mode to COD...")
                
                payment_success = False
                for attempt in range(3):
                    try:
                        # Find the Payment Mode dropdown within Payment Details section
                        payment_mode_dropdown = wait.until(
                            EC.element_to_be_clickable((
                                By.XPATH,
                                "//h3[text()='Payment Details']/ancestor::div[contains(@class, 'ucp__non-oms__card-common-outline')]//button[contains(@class, 'ap-menu-trigger-root')]"
                            ))
                        )
                        
                        # Click to open dropdown
                        driver.execute_script("arguments[0].click();", payment_mode_dropdown)
                        # print("📋 Payment Mode dropdown clicked")
                        
                        time.sleep(1)
                        
                        # Select COD option
                        cod_option = WebDriverWait(driver, 5).until(
                            EC.element_to_be_clickable((
                                By.XPATH,
                                "//button[contains(@class, 'ap-menu-item')]//span[text()='COD']"
                            ))
                        )
                        driver.execute_script("arguments[0].click();", cod_option)
                        print("✅ Payment Mode set to: COD")
                        payment_success = True
                        break

                    except Exception as e:
                        print(f"⚠ Attempt {attempt+1} to set COD failed: {e}")
                        time.sleep(2)
                
                if not payment_success:
                    raise Exception("Failed to set Payment Mode to COD after 3 attempts")
                
                time.sleep(1)


                # ================== OPEN FACILITY DROPDOWN ==================

                wait = WebDriverWait(driver, 30)

                # 1️⃣ Locate the Facility dropdown (parent clickable div)
                facility_dropdown = wait.until(
                    EC.presence_of_element_located((
                        By.XPATH,
                        "//span[normalize-space()='Select Facility']/ancestor::div[contains(@class,'ap-menu-trigger')]"
                    ))
                )

                # 2️⃣ Scroll into view
                driver.execute_script(
                    "arguments[0].scrollIntoView({block:'center'});", facility_dropdown
                )
                time.sleep(0.5)

                # 3️⃣ Force click via JS (works even if overlay exists)
                driver.execute_script("arguments[0].click();", facility_dropdown)

                print("🏭 Facility dropdown opened successfully")

                # ================== TYPE PICKUP ADDRESS ==================

                pickup_input = WebDriverWait(driver, 30).until(
                    EC.presence_of_element_located((
                        By.XPATH,
                        "//input[@placeholder='Search Pickup Locations']"
                    ))
                )

                pickup_input.clear()
                pickup_input.send_keys(pickup_text)

                print(f"✍️ Typed pickup address: {pickup_text}")

                time.sleep(1.2)   # allow dropdown results to load

                # ================== SELECT PICKUP FROM LIST ==================

                # ================== SELECT PICKUP FROM LIST (ROBUST) ==================

                pickup_option = None
                pt_lower = pickup_text.lower()
                
                try:
                    # 1. Try strict/specific match first (span with class 'main')
                    pickup_option = WebDriverWait(driver, 5).until(
                        EC.element_to_be_clickable((
                            By.XPATH, 
                            f"//button[contains(@class,'ap-menu-item')]//span[contains(@class,'main') and contains(translate(text(), 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), '{pt_lower}')]"
                        ))
                    )
                    print("✅ Found pickup via strict match")
                except:
                    print("⚠ Strict match failed, trying loose match...")
                    try:
                        # 2. Try loose match (any text in button)
                        pickup_option = WebDriverWait(driver, 5).until(
                            EC.element_to_be_clickable((
                                By.XPATH,
                                f"//button[contains(@class,'ap-menu-item') and contains(translate(., 'ABCDEFGHIJKLMNOPQRSTUVWXYZ', 'abcdefghijklmnopqrstuvwxyz'), '{pt_lower}')]"
                            ))
                        )
                        print("✅ Found pickup via loose match")
                    except:
                        print("⚠ Loose match failed, trying first available option...")
                        try:
                            time.sleep(2) # Give a moment for options to render
                            # 3. Just pick the first option
                            # (Logic: we typed a specific search term, so the first result is likely the correct one)
                            pickup_option = WebDriverWait(driver, 5).until(
                                EC.element_to_be_clickable((By.XPATH, "//button[contains(@class,'ap-menu-item')]"))
                            )
                            print(f"✅ Selected first available option: {pickup_option.text if pickup_option else 'Unknown'}")
                        except:
                            pass
                
                if not pickup_option:
                    # Debugging: print available options
                    print("❌ Could not find pickup option. Available options:")
                    try:
                        options = driver.find_elements(By.XPATH, "//button[contains(@class,'ap-menu-item')]")
                        for opt in options:
                            txt = opt.text or opt.get_attribute("innerText")
                            print(f" - {txt}")
                    except:
                        print(" (Could not list options)")
                    
                    print(f"❌ Error: Failed to select pickup address for '{pickup_text}'")
                    return {"status": "ERROR", "reason": f"Failed to select pickup address for '{pickup_text}'", "order_id": order_number}

                driver.execute_script("arguments[0].click();", pickup_option)

                print(f"✅ Pickup facility selected: {pickup_text}")

                # -------- ADD CUSTOMER DETAILS --------
                add_customer_btn = wait.until(
                    EC.element_to_be_clickable((
                        By.XPATH,
                        "//div[contains(@class, 'text-cta-primary') and contains(text(), 'Add Customer Details')]"
                    ))
                )
                driver.execute_script("arguments[0].click();", add_customer_btn)
                print("🔵 Add Customer Details clicked")

                time.sleep(3)  # Wait for modal to fully load

                # ================== FILL ADD CUSTOMER MODAL ==================
                print("📝 Filling Add Customer modal...")

                try:
                    # -------- SPLIT NAME INTO FIRST AND LAST --------
                    name_parts = customer_name.split(maxsplit=1)
                    first_name = name_parts[0] if len(name_parts) > 0 else customer_name
                    last_name = name_parts[1] if len(name_parts) > 1 else ""

                    # -------- FIRST NAME --------
                    first_name_input = wait.until(
                        EC.presence_of_element_located((
                            By.XPATH,
                            "//input[@name='customer_first_name' or @placeholder='Enter first name']"
                        ))
                    )
                    first_name_input.clear()
                    first_name_input.send_keys(first_name)
                    print(f"✅ First Name filled: {first_name}")

                    time.sleep(0.3)

                    # -------- LAST NAME --------
                    last_name_input = driver.find_element(
                        By.XPATH,
                        "//input[@name='customer_last_name' or @placeholder='Enter last name']"
                    )
                    last_name_input.clear()
                    if last_name:
                        last_name_input.send_keys(last_name)
                        print(f"✅ Last Name filled: {last_name}")
                    else:
                        print("ℹ Last Name left empty (no last name in source)")

                    time.sleep(0.3)

                    # -------- EMAIL (OPTIONAL) --------
                    email_input = driver.find_element(
                        By.XPATH,
                        "//input[@name='email' or @type='email' or @placeholder='Enter email ID']"
                    )
                    email_input.clear()
                    if customer_email:
                        email_input.send_keys(customer_email)
                        print(f"✅ Email filled: {customer_email}")
                    else:
                        print("ℹ Email left empty (optional)")

                    time.sleep(0.3)

                    # -------- PHONE NUMBER --------
                    phone_input = driver.find_element(
                        By.XPATH,
                        "//input[@name='phone_number' or @placeholder='Enter mobile number']"
                    )
                    phone_input.clear()
                    phone_input.send_keys(customer_phone)
                    print(f"✅ Phone Number filled: {customer_phone}")

                    time.sleep(0.3)

                    # -------- SHIPPING ADDRESS LINE 1 --------
                    address_line1_input = driver.find_element(
                        By.XPATH,
                        "//input[@placeholder='Address Line 1']"
                    )
                    address_line1_input.clear()
                    address_line1_input.send_keys(customer_address)
                    print(f"✅ Shipping Address Line 1 filled: {customer_address}")

                    time.sleep(0.3)

                    # -------- SHIPPING ADDRESS LINE 2 (OPTIONAL) --------
                    # Combine City, Taluka, District for Address Line 2

                    wait = WebDriverWait(driver, 20)

                    # ================== BUILD FORMATTED ADDRESS ==================

                    address_parts = []

                    if customer_city:
                        address_parts.append(f"City : {customer_city}")

                    if customer_taluka:
                        address_parts.append(f"Taluka : {customer_taluka}")

                    if customer_district:
                        address_parts.append(f"District : {customer_district}")

                    address_line2 = " ".join(address_parts)

                    # Example output:
                    # City : Ahmedabad Taluka : Ahmedabad District : Ahmedabad

                    # ================== LOCATE INPUT ==================

                    address_line2_input = wait.until(
                        EC.element_to_be_clickable((
                            By.XPATH, "//input[@placeholder='Address Line 2']"
                        ))
                    )

                    # ================== FOCUS + CLEAR ==================

                    address_line2_input.click()
                    driver.execute_script("arguments[0].value = '';", address_line2_input)

                    # ================== PASTE VALUE (REACT SAFE) ==================

                    if address_line2:
                        driver.execute_script(
                            """
                            arguments[0].value = arguments[1];
                            arguments[0].dispatchEvent(new Event('input', { bubbles: true }));
                            arguments[0].dispatchEvent(new Event('change', { bubbles: true }));
                            """,
                            address_line2_input,
                            address_line2
                        )

                        print(f"✅ Address Line 2 filled: {address_line2}")
                    else:
                        print("ℹ Address Line 2 left empty")

                    time.sleep(0.3)


                    # -------- PINCODE --------
                    pincode_input = driver.find_element(
                        By.XPATH,
                        "//input[@placeholder='Enter pincode']"
                    )
                    pincode_input.clear()
                    pincode_input.send_keys(customer_pincode)
                    print(f"✅ Pincode filled: {customer_pincode}")

                    time.sleep(2)  # Wait for State and City to auto-populate from pincode

                    # -------- VERIFY STATE AND CITY AUTO-FILLED --------
                    try:
                        state_input = driver.find_element(
                            By.XPATH,
                            "//input[@placeholder='Enter state']"
                        )
                        state_value = state_input.get_attribute("value")
                        print(f"ℹ State auto-filled: {state_value if state_value else 'Not filled'}")
                    except:
                        print("ℹ State field not found")

                    try:
                        city_input = driver.find_element(
                            By.XPATH,
                            "//input[@placeholder='Enter City']"
                        )
                        city_value = city_input.get_attribute("value")
                        print(f"ℹ City auto-filled: {city_value if city_value else 'Not filled'}")
                    except:
                        print("ℹ City field not found")

                    # -------- CLICK "ADD CUSTOMER" BUTTON --------
                    try:
                        # Try multiple selectors for the Add Customer button
                        add_customer_btn = None
                        
                        # Try method 1: Exact text match with blue button
                        try:
                            add_customer_btn = WebDriverWait(driver, 5).until(
                                EC.element_to_be_clickable((
                                    By.XPATH,
                                    "//button[contains(@class, 'blue') and contains(., 'Add Customer')]"
                                ))
                            )
                            print("✅ Found 'Add Customer' button (method 1)")
                        except:
                            pass

                        # Try method 2: Just text content
                        if not add_customer_btn:
                            try:
                                add_customer_btn = WebDriverWait(driver, 5).until(
                                    EC.element_to_be_clickable((
                                        By.XPATH,
                                        "//button[normalize-space()='Add Customer']"
                                    ))
                                )
                                print("✅ Found 'Add Customer' button (method 2)")
                            except:
                                pass

                        # Try method 3: Any button with "Add Customer" text
                        if not add_customer_btn:
                            try:
                                add_customer_btn = WebDriverWait(driver, 5).until(
                                    EC.element_to_be_clickable((
                                        By.XPATH,
                                        "//button[contains(text(), 'Add Customer')]"
                                    ))
                                )
                                print("✅ Found 'Add Customer' button (method 3)")
                            except:
                                pass

                        if add_customer_btn:
                            driver.execute_script("arguments[0].scrollIntoView({block:'center'});", add_customer_btn)
                            time.sleep(0.5)
                            driver.execute_script("arguments[0].click();", add_customer_btn)
                            print("✅ 'Add Customer' button clicked")
                            time.sleep(3)  # Wait for modal to close and customer to be added
                        else:
                            print("⚠ Could not find 'Add Customer' button - modal may need manual submit")

                    except Exception as btn_error:
                        print(f"⚠ Error clicking Add Customer button: {btn_error}")

                except Exception as e:
                    print(f"⚠ Error filling customer modal: {e}")
                    import traceback
                    traceback.print_exc()

                # ================== SELECT PACKAGE TYPE: CARDBOARD BOX ==================
                
                print("📦 Selecting Package Type...")
                
                try:
                    # Wait a bit for the page to update after customer is added
                    time.sleep(2)

                    # Find and click the Package Type dropdown
                    package_type_dropdown = wait.until(
                        EC.element_to_be_clickable((
                            By.XPATH,
                            "//button[.//span[text()='Select Package Type']]"
                        ))
                    )
                    
                    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", package_type_dropdown)
                    time.sleep(0.5)
                    driver.execute_script("arguments[0].click();", package_type_dropdown)
                    print("✅ Package Type dropdown opened")
                    
                    time.sleep(1)
                    
                    # Select "Cardboard Box" option
                    cardboard_option = wait.until(
                        EC.element_to_be_clickable((
                            By.XPATH,
                            "//button[contains(@class, 'ap-menu-item')]//span[contains(text(), 'Cardboard Box')]"
                        ))
                    )
                    
                    driver.execute_script("arguments[0].click();", cardboard_option)
                    print("✅ Package Type set to: Cardboard Box")
                    
                    time.sleep(1)
                    

                except Exception as pkg_error:
                    print(f"⚠ Error selecting Package Type: {pkg_error}")
                    print("ℹ Package Type may need to be selected manually")

                print("✅ Delhivery form partially filled. Ready for manual box details entry.")

                # ================== FILL BOX DIMENSIONS (L, B, H) ==================

                wait = WebDriverWait(driver, 20)

                def fill_number_input(el, value):
                    driver.execute_script(
                        """
                        arguments[0].focus();
                        arguments[0].value = '';
                        arguments[0].value = arguments[1];
                        arguments[0].dispatchEvent(new Event('input', { bubbles: true }));
                        arguments[0].dispatchEvent(new Event('change', { bubbles: true }));
                        """,
                        el,
                        value
                    )

                # L
                length_input = wait.until(
                    EC.element_to_be_clickable((By.XPATH, "//input[@placeholder='L']"))
                )
                fill_number_input(length_input, length)

                # B
                breadth_input = wait.until(
                    EC.element_to_be_clickable((By.XPATH, "//input[@placeholder='B']"))
                )
                fill_number_input(breadth_input, breadth)

                # H
                height_input = wait.until(
                    EC.element_to_be_clickable((By.XPATH, "//input[@placeholder='H']"))
                )
                fill_number_input(height_input, height)

                print(f"✅ Box dimensions filled → L={length}, B={breadth}, H={height}")



                # ================== FILL PACKAGED WEIGHT ==================

                wait = WebDriverWait(driver, 20)

                # Round weight and ensure integer (Delhivery prefers whole grams)
                packaged_weight = int(round(per_split_weight)) if per_split_weight else int(round(total_weight))


                weight_input = wait.until(
                    EC.element_to_be_clickable((
                        By.XPATH,
                        "//input[@placeholder='Enter packaged weight']"
                    ))
                )

                driver.execute_script(
                    """
                    arguments[0].focus();
                    arguments[0].value = '';
                    arguments[0].value = arguments[1];
                    arguments[0].dispatchEvent(new Event('input', { bubbles: true }));
                    arguments[0].dispatchEvent(new Event('change', { bubbles: true }));
                    """,
                    weight_input,
                    packaged_weight
                )

                print(f"✅ Packaged weight filled: {packaged_weight} g")

                create_forward_btn = WebDriverWait(driver, 30).until(
                    EC.element_to_be_clickable((
                        By.XPATH,
                        "//button[normalize-space()='Create Forward Shipment']"
                    ))
                )

                driver.execute_script("arguments[0].click();", create_forward_btn)
                print("🚀 Create Forward Shipment clicked")

                time.sleep(3)

                # ================== CLICK PRINT SHIPPING LABEL ==================

                print_label_btn = WebDriverWait(driver, 30).until(
                    EC.element_to_be_clickable((
                        By.XPATH,
                        "//button[normalize-space()='Print Shipping Label']"
                    ))
                )

                driver.execute_script("arguments[0].click();", print_label_btn)
                print("✅ Print Shipping Label button clicked")

                # ================== RENAME DOWNLOADED SHIPPING LABEL ==================

                def wait_and_rename_pdf(download_dir, new_name, timeout=60):
                    print("⏳ Waiting for shipping label download...")

                    before_files = set(os.listdir(download_dir))
                    end_time = time.time() + timeout
                    downloaded_pdf = None

                    while time.time() < end_time:
                        current_files = set(os.listdir(download_dir))
                        new_files = current_files - before_files

                        for f in new_files:
                            if f.lower().endswith(".pdf") and not f.lower().endswith(".crdownload"):
                                downloaded_pdf = f
                                break

                        if downloaded_pdf:
                            break

                        time.sleep(0.5)

                    if not downloaded_pdf:
                        raise Exception("❌ Shipping label PDF did not download")

                    old_path = os.path.join(download_dir, downloaded_pdf)
                    safe_name = new_name.replace("/", "_").replace(" ", "_").strip()
                    if not safe_name.lower().endswith(".pdf"):
                        safe_name += ".pdf"
                    
                    new_path = os.path.join(download_dir, safe_name)

                    if os.path.exists(new_path):
                        os.remove(new_path)

                    os.rename(old_path, new_path)

                    print(f"✅ Shipping label renamed: {safe_name}")
                    return new_path

                # Rename using the SPLIT order ID (e.g. 12345 - 1.pdf)
                renamed_pdf_path = wait_and_rename_pdf(download_dir, split_order_id)
                
                # ================== SEND EMAIL ==================
                if seller_email:
                    print(f"📧 Sending email to {seller_email}...")
                    email_subject = f"Shipping Label for Order {split_order_id}"
                    email_body = f"Please find attached the shipping label for order {split_order_id}.\n\nProduct: {product_description}\n\nRegards,\nAutomation Script"
                    
                    email_helper.send_email_with_attachment(
                        SENDER_EMAIL, 
                        SENDER_PASSWORD, 
                        seller_email, 
                        email_subject, 
                        email_body, 
                        renamed_pdf_path
                    )
                else:
                    print("⚠ Skipping email (no seller email found)")

                # ================== PREPARE FOR NEXT LOOP (USER REQUESTED NAVIGATION) ==================
                # "navigate to that page and again click on Create Forward shipment"
                if i < split_count:
                    print("🔄 Preparing for next split shipment (Explicit Navigation)...")
                    
                    try:
                        # 1. Click Sidebar Toggler (Expand Sidebar)
                        print("🖱 Clicking Sidebar Toggler...")
                        try:
                            sidebar_toggler = WebDriverWait(driver, 5).until(
                                EC.element_to_be_clickable((
                                    By.XPATH,
                                    "//div[contains(@class, 'ap-sidebar-parent--left__toggler')]"
                                ))
                            )
                            driver.execute_script("arguments[0].click();", sidebar_toggler)
                            print("✅ Sidebar Toggler clicked")
                            time.sleep(1)
                        except:
                            print("ℹ Sidebar might already be expanded or toggler not found")

                        # 2. Click "Shipments & Pickups" (Dropdown)
                        print("🖱 Clicking 'Shipments & Pickups' menu...")
                        shipments_menu = WebDriverWait(driver, 10).until(
                            EC.element_to_be_clickable((
                                By.XPATH,
                                "//button[contains(@class, 'ap-sidebar__item') and .//span[contains(text(), 'Shipments & Pickups')]]"
                            ))
                        )
                        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", shipments_menu)
                        time.sleep(0.5)
                        driver.execute_script("arguments[0].click();", shipments_menu)
                        print("✅ 'Shipments & Pickups' clicked")
                        
                        time.sleep(2) # Wait for dropdown to expand

                        # 3. Click "Forward Shipments" Link
                        print("🖱 Clicking 'Forward Shipments' link...")
                        forward_shipments_link = WebDriverWait(driver, 10).until(
                            EC.element_to_be_clickable((
                                By.XPATH,
                                "//a[contains(@class, 'ap-sidebar__item') and .//div[contains(text(), 'Forward Shipments')]]"
                            ))
                        )
                        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", forward_shipments_link)
                        time.sleep(0.5)
                        driver.execute_script("arguments[0].click();", forward_shipments_link)
                        print("✅ 'Forward Shipments' clicked")

                        time.sleep(5) # Wait for page reload

                        # 4. Click Create Forward Shipment Button again
                        print("🔍 Looking for 'Create Forward Shipment' button for next loop...")
                        create_forward_btn_loop = WebDriverWait(driver, 10).until(
                            EC.element_to_be_clickable((
                                By.XPATH, 
                                "//button[contains(@class, 'ap-button') and contains(@class, 'blue') and contains(., 'Create Forward Shipment')]"
                            ))
                        )
                        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", create_forward_btn_loop)
                        time.sleep(0.5)
                        driver.execute_script("arguments[0].click();", create_forward_btn_loop)
                        print("🚀 Create Forward Shipment button clicked (for next part)")
                        
                        time.sleep(3) # Wait for form to open

                    except Exception as e:
                        print(f"❌ Failed to reset form via sidebar navigation: {e}")
                        raise e
                    
                # ✅ Split shipment completed — stop further AWB processing
            return {"status": "SUCCESS", "order_id": order_number, "reason": "Split shipment created"}


        # ================== CLICK AWB DROPDOWN ==================

        awb_dropdown = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((
                By.XPATH,
                "//button[.//span[text()='AWB']]"
            ))
        )

        driver.execute_script("arguments[0].click();", awb_dropdown)
        print("✅ AWB dropdown opened")

        order_id_option = WebDriverWait(driver, 20).until(
        EC.element_to_be_clickable((
            By.XPATH,
            "//button[.//span[text()='Order ID']]"
        ))
        )

        driver.execute_script("arguments[0].click();", order_id_option)
        print("✅ Order ID selected")

        order_search_input = WebDriverWait(driver, 20).until(
        EC.presence_of_element_located((
            By.XPATH,
            "//input[contains(@placeholder,'ORDER ID')]"
        ))
    )

        order_search_input.clear()
        order_search_input.send_keys(order_number)

        print(f"✅ Order ID pasted: {order_number}")

        # ================== CLICK ORDER SEARCH RESULT (WITH RETRY) ==================
        
        time.sleep(3) # Wait for search results to appear

        search_success = False
        for attempt in range(3):
            try:
                print(f"🔍 Waiting for search result (Attempt {attempt+1})...")
                search_result = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((
                        By.XPATH,
                        "//div[contains(@class,'ucp__global-search__results')]"
                        "//div[contains(@class,'cursor-pointer')]"
                        f"[.//span[contains(text(), '{order_number}')]]"
                    ))
                )
                driver.execute_script("arguments[0].click();", search_result)
                print("✅ Order search result clicked")
                search_success = True
                break
            except Exception as e:
                print(f"⚠ Attempt {attempt+1} search failed: {e}")
                if attempt < 2:
                    # Retry entering text
                    order_search_input.clear()
                    time.sleep(0.5)
                    order_search_input.send_keys(order_number)
                    time.sleep(2)
        
        if not search_success:
            print(f"❌ Error: Could not find order {order_number} in Delhivery search results after 3 attempts.")
            return {"status": "ERROR", "reason": "Order not found in Delhivery", "order_id": order_number}

        time.sleep(2)




        # ================== PARTIAL PAYMENT HANDLING IN DELHIVERY ==================

        if is_partial_payment:
            print("⚠ Partial payment flow activated in Delhivery")

            # ---------- SWITCH TO AGRIBEGRI TAB ----------
            driver.switch_to.window(agribegri_order_tab)
            driver.refresh()
            time.sleep(3)

            print("✅ Forced return to Agribegri edit order page")


            # ---------- COPY COD PAYMENT AMOUNT ----------
            cod_amount_elem = WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((
                    By.XPATH,
                    "//p[contains(@class,'p_disp_advance_amt') and contains(.,'COD Payment Amount')]"
                ))
            )

            import re
            cod_text = cod_amount_elem.text
            cod_amount = re.search(r"([\d]+\.\d+)", cod_text).group(1)

            print("💰 Correct COD Amount copied:", cod_amount)

            # ---------- SWITCH BACK TO DELHIVERY ----------
            driver.switch_to.window(driver.window_handles[-1])

            # ---------- CLICK ✏️ EDIT ICON (CORRECT ELEMENT) ----------
            edit_payment_btn = WebDriverWait(driver, 20).until(
                EC.element_to_be_clickable((
                    By.XPATH,
                    "//button[@data-action='edit-payment-mode']"
                ))
            )

            driver.execute_script("arguments[0].click();", edit_payment_btn)
            print("💜 Payment Details edit icon clicked (correct one)")

            time.sleep(0.8)

            # ---------- FORCE-OPEN PAYMENT MODE DROPDOWN ----------
            payment_dropdown = WebDriverWait(driver, 20).until(
                EC.element_to_be_clickable((
                    By.XPATH,
                    "//button[contains(@class,'ap-menu-trigger-root')]"
                ))
            )

            driver.execute_script("arguments[0].click();", payment_dropdown)
            time.sleep(0.5)

            cod_option = WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((
                    By.XPATH,
                    "//div[contains(@class,'ucp__order-creation__select-payment__dropdown-item--cod')]"
                ))
            )

            driver.execute_script("arguments[0].click();", cod_option)
            print("💵 Cash On Delivery selected")
            time.sleep(0.5)


            # ---------- ENTER COLLECTABLE AMOUNT ----------
            collectable_input = WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "input.input[type='number']"))
            )
            collectable_input.clear()
            collectable_input.send_keys(cod_amount)
            print(f"💰 Collectable amount entered: {cod_amount}")

            # ---------- CLICK ✔ SAVE ----------
            save_btn = WebDriverWait(driver, 20).until(
                EC.element_to_be_clickable((By.XPATH, "//button[@title='Save']"))
            )
            driver.execute_script("arguments[0].click();", save_btn)
            print("✅ COD payment saved")

            time.sleep(2)


        # ================== CLICK PRINT SHIPPING LABEL ==================

        print_label_btn = WebDriverWait(driver, 30).until(
            EC.element_to_be_clickable((
                By.XPATH,
                "//button[normalize-space()='Print Shipping Label']"
            ))
        )

        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", print_label_btn)
        time.sleep(0.5)

        driver.execute_script("arguments[0].click();", print_label_btn)
        print("✅ Print Shipping Label button clicked")

        time.sleep(2)

        # ================== WAIT FOR PDF DOWNLOAD & RENAME ==================

        # def wait_and_rename_pdf(download_dir, order_id, timeout=30):
        #     start_time = time.time()
        #     downloaded_file = None

        #     while time.time() - start_time < timeout:
        #         files = os.listdir(download_dir)

        #         # find completed PDF (ignore .crdownload)
        #         pdfs = [
        #             f for f in files
        #             if f.lower().endswith(".pdf") and not f.lower().endswith(".crdownload")
        #         ]

        #         if pdfs:
        #             # assume latest downloaded file
        #             pdfs.sort(
        #                 key=lambda x: os.path.getmtime(os.path.join(download_dir, x)),
        #                 reverse=True
        #             )
        #             downloaded_file = pdfs[0]
        #             break

        #         time.sleep(1)

        #     if not downloaded_file:
        #         raise Exception("❌ PDF download timed out")

        #     old_path = os.path.join(download_dir, downloaded_file)
        #     new_filename = f"{order_id}.pdf"
        #     new_path = os.path.join(download_dir, new_filename)

        #     # if file with same name exists, remove it
        #     if os.path.exists(new_path):
        #         os.remove(new_path)

        #     os.rename(old_path, new_path)
        #     print(f"✅ PDF renamed to: {new_filename}")

        def wait_and_rename_pdf(download_dir, order_id, timeout=30):
            start_time = time.time()
            downloaded_file = None

            while time.time() - start_time < timeout:
                files = os.listdir(download_dir)

                pdfs = [
                    f for f in files
                    if f.lower().endswith(".pdf") and not f.lower().endswith(".crdownload")
                ]

                if pdfs:
                    pdfs.sort(
                        key=lambda x: os.path.getmtime(os.path.join(download_dir, x)),
                        reverse=True
                    )
                    downloaded_file = pdfs[0]
                    break

                time.sleep(1)

            if not downloaded_file:
                raise Exception("❌ PDF download timed out")

            old_path = os.path.join(download_dir, downloaded_file)
            new_path = os.path.join(download_dir, f"{order_id}.pdf")

            if os.path.exists(new_path):
                os.remove(new_path)

            os.rename(old_path, new_path)
            print(f"✅ PDF renamed to: {order_id}.pdf")


            return new_path


        # call the function
        wait_and_rename_pdf(download_dir, order_number)

        # ================== SWITCH BACK TO AGRIBEGRI ADMIN TAB ==================
        driver.switch_to.window(agribegri_order_tab)
        driver.refresh()
        time.sleep(2)

        print("✅ Switched back to Agribegri edit order page (final)")

        # ================== SET ORDER STATUS TO PACKED ==================

        # wait for Order Status dropdown
        order_status_select = WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.ID, "abo_status"))
        )

        # use Select for <select> tag
        select = Select(order_status_select)

        # select "Packed"
        select.select_by_visible_text("Packed")

        print("✅ Order status set to Packed")

        time.sleep(1)

        # ================== SET PACKED REASON TO CL SURFACE ==================

        packed_reason_select = WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.ID, "abo_packed_reason"))
        )

        packed_reason = Select(packed_reason_select)

        packed_reason.select_by_visible_text("CL Surface")

        print("✅ Packed reason set to CL Surface")

        time.sleep(1)


        # ================== SET PAYMENT STATUS IF PARTIAL PAYMENT ==================

        if is_partial_payment:
            print("⚠ Setting payment status: Bank Partial Payment Recieved")

            payment_status_select = WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((By.ID, "abo_payment_type"))
            )

            payment_select = Select(payment_status_select)

            payment_select.select_by_visible_text("Bank Partial Payment Recieved")

            print("✅ Payment status set to Bank Partial Payment Recieved")

            time.sleep(1)
        else:
            print("💳 Skipping payment status change (not partial payment)")

        time.sleep(1)
        # ================== CLICK SUBMIT BUTTON ==================

        submit_btn = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((
                By.NAME, "update_order_status"
            ))
        )

        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", submit_btn)
        time.sleep(0.5)

        driver.execute_script("arguments[0].click();", submit_btn)
        print("✅ Submit button clicked successfully")

        time.sleep(2)
        # ================== CLICK OK ON SUCCESS POPUP ==================

        ok_btn = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.ID, "popup_ok"))
        )

        driver.execute_script("arguments[0].click();", ok_btn)
        print("✅ Success popup OK clicked")

        time.sleep(1.5)

        # ================== UPLOAD LABEL PDF ==================

        label_input = WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.NAME, "label_file"))
        )

        label_pdf_path = os.path.join(download_dir, f"{order_number}.pdf")

        if not os.path.exists(label_pdf_path):
            raise Exception(f"❌ Label PDF not found: {label_pdf_path}")

        label_input.send_keys(label_pdf_path)
        print(f"✅ Label PDF uploaded: {label_pdf_path}")

        time.sleep(1)
            # ================== CLICK SUBMIT BUTTON ==================

        submit_btn = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.NAME, "update_manifest_file"))
        )

        submit_btn.click()
        print("✅ Submit button clicked")


        # ================== SUBMIT LABEL UPLOAD ==================

        time.sleep(1)

        # ================== CLICK OK ON POPUP ==================

        ok_btn = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.ID, "popup_ok"))
        )

        ok_btn.click()
        print("✅ OK button clicked on popup")

        time.sleep(1)
        return {"status": "SUCCESS", "order_id": order_number, "reason": "Processed Successfully"}

    
    except Exception as e:
        print(f"❌ Error inside click_truck_icons_one_by_one: {e}")
        import traceback
        traceback.print_exc()
        return {
            "status": "ERROR",
            "order_id": order_number if 'order_number' in locals() else "Unknown",
            "reason": str(e)
        }



def main_workflow():
    login_to_agribegri()
    apply_filter()

    print("Filtering done. Checking total orders...")

    # ================== GET TOTAL COUNT ==================
    try:
        total_count_elem = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.XPATH, "//*[@id='seller_frm']/div/span[1]"))
        )
        total_text = total_count_elem.text.strip()
        # Extract number from string like "Total No of Orders : 75 "
        import re
        match = re.search(r"(\d+)", total_text)
        if match:
            total_orders = int(match.group(1))
        else:
            total_orders = 50 # Fallback default
            print("⚠ Could not parse total count, defaulting to 50")
            
        print(f"📊 Total Orders Found: {total_orders}")

    except Exception as e:
        print(f"⚠ Error reading total count: {e}. Defaulting to single page processing.")
        total_orders = 50

    
    import math
    ORDERS_PER_PAGE = 50
    total_pages = math.ceil(total_orders / ORDERS_PER_PAGE)
    print(f"📄 Total Pages to process: {total_pages}")


    processed_count = 0
    total_processed_global = 0

    # ================== PAGE LOOP ==================
    for page in range(1, total_pages + 1):
        print(f"\n📂 --- STARTING PAGE {page} / {total_pages} ---")
        
        # Determine rows on this page (might be less than 50 on last page)
        # But actually we can just loop until "END" (index out of bounds) is returned
        
        current_index = 0

        while True:
            # Check Global Limit (if set)
            if ROW_LIMIT > 0 and total_processed_global >= ROW_LIMIT:
                print(f"🛑 Global Limit of {ROW_LIMIT} rows reached.")
                return

            print(f"\n🔄 --- Processing Row Index: {current_index} (Page {page}) ---")
            
            try:
                # Call the function for the current index
                result = click_truck_icons_one_by_one(current_index)

                # Check for End of List on this page
                if isinstance(result, dict) and result.get("status") == "END":
                    print(f"🏁 End of rows on Page {page}.")
                    break
                
                # Retrieve Order ID safely
                order_id_log = result.get("order_id", "Unknown") if isinstance(result, dict) else "N/A"

                # Check Success
                if isinstance(result, dict) and result.get("status") == "SUCCESS":
                    print(f"✅ Row {current_index} processed successfully.")
                    log_to_excel(current_index, order_id_log, "SUCCESS", "Processed")
                    processed_count += 1
                    total_processed_global += 1
                
                # Check Skipped
                elif isinstance(result, dict) and result.get("status") == "SKIPPED":
                    reason = result.get("reason", "Skipped")
                    print(f"⏩ Row {current_index} skipped: {reason}")
                    log_to_excel(current_index, "N/A", "SKIPPED", reason)
                
                elif isinstance(result, dict) and result.get("status") == "ERROR":
                    reason = result.get("reason", "Unknown Error")
                    order_id = result.get("order_id", "N/A")
                    print(f"❌ Row {current_index} failed: {reason} (Order: {order_id})")
                    log_to_excel(current_index, order_id, "ERROR", reason)
                
                else:
                    # Fallback for unexpected return
                    print(f"⏩ Row {current_index} skipped/unknown status.")
                    log_to_excel(current_index, "N/A", "UNKNOWN", "Unexpected return value")
                    
                current_index += 1


            except Exception as e:
                print(f"❌ Error processing row {current_index}: {e}")
                import traceback
                traceback.print_exc()
                log_to_excel(current_index, "Error", "ERROR", str(e))
                current_index += 1 # Skip on error
                
            # ================== CLEANUP TABS ==================
            try:
                while len(driver.window_handles) > 1:
                    driver.switch_to.window(driver.window_handles[-1])
                    print("🧹 Closing extra tab...")
                    driver.close()
                
                if len(driver.window_handles) > 0:
                    driver.switch_to.window(driver.window_handles[0])
                    # print("🔙 Switched back to Main Window")
            except Exception as cleanup_err:
                print(f"⚠ Cleanup warning: {cleanup_err}")

        # ================== NEXT PAGE LOGIC ==================
        if page < total_pages:
            print(f"➡ Page {page} done. Moving to Page {page + 1}...")
            
            try:
                # Find NEXT button
                # Look for <a> with class 'cus_page_act' and text 'Next'
                next_btn = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((
                        By.XPATH, 
                        "//a[contains(@class, 'cus_page_act') and contains(text(), 'Next')]"
                    ))
                )
                
                driver.execute_script("arguments[0].scrollIntoView({block:'center'});", next_btn)
                time.sleep(0.5)
                driver.execute_script("arguments[0].click();", next_btn)
                print("✅ Next button clicked")
                
                # Wait for table to reload (simple pause usually enough, or wait for loader)
                print("⏳ Waiting for next page to load...")
                time.sleep(5) 
                
            except Exception as e:
                print(f"❌ Could not click Next button: {e}")
                break
        else:
             print("🏁 All pages processed.")
    
    print(f"\n🎉 Workflow completed. Total rows processed: {total_processed_global}")


if __name__ == "__main__":
    try:
        main_workflow()
    except Exception as e:
        print(f"\n❌ CRITICAL ERROR: {e}")
        import traceback
        traceback.print_exc()
    finally:
        print("\n🏁 Script finished. Closing browser in 5 seconds...")
        time.sleep(5)
        # driver.quit() # Optional: keep open for inspection if needed, or close.
        # But for automation, closing is better.
        # driver.quit()
