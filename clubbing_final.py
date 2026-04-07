import traceback
# Removed invalid import: from posthog import page
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import Select
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
import time
import os
import openpyxl
import smtplib
import os
from email.message import EmailMessage
from datetime import datetime, timedelta
import re
from collections import Counter
from selenium.common.exceptions import TimeoutException
import os
import time
import logging

# ================= LOGGING SETUP =================

log_dir = "logs"
if not os.path.exists(log_dir):
    os.makedirs(log_dir)

logging.basicConfig(
    filename=os.path.join(log_dir, "automation_clubbing.log"),   # log file name
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(message)s",
    filemode="a",   # append mode
    force=True      # Reset existing logging
)

logger = logging.getLogger()
download_dir = r"C:\Users\ghara\Documents\agribegri_UI\clubbing"

from openpyxl import Workbook

log_file = "order_processing_log.xlsx"
CLUBBING_HISTORY_FILE = "clubbing_processed_orders.json"
import json

def update_clubbing_history(entry):
    history = []
    if os.path.exists(CLUBBING_HISTORY_FILE):
        with open(CLUBBING_HISTORY_FILE, "r") as f:
            try: history = json.load(f)
            except: pass
    entry["timestamp"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    history.append(entry)
    with open(CLUBBING_HISTORY_FILE, "w") as f:
        json.dump(history[-1000:], f, indent=4)

def initialize_log_file():
    if not os.path.exists(log_file):
        wb = Workbook()
        ws = wb.active
        ws.title = "Order Log"
        ws.append(["Timestamp", "Order Number", "Status", "Message"])
        wb.save(log_file)
        print(" Log file created")

def log_order_status(order_no, status, reason=""):
    try:
        if not os.path.exists(log_file):
            initialize_log_file()
        
        wb = openpyxl.load_workbook(log_file)
        ws = wb.active
        ws.append([datetime.now().strftime("%Y-%m-%d %H:%M:%S"), order_no, status, reason])
        wb.save(log_file)
        logger.info(f" Logged: {order_no} → {status}")
        
        # Add to JSON history for dashboard
        update_clubbing_history({
            "order_no": str(order_no),
            "status": status,
            "message": reason
        })
    except PermissionError:
        logger.error(f" CRITICAL: Could not save to {log_file}. Please CLOSE the Excel file!")
        print(f"\n !!! CLOSE EXCEL FILE: {log_file} !!!")
    except Exception as e:
        logger.error(f" Error logging status: {e}")

def log_manual_required(order_nos, reason):
    if isinstance(order_nos, str):
        order_nos = [order_nos]
    for order_no in order_nos:
        log_order_status(order_no, "MANUALLY REQUIRED", reason)


# driver will be initialized in the main loop below

# ================= TAB TRACKING =================
main_tab = None
buyer_status_tab = None
delhivery_tab = None


def load_seller_pickup_mapping():
    file_path = r"Pickup Address & Seller Name .xlsx"

    wb = openpyxl.load_workbook(file_path)
    sheet = wb.active

    mapping = {}

    for row in sheet.iter_rows(min_row=2, values_only=True):
        seller_name = str(row[0]).strip()
        pickup_name = str(row[1]).strip()

        mapping[seller_name] = pickup_name

    print(" Seller → Pickup mapping loaded")
    return mapping
def split_customer_name(full_name):
    name_parts = full_name.strip().split()

    if len(name_parts) >= 2:
        first_name = name_parts[0]
        last_name = name_parts[1]
    elif len(name_parts) == 1:
        first_name = name_parts[0]
        last_name = name_parts[0]   # fallback
    else:
        first_name = ""
        last_name = ""

    return first_name, last_name

def split_dimension(dimension_string):
    # Example: "20.00 X 15.00 X 15.00"
    parts = dimension_string.replace(" ", "").split("X")

    length = int(float(parts[0]))
    breadth = int(float(parts[1]))
    height = int(float(parts[2]))

    return length, breadth, height

# ================= LOGIN FUNCTION =================


def step_wait(seconds=5, msg=""):
    if msg:
        print(f" Waiting {seconds}s → {msg}")
    time.sleep(seconds)

import sys

def login_to_agribegri(username='Clubbed', password='Clubbed@022026', otp='123456'):
    driver.get('https://agribegri.com/admin/')
    driver.maximize_window()

    username_field = driver.find_element(By.ID, "username")
    username_field.clear()
    username_field.send_keys(username)
    next_button = driver.find_element(By.ID, "btnSubmit")
    next_button.click()
    time.sleep(3)

    password_field = driver.find_element(By.ID, "password")
    password_field.clear()
    password_field.send_keys(password)
    pass_next_button = driver.find_element(By.ID, "btnSubmit")
    pass_next_button.click()
    time.sleep(3)

    otp_field = driver.find_element(By.ID, "otp")
    otp_field.clear()
    otp_field.send_keys(otp)
    sign_in = driver.find_element(By.ID, "btnSubmit")
    sign_in.click()
    time.sleep(2)

# ================= CALL =================


def click_manage_orders_nav():
    manage_orders = wait.until(
        EC.element_to_be_clickable((
            By.XPATH,
            "//a[@href='manage_orders.php' and contains(., 'Manage Orders')]"
        ))
    )

    driver.execute_script(
        "arguments[0].scrollIntoView({block:'center'});", manage_orders
    )
    manage_orders.click()

    print(" Clicked Manage Orders from navigation")

# def click_to_date_today():
#     # Click To Date input
#     to_date = wait.until(
#         EC.element_to_be_clickable((By.ID, "to_date"))
#     )
#     to_date.click()

#     print(" To Date calendar opened")

#     # Click today's date in datepicker
#     today_date = wait.until(
#         EC.element_to_be_clickable((
#             By.XPATH,
#             "//td[contains(@class,'ui-datepicker-today')]/a"
#         ))
#     )
#     today_date.click()

#     print(" Today's date selected in To Date")

def click_to_date_today():
    today = datetime.now().strftime("%d-%m-%Y")

    to_date = wait.until(
        EC.presence_of_element_located((By.ID, "to_date"))
    )

    # Remove readonly and set value
    driver.execute_script("""
        arguments[0].removeAttribute('readonly');
        arguments[0].value = arguments[1];
        arguments[0].dispatchEvent(new Event('change'));
    """, to_date, today)

    print(f" To Date set to today via JS: {today}")

def click_from_date_30_days_ago():
    target_date = (datetime.now() - timedelta(days=30)).strftime("%d-%m-%Y")

    from_date = wait.until(
        EC.presence_of_element_located((By.ID, "from_date"))
    )

    # Remove readonly and set value directly using JS
    driver.execute_script("""
        arguments[0].removeAttribute('readonly');
        arguments[0].value = arguments[1];
        arguments[0].dispatchEvent(new Event('change'));
    """, from_date, target_date)

    logger.info(f" From Date set via JS: {target_date}")

    # phone_input = wait.until(
    # EC.presence_of_element_located((By.ID, 'srchby_phone'))
    # )
    # phone_input.clear()
    # phone_input.send_keys('7227029400')


def select_order_status_confirm():
    # 1️ Click the multiselect dropdown button
    status_dropdown_btn = wait.until(
        EC.element_to_be_clickable((
            By.XPATH,
            "//button[contains(@class,'multiselect') and .//span[contains(text(),'Select Order Status')]]"
        ))
    )
    status_dropdown_btn.click()

    logger.info("Order Status dropdown opened")

    # 2️ Click the Confirm checkbox
    confirm_checkbox = wait.until(
        EC.element_to_be_clickable((
            By.XPATH,
            "//label[normalize-space()='Confirm']/input"
        ))
    )

    # Click via JS (more reliable for bootstrap)
    driver.execute_script("arguments[0].click();", confirm_checkbox)

    logger.info("'Confirm' status selected")

    
def select_shipping_through_courier():
    # 1️ Open Shipping Through dropdown
    shipping_dropdown_btn = wait.until(
        EC.element_to_be_clickable((
            By.XPATH,
            "//button[contains(@class,'multiselect') and .//span[contains(text(),'Select Shipping Through')]]"
        ))
    )
    shipping_dropdown_btn.click()

    logger.info("Shipping Through dropdown opened")

    # 2️ Click "Shipping Through Courier" checkbox
    courier_checkbox = wait.until(
        EC.element_to_be_clickable((
            By.XPATH,
            "//label[normalize-space()='Shipping Through Courier']/input"
        ))
    )

    driver.execute_script("arguments[0].click();", courier_checkbox)

    logger.info("'Shipping Through Courier' selected")



def click_search_and_wait():
    # Click Search button
    search_btn = wait.until(
        EC.element_to_be_clickable((By.ID, "srchSubmit"))
    )
    search_btn.click()

    logger.info("Search button clicked")

    # Wait for results table to load / refresh
    wait.until(
        EC.presence_of_element_located((By.CSS_SELECTOR, "table tbody tr"))
    )

    logger.info("Search results loaded")

def type_search_number(driver):
    number = "852592GP"
    
    search_box = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.XPATH, "//input[@aria-controls='dyntable']"))
    )
    
    search_box.clear()
    search_box.send_keys(number)

def select_order_status_packed():
    # 1 Open dropdown
    status_dropdown_btn = wait.until(
        EC.element_to_be_clickable((
            By.XPATH,
            "//button[contains(@class,'multiselect') and .//span[contains(text(),'Select Order Status')]]"
        ))
    )
    status_dropdown_btn.click()

    logger.info("Order Status dropdown opened")

    # 2 Click Packed checkbox
    packed_checkbox = wait.until(
        EC.element_to_be_clickable((
            By.XPATH,
            "//label[normalize-space()='Packed']/input"
        ))
    )

    driver.execute_script("arguments[0].click();", packed_checkbox)

    logger.info("Packed status selected")


def get_total_orders():
    # Ensure table is already loaded
    wait.until(
        EC.presence_of_element_located((By.CSS_SELECTOR, "table tbody tr"))
    )

    try:
        total_span = wait.until(
            EC.visibility_of_element_located((
                By.XPATH,
                "//span[.//b[contains(text(),'Total No of Orders')]]"
            ))
        )

        total_text = total_span.text
        total_orders = int(re.search(r"\d+", total_text).group())

        logger.info(f"Total orders found: {total_orders}")
        return total_orders

    except Exception as e:
        logger.warning("Could not read total orders, falling back to pagination only")
        return None


from collections import defaultdict

def get_duplicate_orders_from_modal():

    wait.until(
        EC.presence_of_element_located((By.CSS_SELECTOR, "#pending_orders_data"))
    )

    time.sleep(1)

    rows = driver.find_elements(By.CSS_SELECTOR, "#pending_orders_data tr")

    row_data = []

    for i in range(len(rows)):
        try:
            rows = driver.find_elements(By.CSS_SELECTOR, "#pending_orders_data tr")
            r = rows[i]

            tds = r.find_elements(By.TAG_NAME, "td")

            if len(tds) < 9:
                continue

            seller = tds[6].text.replace("\xa0", " ").strip()
            ord_by_raw = tds[7].text.strip()
            ord_by = ord_by_raw.split("(")[0].strip()
            order_no = tds[3].text.strip().split("\n")[0]
            buyer_link = tds[8].find_element(By.TAG_NAME, "a")

            row_data.append((seller, ord_by, order_no, buyer_link))

        except Exception as e:
            logger.warning(f"Error processing row {i+1}: {e}")
            continue

    #  GROUP BY cleaned (seller, ord_by)
    grouped = defaultdict(list)
    
    for seller, ord_by, order_no, buyer_link in row_data:
        # Clean for robust matching
        clean_seller = re.sub(r"[^\w\s]", "", seller).lower().strip()
        clean_ord_by = re.sub(r"[^\w\s]", "", ord_by).lower().strip()
        
        grouped[(clean_seller, clean_ord_by)].append((order_no, buyer_link))

    #  ONLY GROUPS WITH >=2
    valid_groups = {
        key: value
        for key, value in grouped.items()
        if len(value) >= 2
    }

    if not valid_groups:
        return False, {}

    # Pick first valid group
    selected_group = list(valid_groups.values())[0]

    order_dict = {
        order_no: buyer_link
        for order_no, buyer_link in selected_group
    }
    import json

    logger.info("Duplicate Order Dictionary:\n%s",
                json.dumps({k: v.get_attribute("href") for k, v in order_dict.items()}, indent=4))
    return True, order_dict


def process_all_orders_with_custom_pagination(total_pages, browser_start_time):
    global seller_pickup_map, buyer_status_tab
    completed_orders_count = 0  # Counter for completed orders

    page = 1

    for page in range(1, total_pages + 1):
        logger.info(f"\nProcessing page {page}")

        rows = wait.until(
            EC.presence_of_all_elements_located(
                (By.CSS_SELECTOR, "table tbody tr")
            )
        )

        logger.info(f"Found {len(rows)} orders on this page")

        i = 0

        while True:
            rows = driver.find_elements(By.CSS_SELECTOR, "table tbody tr")

            if i >= len(rows):
                break

            order_action_dict = None

            row = rows[i]
            i += 1


            try:
                confirm_links = row.find_elements(
                    By.XPATH,
                    ".//td//a[contains(@onclick,'getConfirmOrders')]"
                )

                if not confirm_links:
                    logger.warning(f"No confirm-orders link for row {i+1}, skipping")
                    continue

                #  Read confirm orders count
                count_text = confirm_links[0].text.strip()

                if not count_text.isdigit():
                    logger.warning(f"Invalid confirm count '{count_text}', skipping row {i+1}")
                    continue

                confirm_count = int(count_text)

                #  Skip if only 1 confirm order
                if confirm_count == 1:
                    logger.warning(f"Confirm Orders = 1 for row {i+1}, skipping")
                    continue

                # 🔹 COPY ORDER NUMBER BEFORE EXPAND
                tds = row.find_elements(By.TAG_NAME, "td")

                if len(tds) > 4:

                    main_order_number = tds[4].text.strip().split("\n")[0]

                    action_td = tds[-1]

                    remark_links = action_td.find_elements(
                        By.XPATH,
                        ".//a[contains(@class,'get_id')]"
                    )

                    if not remark_links:
                        logger.warning(f"No remark link found for order {main_order_number}")
                        continue

                    remark_link_element = remark_links[0]
                    remark_id = remark_link_element.get_attribute("id")

                    remark_id = remark_link_element.get_attribute("id")

                    order_action_dict = {
                        "order_no": main_order_number,
                        "remark_id": remark_id,
                        "page": page
                    }



                else:
                    main_order_number = f"ROW_{i}"



                # Expand confirm modal
                driver.execute_script("arguments[0].click();", confirm_links[0])
                logger.info(f"Expanded order {main_order_number} (Confirm Orders = {confirm_count}) on page {page}")

                logger.info(f"Expanded order {i+1} (Confirm Orders = {confirm_count}) on page {page}")
                # wait until modal table rows appear
                wait.until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, "#pending_orders_data tr"))
                )

                is_valid, order_dict = get_duplicate_orders_from_modal()

                if not is_valid:
                    for order_no, buyer_link in order_dict.items():

                        seller_name = buyer_link.get_attribute("data-seller") if buyer_link else "Unknown Seller"
                        ord_by_name = buyer_link.get_attribute("data-ordby") if buyer_link else "Unknown ORD.BY"

                        reason = f"Seller/OrdBy mismatch | Seller: {seller_name} | ORD.BY: {ord_by_name}"

                        log_order_status(order_no, "SKIPPED", reason)

                    logger.warning("Seller / Ord.By mismatch → skipping this order")
                    driver.execute_script("arguments[0].click();", confirm_links[0])
                    time.sleep(1)
                    continue


                logger.info("Duplicate Seller + Ord.By found.")

                order_nos = list(order_dict.keys())

                # SAFETY CHECK
                if len(order_nos) < 2:
                    logger.warning("Less than 2 matching orders after filtering — skipping")
                    driver.execute_script("arguments[0].click();", confirm_links[0])
                    time.sleep(1)
                    continue



                # 2️ Buyer Status link (same for all rows)
                buyer_href = list(order_dict.values())[0].get_attribute("href")

                # 3️ Open Buyer Status ONCE
                logger.info("Opening Buyer Status page")
                driver.execute_script("window.open(arguments[0], '_blank');", buyer_href)
                driver.switch_to.window(driver.window_handles[-1])

                buyer_status_tab = driver.current_window_handle

                wait.until(EC.url_contains("manage_orders_user.php"))
                logger.info("Buyer Status page opened")

                # 4️ Loop order numbers ONE BY ONE
                all_orders_data = {}

                for order_no in order_nos:
                    logger.info(f"Searching order: {order_no}")

                    # search_input = wait.until(
                    #     EC.element_to_be_clickable((By.XPATH, "//input[@aria-controls='dyntable']"))
                    # )
                    # search_input.clear()
                    # search_input.send_keys(order_no)

                    # Locate search box AFTER table loads
                    search_input = wait.until(
                        EC.presence_of_element_located((By.XPATH, "//input[@aria-controls='dyntable']"))
                    )

                    # Re-fetch element to avoid stale reference
                    search_input = driver.find_element(By.XPATH, "//input[@aria-controls='dyntable']")

                    driver.execute_script("arguments[0].value = '';", search_input)

                    search_input.send_keys(order_no)

                    driver.execute_script("""
                        arguments[0].dispatchEvent(new Event('input'));
                    """, search_input)

                    #  trigger JS input event
                    driver.execute_script("""
                        arguments[0].dispatchEvent(new Event('input'));
                    """, search_input)

                    wait.until(
                        EC.presence_of_element_located((By.XPATH, "//table[@id='dyntable']/tbody/tr"))
                    )

                    table_data = collect_manage_user_orders_table()

                    valid_table_data = []

                    for order in table_data:

                        logger.info("\n---------------------------------")
                        logger.info(" Table Row Found")
                        logger.info(f"Order No: {order['order_no']}")
                        logger.info(f"Seller: {order['seller_name']}")
                        logger.info(f"ORD.BY: {order['ord_by']}")
                        logger.info("---------------------------------")

                        view_link = order["actions"]["view"]

                        if not view_link:
                            continue

                        view_details, is_valid = fetch_view_order_details(view_link)

                        if not is_valid:
                            reason = "Individual order exceeds weight (>13000g) or price (>8000) limit"
                            log_order_status(order['order_no'], "SKIPPED", reason)
                            logger.warning(f"Skipping {order['order_no']} - {reason}")
                            continue

                        order["view_details"] = view_details
                        valid_table_data.append(order)

                    # Only store VALID orders
                    if valid_table_data:
                        all_orders_data[order_no] = valid_table_data
                        logger.info(f" Valid Data collected for {order_no}")
                    else:
                        logger.warning(f" No valid data for {order_no}")


                logger.info(" ALL ORDERS PROCESSED")

                import json

                print("\n=========== ALL ORDERS DATA ===========")
                print(json.dumps(all_orders_data, indent=4, default=str))
                print("=======================================\n")
                # Remove empty orders
                all_orders_data = {
                    k: v for k, v in all_orders_data.items() if v
                }

                # CRITICAL: Only use the order numbers that actually passed validation
                order_nos = list(all_orders_data.keys())

                if len(all_orders_data) < 2:
                        logger.warning(" Less than 2 valid orders after filtering → Skipping clubbing")

                        # Close Buyer Status tab
                        if buyer_status_tab in driver.window_handles:
                            driver.close()

                        # Switch back to main tab
                        driver.switch_to.window(main_tab)

                        i += 1   #  Move to next row safely
                        continue

                # Calculate totals for the manifest and collection
                combined_total_valuation = 0  # True worth of all goods (for Shipment/Tax/Total)
                combined_total_weight = 0
                total_collectable_amount = 0 # Amount to be collected (COD + Partial)
                any_cod_or_partial = False
                
                # Payment classification lists
                prepaid_list = ["payu payment received", "bill desk payment received", "phonepe payment received", "bank payment received", "settlement for non-cod", "bank payment recieved"]
                cod_list = ["cash on delivery", "cash on delhivery", "settlement for cod"]
                partial_list = ["bank partial payment received", "bank partial payment recieved"]

                for order_id, order_list in all_orders_data.items():
                    for order in order_list:

                        payment_status = order["payment_status"].strip().lower()

                        if payment_status in cod_list or payment_status in partial_list:
                            any_cod_or_partial = True

                        order_total = 0

                        if "view_details" in order:
                            for item in order["view_details"]:

                                item_price = float(item.get("price", 0))
                                item_qty = int(item.get("quantity", 0))
                                item_weight = float(item.get("weight", 0))

                                order_total += item_price * item_qty
                                combined_total_weight += item_weight

                        combined_total_valuation += order_total

                        # Calculate how much of this order is collectable
                        if payment_status in cod_list:
                            total_collectable_amount += order_total
                        elif payment_status in partial_list:

                            advance_text = str(order.get("advance_payment", ""))

                            match = re.search(r"\d+(\.\d+)?", advance_text)

                            if match:
                                advance_payment = float(match.group())
                            else:
                                advance_payment = 0

                            remaining = order_total - advance_payment

                            total_collectable_amount += max(0, remaining)

                logger.info(f" Combined Total Valuation: {combined_total_valuation}")
                logger.info(f" Combined Total Weight: {combined_total_weight}")
                logger.info(f" Total Collectable (COD): {total_collectable_amount}")
                logger.info(f" Any COD or Partial: {any_cod_or_partial}")

                if combined_total_valuation > 8000 or combined_total_weight > 13000:
                    reason = f"Combined totals exceed limit (Price: {combined_total_valuation}, Weight: {combined_total_weight}g)"
                    logger.warning(f"!!! Skipping clubbing: {reason}")
                    
                    for ord_no in all_orders_data.keys():
                        log_order_status(ord_no, "SKIPPED", reason)

                    if buyer_status_tab in driver.window_handles:
                        driver.close()
                    driver.switch_to.window(main_tab)
                    i += 1
                    continue


                # -------- STEP 1 : OPEN DELHIVERY AND START SHIPMENT --------
                logger.info(" Opening Delhivery now...")

                open_delhivery_and_login()
                select_agribegri_surface()
                time.sleep(2)

                open_ready_to_ship_page()
                time.sleep(2)

                click_create_forward_shipment()

                combined_order = enter_combined_order_number(order_nos)

                # Collect product names from all valid orders
                product_names = []
                for order_list in all_orders_data.values():
                    for order in order_list:
                        if "view_details" in order:
                            for item in order["view_details"]:
                                name = item["product_name"].strip()
                                product_names.append(name)

                unique_products = list(set(product_names))
                combined_product_name = " , ".join(unique_products)

                logger.info(f" Combined Clean Product Name: {combined_product_name}")

                enter_product_name(combined_product_name)
                enter_product_category()

                # Enter the full valuation of goods into Delhivery fields
                # Decide correct shipment value
                if any_cod_or_partial:
                    shipment_value = total_collectable_amount
                else:
                    shipment_value = combined_total_valuation

                logger.info(f"Final Shipment Value sent to Delhivery: {shipment_value}")

                enter_shipment_values(shipment_value)
                # Select Payment Mode in Delhivery
                # If any order is COD or Partial, final mode is COD.
                final_mode = "COD" if any_cod_or_partial else "PREPAID"
                is_cod_or_partial = any_cod_or_partial # Update flag for COD entry logic

                # Use a dummy payment_status to trigger the right mode in the helper
                select_payment_mode_based_on_status(final_mode)


                                #  Get seller name (from your collected data)
                company_name = None

                for order_list in all_orders_data.values():
                    for order in order_list:
                        if "view_details" in order and order["view_details"]:
                            company_name = order["view_details"][0].get("company")
                            break
                    if company_name:
                        break
                for order_list in all_orders_data.values():
                    for order in order_list:
                        seller_name = order.get("seller_name")
                        break
                    if seller_name:
                        break

                if not seller_name:
                    logger.warning(" Could not determine seller name")
                    return


                pickup_name = seller_pickup_map.get(seller_name)

                if pickup_name:
                    select_pickup_location(pickup_name)
                else:
                    logger.warning(f" Pickup not found in Excel for seller: {seller_name}")
                
                time.sleep(2)

                # ADD THIS HERE
                click_add_seller_details()

                time.sleep(2)

                logger.info(" Delhivery flow completed.")

                #  Switch back to Buyer Status tab (Agribegri side)
                driver.switch_to.window(buyer_status_tab)

                #  Get any one View link from dictionary
                first_view_link = None

                for order_list in all_orders_data.values():
                    for order in order_list:
                        first_view_link = order["actions"]["view"]
                        break
                    if first_view_link:
                        break

                if first_view_link:
                    logger.info(f" Opening one View page: {first_view_link}")
                    driver.get(first_view_link)

                    wait.until(
                        EC.presence_of_element_located(
                            (By.XPATH, "//h5[contains(text(),'Shipping Address')]")
                        )
                    )


                    full_page_data = extract_full_order_page_details()


                    #  Switch back to Delhivery tab
                    if delhivery_tab and delhivery_tab in driver.window_handles:
                        logger.info(" Switching back to Delhivery...")
                        driver.switch_to.window(delhivery_tab)

                        time.sleep(2)

                        # GET seller name from extracted dictionary
                        seller_name = full_page_data["seller_details"]["name"]

                        enter_seller_name_in_delhivery(seller_name)

                        time.sleep(2)

                        click_add_seller_address_checkbox()
                        time.sleep(2)

                        fill_seller_address_in_delhivery(full_page_data)
                        time.sleep(2)

                        click_final_add_seller_details_button()
                        time.sleep(3)

                        click_add_customer_details()
                        time.sleep(2)

                        fill_customer_details_in_delhivery(full_page_data)
                        time.sleep(2)

                        click_final_add_customer_button()
                        time.sleep(3)

                        open_package_type_dropdown()
                        time.sleep(1)
                        select_cardboard_box()
                        time.sleep(1)

                        dimension_string = get_highest_price_dimension(all_orders_data)

                        length, breadth, height = split_dimension(dimension_string)

                        fill_package_dimensions(length, breadth, height)

                        total_weight = get_total_clubbed_weight_grams(all_orders_data)

                        logger.info(
                            f"Weight Debug → Combined Calc: {combined_total_weight}g | "
                            f"Final Clubbed Weight: {total_weight}g"
                        )

                        fill_packaged_weight(total_weight)


                        total_qty = get_total_product_qty(all_orders_data)

                        print("TOTAL QTY:", total_qty)
                        logger.info(f"TOTAL QTY BEFORE SET: {total_qty}")

                        set_item_count(total_qty)
                        time.sleep(2)


                        # =====================================================
                        #  FINAL COD CHECK BEFORE CREATE FORWARD SHIPMENT
                        # =====================================================


                        try:
                            time.sleep(2)

                            click_final_create_forward_shipment()
                            time.sleep(5)
                            download_start = time.time()
                            click_print_shipping_label()

                            label_success = wait_for_pdf_download_and_rename(combined_order, download_start)

                            if not label_success:
                                logger.error(" Shipping label PDF not downloaded!")

                                log_manual_required(
                                    order_nos,
                                    "Shipping Label PDF not downloaded"
                                )

                                # Close Delhivery tab safely
                                if delhivery_tab in driver.window_handles:
                                    driver.close()

                                driver.switch_to.window(main_tab)
                                continue

                        except Exception as shipment_error:
                            logger.error(f" Shipment Creation Failed: {shipment_error}")

                            #  Log all orders as manual required
                            log_manual_required(
                                order_nos,
                                "Problem in Create Forward Shipment / Print Label"
                            )

                            #  Close Delhivery safely
                            if delhivery_tab in driver.window_handles:
                                driver.close()

                            #  Switch back to main tab
                            driver.switch_to.window(main_tab)

                            logger.info(" Skipping this order and moving to next...")
                            continue


                        # ======================================================
                        #  NOW GENERATE INVOICE FOR EACH CLUBBED ORDER
                        # ======================================================

                        logger.info(" Starting Invoice Generation for all clubbed orders...")

                        invoice_index = 1

                        #  Switch back to Buyer Status tab
                        driver.switch_to.window(buyer_status_tab)

                        invoice_failed = False

                        for order_no, order_list in all_orders_data.items():

                            for order in order_list:

                                view_link = order["actions"]["view"]

                                if not view_link:
                                    continue

                                success = process_single_view_link(
                                    view_link,
                                    combined_order,
                                    invoice_index
                                )

                                if not success:
                                    invoice_failed = True
                                    logger.error(" Invoice PDF failed for %s", order_no)
                                    break

                                invoice_index += 1
                                break

                            if invoice_failed:
                                break


                        if invoice_failed:
                            log_manual_required(
                                order_nos,
                                "Invoice PDF not downloaded"
                            )

                            driver.switch_to.window(main_tab)
                            continue  #  ensures only one invoice per order_no

                        seller_email = full_page_data["seller_details"]["email"]

                        # ==================== STEP 4: REMARKS (CRITICAL) ====================
                        logger.info(f" Adding remarks for all {len(all_orders_data)} clubbed orders...")

                        for club_order_no, order_list_for_remark in all_orders_data.items():
                            for order in order_list_for_remark:
                                remark_url = order["actions"].get("remark")
                                if remark_url:
                                    try:
                                        remark_process_in_new_tab(
                                            club_order_no,
                                            combined_order,
                                            target_url=remark_url
                                        )
                                        break
                                    except Exception as exc_rem:
                                        logger.error(f" Remark failed for {club_order_no}: {exc_rem}")

                        # ==================== STEP 5: EMAIL ====================
                        try:
                            email_success = send_last_pdfs_to_seller(
                                seller_email,
                                combined_order,
                                len(all_orders_data),
                                seller_name,
                                company_name
                            )

                            if not email_success:
                                log_manual_required(order_nos, "Email sending failed")
                        except Exception as ee:
                            logger.error(f" Email Process Error: {ee}")
                            log_manual_required(order_nos, f"Email error: {ee}")

                        for order_no in all_orders_data.keys():
                            log_order_status(order_no, "PROCESSED", "Shipment + Invoice + Email Done")

                    else:
                        logger.warning(" Delhivery tab not found!")


                else:
                    logger.warning(" No View link found")

                logger.info(" Order completed. Continuing next rows...")
                driver.switch_to.window(main_tab)

                # Restart Chrome after 10 orders completed (user request)
                completed_orders_count += 1
                if completed_orders_count > 0 and completed_orders_count % 10 == 0:
                    logger.info(f"Closing Chrome after 10 orders completed. Total processed: {completed_orders_count}")
                    return True # Indicate needs restart

                continue


            
            except Exception as e:
                error_message = str(e)
                full_trace = traceback.format_exc()

                logger.error(f"\n FULL ERROR TRACE: {full_trace}")
                print(full_trace)

                # Try to get order number safely
                try:
                    current_order = order_no
                except:
                    try:
                        current_order = list(order_dict.keys())[0]
                    except:
                        current_order = f"PAGE_{page}_ROW_{i}"

                log_order_status(
                    current_order,
                    "ERROR",
                    error_message[:500]   # limit length for Excel
                )

                continue

        # Move to next page ONLY if not last page
        # Move to next page ONLY if not last page
        if page < total_pages:

            next_btn = wait.until(
                EC.element_to_be_clickable((
                    By.XPATH,
                    "//a[contains(@class,'cus_page_act') and text()='Next']"
                ))
            )

            driver.execute_script("arguments[0].click();", next_btn)

            wait.until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "table tbody tr"))
            )

            logger.info(f" Moved to page {page + 1}")
    # AFTER ALL PAGES FINISHED
    logger.info("\n All pages processed. Returning to First page...")

    first_btn = wait.until(
        EC.element_to_be_clickable((
            By.XPATH,
            "//a[contains(@class,'cus_page_act') and text()='First']"
        ))
    )

    driver.execute_script("arguments[0].click();", first_btn)

    wait.until(
        EC.presence_of_element_located((By.CSS_SELECTOR, "table tbody tr"))
    )

    logger.info(" Returned to Page 1 successfully")
    return False


def collect_manage_user_orders_table():
    rows = wait.until(
        EC.presence_of_all_elements_located(
            (By.XPATH, "//table[@id='dyntable']/tbody/tr")
        )
    )

    orders = []

    for r in rows:
        tds = r.find_elements(By.TAG_NAME, "td")

        # ---------------- ACTION LINKS ----------------
        actions_td = tds[15]   # Actions column (last)

        remark_link = None
        view_link = None
        truck_link = None

        links = actions_td.find_elements(By.TAG_NAME, "a")
        for a in links:
            href = a.get_attribute("href")
            if not href:
                continue

            if "#modal" in href:
                remark_link = href
            elif "edit_order.php" in href:
                view_link = href
            elif "#surface_modal" in href:
                truck_link = href

        # ---------------- ORDER DATA ----------------
        orders.append({
            "order_status": tds[0].text.strip(),
            "seller_name": tds[6].text.strip(),
            "order_no": tds[3].text.strip().split("\n")[0],
            "product_name": tds[5].text.strip(),
            "ord_by": tds[7].text.strip(),
            "payment_status": tds[11].text.strip(),
            "advance_payment": tds[13].text.strip(),
            "total_for_investigation": tds[14].text.strip(),

            #  ACTION LINKS
            "actions": {
                "remark": remark_link,
                "view": view_link,
                "truck": truck_link
            }
        })

    return orders

def fetch_view_order_details(view_url):
    logger.info("\n====================================")
    logger.info(" FETCHING VIEW PAGE DETAILS")
    logger.info(f"URL: {view_url}")
    logger.info("====================================\n")


    buyer_status_tab = driver.current_window_handle

    driver.execute_script("window.open(arguments[0], '_blank');", view_url)
    driver.switch_to.window(driver.window_handles[-1])

    wait.until(
        EC.presence_of_element_located((By.ID, "dyntable"))
    )

    rows = driver.find_elements(By.XPATH, "//table[@id='dyntable']/tbody/tr")

    products = []

    total_price_value = 0
    total_weight_value = 0

    for r in rows:
        tds = r.find_elements(By.TAG_NAME, "td")

        # ---- PRICE ----
        price_text = tds[9].text.strip().replace(",", "")
        price_clean = re.sub(r"[^\d.]", "", price_text)
        price_value = float(price_clean) if price_clean else 0

        # ---- WEIGHT ----
        # weight_text = tds[4].text.strip().lower()
        # weight_num_match = re.search(r"(\d+\.?\d*)", weight_text)
        # weight_num = float(weight_num_match.group(1)) if weight_num_match else 0

        weight_text = tds[4].text.strip().lower()

        # remove commas before regex
        weight_text_clean = weight_text.replace(",", "")

        weight_num_match = re.search(r"(\d+\.?\d*)", weight_text_clean)
        weight_num = float(weight_num_match.group(1)) if weight_num_match else 0
        
        # Check for Kg or Ltr units (both are * 1000)
        if any(unit in weight_text for unit in ["kg", "ltr", "liter", "litre"]):
            weight_per_unit = weight_num * 1000
        else:
            weight_per_unit = weight_num
            
        logger.info(f" Weight Text: '{weight_text}' -> Num: {weight_num} -> Per Unit: {weight_per_unit}g")

        # ---- QUANTITY ----
        qty_text = tds[7].text.strip()
        qty_value = int(float(re.sub(r"[^\d.]", "", qty_text)))

        weight_value = weight_per_unit * qty_value

        total_price_value += price_value * qty_value
        total_weight_value += weight_value

        try:
            dimension = tds[5].find_element(By.CLASS_NAME, "dimension_lbl").text.strip()
        except:
            dimension = tds[5].text.strip()

        products.append({
            "product_name": tds[1].text.strip(),
            "company": tds[2].text.strip(),
            "seller": tds[3].text.strip(),
            "weight": weight_value,
            "dimension": dimension,
            "courier_charge": tds[6].text.strip(),
            "quantity": qty_value,
            "price": price_value
        })

    print(f" Total Price: {total_price_value}")
    print(f"Total Weight: {total_weight_value}")

    #  WEIGHT CHECK
    if total_weight_value < 13000:
        logger.info(" Weight Under Control")
    else:
        logger.warning(f" Weight Exceeds Limit (>= 13000g) Current: {total_weight_value}")

    #  PRICE CHECK
    if total_price_value < 8000:
        logger.info(" Price Under Control")
    else:
        logger.warning(f" Price Exceeds Limit (>= 8000) Current: {total_price_value}")

    # FINAL VALIDATION
    is_valid_order = not (
        total_price_value > 8000 or total_weight_value > 13000
    )



    driver.close()
    driver.switch_to.window(buyer_status_tab)

   #  wait for buyer table to reload
    wait.until(
        EC.presence_of_element_located((By.XPATH, "//table[@id='dyntable']"))
    )

    return products, is_valid_order

def extract_full_order_page_details():
    logger.info(" Extracting Full Order Page Details...")

    details = {}

    # ================= SHIPPING ADDRESS =================
    shipping = {}

    shipping["name"] = driver.find_element(
        By.CSS_SELECTOR, ".name_lbl"
    ).text.strip()

    shipping["email"] = driver.find_element(
        By.XPATH, "//td[text()='Email :']/following-sibling::td"
    ).text.strip()

    shipping["address"] = driver.find_element(
        By.CSS_SELECTOR, ".address_lbl"
    ).text.strip()

    shipping["landmark"] = driver.find_element(
        By.CSS_SELECTOR, ".landmark_lbl"
    ).text.strip()

    shipping["pincode"] = driver.find_element(
        By.CSS_SELECTOR, ".zipcode_lbl"
    ).text.strip()

    shipping["phone"] = driver.find_element(
        By.CSS_SELECTOR, ".phone_lbl"
    ).text.strip()

    shipping["city"] = driver.find_element(
        By.CSS_SELECTOR, ".city_lbl"
    ).text.strip()

    shipping["taluka"] = driver.find_element(
        By.CSS_SELECTOR, ".taluka_lbl"
    ).text.strip()

    shipping["district"] = driver.find_element(
        By.CSS_SELECTOR, ".district_lbl"
    ).text.strip()

    shipping["state"] = driver.find_element(
        By.CSS_SELECTOR, ".state_lbl"
    ).text.strip()

    details["shipping_address"] = shipping


    # ================= SELLER DETAILS =================
    seller = {}

    seller["name"] = driver.find_element(
        By.XPATH, "//h5[text()='Seller Details']/following::table[1]//tr[1]/td[2]"
    ).text.strip()

    seller["email"] = driver.find_element(
        By.XPATH, "//h5[text()='Seller Details']/following::table[1]//tr[2]/td[2]"
    ).text.strip()

    seller["phone"] = driver.find_element(
        By.XPATH, "//h5[text()='Seller Details']/following::table[1]//tr[3]/td[2]"
    ).text.strip()

    seller["city"] = driver.find_element(
        By.XPATH, "//h5[text()='Seller Details']/following::table[1]//tr[4]/td[2]"
    ).text.strip()

    seller["postcode"] = driver.find_element(
        By.XPATH, "//h5[text()='Seller Details']/following::table[1]//tr[5]/td[2]"
    ).text.strip()

    details["seller_details"] = seller


    # ================= PICKUP ADDRESS =================
    pickup = {}

    pickup["address"] = driver.find_element(
        By.XPATH, "//h5[text()='Pickup Address']/following::table[1]//tr[1]/td[2]"
    ).text.strip()

    pickup["city"] = driver.find_element(
        By.XPATH, "//h5[text()='Pickup Address']/following::table[1]//tr[2]/td[2]"
    ).text.strip()

    pickup["state"] = driver.find_element(
        By.XPATH, "//h5[text()='Pickup Address']/following::table[1]//tr[3]/td[2]"
    ).text.strip()

    pickup["pincode"] = driver.find_element(
        By.XPATH, "//h5[text()='Pickup Address']/following::table[1]//tr[4]/td[2]"
    ).text.strip()

    details["pickup_address"] = pickup

    logger.info(" Full Order Page Data Extracted")
    logger.info(details)

    return details


def open_delhivery_and_login():
    global delhivery_tab

    if delhivery_tab and delhivery_tab in driver.window_handles:
        logger.info(" Delhivery already open. Switching...")
        driver.switch_to.window(delhivery_tab)
        return

    logger.info(" Opening Delhivery...")

    driver.execute_script(
        "window.open('https://one.delhivery.com/v2/login','_blank')"
    )

    driver.switch_to.window(driver.window_handles[-1])
    delhivery_tab = driver.current_window_handle


    try:
        WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.XPATH, "//div[contains(text(),'Domestic')]"))
        )
        logger.info(" Already logged in - Skipping login steps")

    except TimeoutException:
        logger.info(" Not logged in. Proceeding with login...")

        # Enter email
        email_input = wait.until(
            EC.visibility_of_element_located((By.NAME, "email"))
        )
        email_input.clear()
        email_input.send_keys("complain@agribegri.com")
        logger.info(" Email entered")

        # Click Continue
        continue_btn = wait.until(
            EC.presence_of_element_located(
                (By.XPATH, "//button[contains(text(),'Continue')]")
            )
        )
        driver.execute_script("arguments[0].click();", continue_btn)
        logger.info(" Continue clicked")

        # Enter password
        password_input = wait.until(
            EC.visibility_of_element_located((By.XPATH, "/html/body/div/div/div[2]/div[2]/div[2]/form/div[2]/div/div/section/input"))
        )
        password_input.send_keys("Agribegri@CL#26")
        logger.info(" Password entered")

        # Click Login
        login_btn = wait.until(
            EC.presence_of_element_located((By.XPATH, "/html/body/div/div/div[2]/div[2]/div[2]/form/button"))
        )
        driver.execute_script("arguments[0].click();", login_btn)
        logger.info(" Login button clicked")

        # Wait for dashboard
        wait.until(
            EC.presence_of_element_located((By.XPATH, "//div[contains(text(),'Domestic')]"))
        )

        logger.info(" Delhivery Login Successful")

def select_agribegri_surface():
    logger.info(" Selecting AGRIBEGRI SURFACE...")

    # Wait for dashboard
    wait.until(
        EC.presence_of_element_located(
            (By.XPATH, "//div[contains(text(),'Domestic')]")
        )
    )

    # ================== CLICK DOMESTIC DROPDOWN ==================

    domestic_dropdown = wait.until(
        EC.element_to_be_clickable((
            By.XPATH,
            "//button[contains(@class,'ap-menu-trigger-root') and .//i[contains(@class,'fa-truck')]]"
        ))
    )

    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", domestic_dropdown)
    driver.execute_script("arguments[0].click();", domestic_dropdown)

    logger.info(" Domestic dropdown clicked")

    # ================== SELECT AGRIBEGRI SURFACE ==================

    agribegri_surface = wait.until(
        EC.element_to_be_clickable((
            By.XPATH,
            "//button[contains(@class,'ap-menu-item')]//div[normalize-space()='AGRIBEGRI SURFACE']/ancestor::button"
        ))
    )

    driver.execute_script("arguments[0].click();", agribegri_surface)

    logger.info(" AGRIBEGRI SURFACE selected")

    time.sleep(1.5)

def open_ready_to_ship_page():
    logger.info(" Opening Ready To Ship page...")

    driver.get("https://one.delhivery.com/shipments/forward/ready-to-ship")

    # wait for Create Shipment button instead of text
    wait.until(
        EC.presence_of_element_located((
            By.XPATH,
            "//button[@data-action='create-shipment']"
        ))
    )

    logger.info(" Ready To Ship page loaded successfully")

def click_create_forward_shipment():
    logger.info(" Clicking Create Forward Shipment...")

    try:
        create_btn = wait.until(
            EC.element_to_be_clickable((
                By.XPATH,
                "//button[@data-action='create-shipment']"
            ))
        )

        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", create_btn)
        time.sleep(1)

        driver.execute_script("arguments[0].click();", create_btn)

        logger.info(" Create Forward Shipment clicked")

    except Exception as e:
        logger.error(" Failed to click Create Shipment: %s", e)

def enter_combined_order_number(order_nos):
    """
    order_nos example:
    ['845010YO - W', '845015GD - W']
    """

    logger.info(" Entering combined Order ID (without -W)...")

    cleaned_orders = []

    for order in order_nos:
        # Remove space + dash + anything after
        clean = order.split("-")[0].strip()
        cleaned_orders.append(clean)

    combined_order = "-".join(cleaned_orders)

    logger.info(f" Combined Order ID: {combined_order}")

    order_input = wait.until(
        EC.element_to_be_clickable((
            By.XPATH,
            "//input[@placeholder='Enter Order ID / Reference Number']"
        ))
    )

    order_input.clear()
    time.sleep(0.5)

    order_input.send_keys(combined_order)

    logger.info(" Order ID entered successfully")

    return combined_order


def enter_product_name(product_name):
    logger.info(" Entering Shipment Description...")

    product_input = wait.until(
        EC.element_to_be_clickable((
            By.XPATH,
            "//input[@placeholder='Enter a description of the item']"
        ))
    )

    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", product_input)
    time.sleep(0.5)

    product_input.clear()
    time.sleep(0.5)

    product_input.send_keys(product_name)

    logger.info(" Shipment Description entered successfully")

def enter_product_category():
    logger.info(" Entering Product Category...")

    category_input = wait.until(
        EC.element_to_be_clickable((
            By.XPATH,
            "//input[@placeholder='Select Product Category']"
        ))
    )

    category_input.clear()
    time.sleep(0.5)

    category_input.send_keys("Agriculture")

    time.sleep(1)  # wait for dropdown suggestions

    #  Select first dropdown option
    category_input.send_keys("\n")

    logger.info(" Product Category selected: Agriculture")

def enter_shipment_values(shipment_value):
    logger.info(" Entering Shipment / Tax / Total values...")

    shipment_value = float(shipment_value)
    tax_value = 0
    total_value = shipment_value + tax_value

    # Shipment Value
    shipment_input = wait.until(
        EC.element_to_be_clickable((
            By.XPATH,
            "//input[@placeholder='Enter Item Value']"
        ))
    )

    driver.execute_script("""
        arguments[0].value = arguments[1];
        arguments[0].dispatchEvent(new Event('input'));
    """, shipment_input, shipment_value)

    logger.info(" Shipment Value entered")

    # Tax Value
    tax_input = wait.until(
        EC.element_to_be_clickable((
            By.XPATH,
            "//input[@placeholder='Enter Tax Value']"
        ))
    )

    driver.execute_script("""
        arguments[0].value = arguments[1];
        arguments[0].dispatchEvent(new Event('input'));
    """, tax_input, tax_value)

    logger.info(" Tax Value entered")

    # Total Value
    total_input = wait.until(
        EC.element_to_be_clickable((
            By.XPATH,
            "//input[@placeholder='Enter Total Value']"
        ))
    )

    driver.execute_script("""
        arguments[0].value = arguments[1];
        arguments[0].dispatchEvent(new Event('input'));
    """, total_input, total_value)

    logger.info(" Total Value entered successfully")



def get_payment_type_from_view_page(view_url):
    logger.info(" Checking Payment Status from View Page...")

    current_tab = driver.current_window_handle

    driver.execute_script("window.open(arguments[0], '_blank');", view_url)
    driver.switch_to.window(driver.window_handles[-1])

    wait.until(
        EC.presence_of_element_located((By.ID, "abo_payment_type"))
    )

    payment_dropdown = Select(driver.find_element(By.ID, "abo_payment_type"))
    selected_payment = payment_dropdown.first_selected_option.text.strip()

    logger.info(" Payment Status Found: %s", selected_payment)

    driver.close()
    driver.switch_to.window(current_tab)

    return selected_payment

def select_payment_mode_based_on_status(payment_status):

    logger.info(" Selecting Payment Mode in Delhivery...")

    payment_status = payment_status.strip().upper()

    if payment_status == "COD":
        mode_to_select = "COD"
    elif payment_status == "PREPAID":
        mode_to_select = "PREPAID"
    else:
        # Fallback for individual values if passed directly
        prepaid_list = ["payu payment received", "bill desk payment received", "phonepe payment received", "bank payment received", "settlement for non-cod", "bank payment recieved"]
        if payment_status.lower() in prepaid_list:
            mode_to_select = "PREPAID"
        else:
            mode_to_select = "COD"

    #  Click dropdown
    #  Click Payment Mode dropdown (STRUCTURE SAFE)
    dropdown = wait.until(
        EC.element_to_be_clickable((
            By.XPATH,
            "//span[normalize-space()='Payment Mode']"
            "/ancestor::div[contains(@class,'ap-meta-label')]"
            "//div[contains(@class,'ap-menu-trigger')]"
        ))
    )

    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", dropdown)
    time.sleep(1)
    driver.execute_script("arguments[0].click();", dropdown)

    logger.info(" Payment dropdown opened")
    #  Select option (CORRECT STRUCTURE)
    option = wait.until(
        EC.element_to_be_clickable((
            By.XPATH,
            f"//ul[contains(@class,'ap-menu-items')]//span[@class='text main' and normalize-space()='{mode_to_select}']/ancestor::button"
        ))
    )

    driver.execute_script("arguments[0].click();", option)

    logger.info("%s selected successfully", mode_to_select)


def select_pickup_location(pickup_name):
    logger.info(" Selecting Pickup Location: %s", pickup_name)

    # 1️ Wait until Select Facility text is visible
    dropdown_btn = wait.until(
        EC.element_to_be_clickable((
            By.XPATH,
            "//button[contains(@class,'ucp__inputs--list-search__trigger') "
            "and .//span[normalize-space()='Select Facility']]"
        ))
    )

    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", dropdown_btn)
    time.sleep(1)

    # Try normal click first
    try:
        dropdown_btn.click()
    except:
        driver.execute_script("arguments[0].click();", dropdown_btn)

    logger.info(" Facility dropdown opened")

    # 2️ Wait for search box (IMPORTANT: overlay appears in body, not inside button)
    search_input = wait.until(
        EC.visibility_of_element_located((
            By.XPATH,
            "//input[@placeholder='Search Pickup Locations']"
        ))
    )

    search_input.clear()
    time.sleep(0.5)
    search_input.send_keys(pickup_name)

    logger.info(" Searching pickup location: %s", pickup_name)

    time.sleep(2)

    # 3️ Click exact matching result from overlay
    pickup_option = wait.until(
        EC.element_to_be_clickable((
            By.XPATH,
            f"//div[contains(@class,'ap-menu')]//span[normalize-space()='{pickup_name}']"
        ))
    )

    driver.execute_script("arguments[0].click();", pickup_option)

    logger.info(" Pickup location selected successfully")

def click_add_seller_details():
    logger.info(" Clicking Add Seller Details...")

    add_btn = wait.until(
        EC.element_to_be_clickable((
            By.XPATH,
            "//button[.//text()[contains(.,'Add Seller Details')]]"
        ))
    )

    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", add_btn)
    time.sleep(1)

    driver.execute_script("arguments[0].click();", add_btn)

    logger.info(" Add Seller Details clicked successfully")

def enter_seller_name_in_delhivery(seller_name):
    logger.info("Entering Seller Name: %s", seller_name)

    # Wait for Seller Name input box
    name_input = wait.until(
        EC.element_to_be_clickable((
            By.XPATH,
            "//input[@placeholder='Enter name']"
        ))
    )

    driver.execute_script(
        "arguments[0].scrollIntoView({block:'center'});", name_input
    )

    time.sleep(1)

    name_input.clear()
    name_input.send_keys(seller_name)

    logger.info(" Seller Name entered successfully")


def click_add_seller_address_checkbox():
    logger.info(" Clicking 'Add seller address' checkbox...")

    checkbox_label = wait.until(
        EC.element_to_be_clickable((
            By.XPATH,
            "//label[.//span[normalize-space()='Add seller address']]"
        ))
    )

    driver.execute_script(
        "arguments[0].scrollIntoView({block:'center'});", checkbox_label
    )

    time.sleep(1)

    driver.execute_script("arguments[0].click();", checkbox_label)

    logger.info(" Checkbox clicked successfully")


def fill_seller_address_in_delhivery(full_page_data):
    logger.info(" Filling Seller Address...")

    pickup_data = full_page_data["pickup_address"]

    full_address = pickup_data["address"]
    pincode = pickup_data["pincode"]

    #  Remove pincode from end of address (if present)
    cleaned_address = re.sub(rf"\s*-?\s*{pincode}\s*$", "", full_address).strip()

    logger.info(" Cleaned Address: %s", cleaned_address)
    logger.info(" Pincode: %s", pincode)

    # ---------------- Address Line 1 ----------------
    address_input = wait.until(
        EC.element_to_be_clickable((
            By.XPATH,
            "//input[@placeholder='Address Line 1']"
        ))
    )

    driver.execute_script(
        "arguments[0].scrollIntoView({block:'center'});", address_input
    )

    address_input.clear()
    time.sleep(1)
    address_input.send_keys(cleaned_address)

    logger.info(" Address Line 1 filled")

    # ---------------- Pincode ----------------
    pincode_input = wait.until(
        EC.element_to_be_clickable((
            By.XPATH,
            "//input[@placeholder='Enter pincode']"
        ))
    )

    pincode_input.clear()
    time.sleep(1)
    pincode_input.send_keys(pincode)

    logger.info(" Pincode filled successfully")

def click_final_add_seller_details_button():
    logger.info(" Clicking final 'Add Seller Details' button...")

    add_button = wait.until(
        EC.element_to_be_clickable((
            By.XPATH,
            "//button[@label='Add Seller Details']"
        ))
    )

    driver.execute_script(
        "arguments[0].scrollIntoView({block:'center'});", add_button
    )

    time.sleep(1)

    driver.execute_script("arguments[0].click();", add_button)

    logger.info(" Final Add Seller Details button clicked successfully")

def click_add_customer_details():
    logger.info(" Clicking 'Add Customer Details'...")

    customer_btn = wait.until(
        EC.visibility_of_element_located((
            By.XPATH,
            "//div[@showshippingaddress='false' and normalize-space()='Add Customer Details']"
        ))
    )

    actions = ActionChains(driver)
    actions.move_to_element(customer_btn).pause(1).click().perform()

    logger.info(" Add Customer Details clicked successfully")

    time.sleep(3)

def fill_customer_details_in_delhivery(full_page_data):
    logger.info(" Filling Customer Details Form...")

    shipping = full_page_data["shipping_address"]

    full_name = shipping["name"]
    phone = shipping["phone"]
    address = shipping["address"]
    landmark = shipping.get("landmark", "")
    pincode = shipping["pincode"]
    city = shipping["city"]
    taluka = shipping["taluka"]
    district = shipping["district"]
    state = shipping["state"]

    # -------- Split Name --------
    name_parts = full_name.strip().split()
    first_name = name_parts[0]
    last_name = name_parts[1] if len(name_parts) > 1 else name_parts[0]

    # ---------------- FIRST NAME ----------------
    wait.until(EC.element_to_be_clickable(
        (By.NAME, "customer_first_name")
    )).send_keys(first_name)

    # ---------------- LAST NAME ----------------
    wait.until(EC.element_to_be_clickable(
        (By.NAME, "customer_last_name")
    )).send_keys(last_name)

    logger.info(" Name filled")

    # ---------------- PHONE ----------------
    phone_input = wait.until(
        EC.element_to_be_clickable((By.NAME, "phone_number"))
    )
    phone_input.clear()
    phone_input.send_keys(phone)

    logger.info(" Phone filled")

    # ---------------- ADDRESS LINE 1 ----------------
    full_address = f"{address}, {landmark}".strip(", ")

    address_input = wait.until(
        EC.element_to_be_clickable((
            By.XPATH,
            "//input[@placeholder='Address Line 1']"
        ))
    )
    address_input.clear()
    address_input.send_keys(full_address)

    logger.info(" Address Line 1 filled")

    # ---------------- CITY / TALUKA / DISTRICT ----------------
    # ---------------- ADDRESS LINE 2 ----------------
    address_line_2 = f"{city}, {taluka}, {district}, {state}"

    address2_input = wait.until(
        EC.element_to_be_clickable((
            By.XPATH,
            "//input[@placeholder='Address Line 2']"
        ))
    )

    address2_input.clear()
    address2_input.send_keys(address_line_2)

    logger.info(" Address Line 2 filled")


    # ---------------- PINCODE ----------------
    pincode_input = wait.until(
        EC.element_to_be_clickable((
            By.XPATH,
            "//input[@placeholder='Enter pincode']"
        ))
    )
    pincode_input.clear()
    pincode_input.send_keys(pincode)

    logger.info(" Pincode filled")

    logger.info(" Customer Details Form Completed")

def click_final_add_customer_button():
    logger.info(" Clicking final 'Add Customer' button...")

    add_customer_btn = wait.until(
        EC.presence_of_element_located((
            By.XPATH,
            "//button[contains(@class,'blue') and contains(.,'Add Customer')]"
        ))
    )

    driver.execute_script("""
        arguments[0].scrollIntoView({block:'center'});
        arguments[0].click();
    """, add_customer_btn)

    logger.info(" Final Add Customer button clicked successfully")

    time.sleep(3)

def get_highest_price_dimension(all_orders_data):
    highest_price = 0
    selected_dimension = None

    for order_list in all_orders_data.values():
        for order in order_list:
            if "view_details" in order:
                for item in order["view_details"]:
                    if item["price"] > highest_price:
                        highest_price = item["price"]
                        selected_dimension = item["dimension"]

    logger.info(" Highest Price: %s", highest_price)
    logger.info(" Selected Dimension: %s", selected_dimension)

    return selected_dimension
def fill_package_dimensions(length, breadth, height):
    logger.info(" Filling Package Dimensions...")

    dimension_inputs = wait.until(
        EC.presence_of_all_elements_located((
            By.XPATH,
            "//input[@placeholder='L' or @placeholder='B' or @placeholder='H']"
        ))
    )

    # Ensure order L B H
    for input_box in dimension_inputs:
        placeholder = input_box.get_attribute("placeholder")

        if placeholder == "L":
            input_box.clear()
            input_box.send_keys(str(length))

        elif placeholder == "B":
            input_box.clear()
            input_box.send_keys(str(breadth))

        elif placeholder == "H":
            input_box.clear()
            input_box.send_keys(str(height))

    logger.info(" Dimensions filled successfully")


from selenium.webdriver.common.action_chains import ActionChains

def open_package_type_dropdown():
    logger.info(" Opening Package Type dropdown...")

    dropdown = wait.until(
        EC.element_to_be_clickable((
            By.XPATH,
            "//span[text()='Package Type']/ancestor::div[contains(@class,'ap-meta-label')]"
            "//div[contains(@class,'ap-menu-trigger')]"
        ))
    )

    driver.execute_script("""
        arguments[0].scrollIntoView({block:'center'});
        arguments[0].click();
    """, dropdown)

    logger.info(" Correct Package Type dropdown opened")

    time.sleep(2)

def select_cardboard_box():
    logger.info(" Selecting 'Cardboard Box'...")

    option = wait.until(
        EC.visibility_of_element_located((
            By.XPATH,
            "//div[contains(@class,'section')]//span[normalize-space()='Cardboard Box']"
        ))
    )

    ActionChains(driver).move_to_element(option).pause(0.5).click().perform()

    logger.info(" Cardboard Box selected successfully")

    time.sleep(2)


def get_total_clubbed_weight_grams(all_orders_data):
    total_weight_grams = 0

    for order_no, order_list in all_orders_data.items():

        for order in order_list:

            # Only Confirm
            if not order["order_status"].startswith("Confirm"):
                continue

            #  If view_details exists → use it
            if "view_details" in order:
                for item in order["view_details"]:
                    total_weight_grams += item.get("weight", 0)

            else:
                #  If view_details missing (skipped by validation)
                # Extract weight from product_name text

                product_text = order.get("product_name", "")

                # Example: "UPL Lancer Gold Insecticide - 10 Kg (1 Kg x 10 Qty)"
                match = re.search(r"\((.*?)\)", product_text)

                if match:
                    bracket_text = match.group(1)

                    # Extract weight per unit
                    weight_match = re.search(r"(\d+(?:\.\d+)?)\s*(Kg|Ltr|Liter|Litre)", bracket_text, re.IGNORECASE)
                    qty_match = re.search(r"x\s*(\d+)", bracket_text)

                    if weight_match and qty_match:
                        weight_num = float(weight_match.group(1))
                        weight_per_unit = weight_num * 1000
                        qty = int(qty_match.group(1))

                        calculated_weight = weight_per_unit * qty
                        total_weight_grams += calculated_weight

                        logger.info("Recovered weight from text for %s: %d (Unit: %s)", order_no, calculated_weight, weight_match.group(2))

    total_weight_grams = int(round(total_weight_grams))

    logger.info("FINAL Total Weight (grams): %d", total_weight_grams)

    return total_weight_grams

def fill_packaged_weight(weight_grams):
    logger.info(" Filling Packaged Weight (grams)...")

    weight_input = wait.until(
        EC.element_to_be_clickable((
            By.XPATH,
            "//input[@placeholder='Enter packaged weight']"
        ))
    )

    weight_input.clear()
    weight_input.send_keys(str(weight_grams))

    logger.info(" Packaged Weight entered successfully")

def get_total_product_qty(all_orders_data):

    total_qty = 0

    for order_no, order_list in all_orders_data.items():
        for order in order_list:

            if "view_details" in order:

                for item in order["view_details"]:

                    qty = int(item.get("quantity", 1))
                    product_name = item.get("product_name", "")

                    # Count bundle items
                    if "+" in product_name:
                        bundle_count = len(product_name.split("+"))
                    else:
                        bundle_count = 1

                    total_qty += qty * bundle_count

    logger.info(" FINAL Total Product Qty: %d", total_qty)

    return total_qty

def set_item_count(total_qty):
    logger.info(" Setting Item Count directly (final fix)...")

    #  Locate Item Count number input ONLY
    item_input = wait.until(
        EC.element_to_be_clickable((
            By.XPATH,
            "//span[normalize-space()='Item Count']"
            "/ancestor::label//input[@type='number']"
        ))
    )

    driver.execute_script("""
        arguments[0].focus();
        arguments[0].value = arguments[1];
        arguments[0].dispatchEvent(new Event('input', { bubbles: true }));
        arguments[0].dispatchEvent(new Event('change', { bubbles: true }));
    """, item_input, total_qty)

    logger.info(" Item Count set to %d successfully", total_qty)


def click_final_create_forward_shipment():
    logger.info(" Clicking final 'Create Forward Shipment' button...")

    create_btn = wait.until(
        EC.element_to_be_clickable((
            By.XPATH,
            "//button[normalize-space()='Create Forward Shipment']"
        ))
    )

    driver.execute_script("""
        arguments[0].scrollIntoView({block:'center'});
        arguments[0].click();
    """, create_btn)

    logger.info(" Final Create Forward Shipment clicked successfully")
def click_print_shipping_label():
    logger.info(" Clicking 'Print Shipping Label'...")

    print_btn = wait.until(
        EC.element_to_be_clickable((
            By.XPATH,
            "//button[normalize-space()='Print Shipping Label']"
        ))
    )

    driver.execute_script("arguments[0].click();", print_btn)

    logger.info(" Print Shipping Label clicked")


def wait_for_pdf_download_and_rename(order_reference, download_start_time):

    download_folder = download_dir
    logger.info("Waiting for PDF download...")

    timeout = 60
    start_time = time.time()
    downloaded_file = None

    while time.time() - start_time < timeout:

        files = os.listdir(download_folder)

        # Wait while Chrome downloading
        if any(f.endswith(".crdownload") for f in files):
            time.sleep(1)
            continue

        for file in files:
            if not file.lower().endswith(".pdf"):
                continue

            full_path = os.path.join(download_folder, file)

            #  IMPORTANT FIX
            if os.path.getmtime(full_path) >= download_start_time:

                size1 = os.path.getsize(full_path)
                time.sleep(1)
                size2 = os.path.getsize(full_path)

                if size1 == size2 and size1 > 0:
                    downloaded_file = full_path
                    break

        if downloaded_file:
            break

        time.sleep(1)

    if not downloaded_file:
        logger.warning("PDF download not detected within timeout!")
        return False

    new_path = os.path.join(download_folder, f"{order_reference}.pdf")

    if os.path.exists(new_path):
        os.remove(new_path)

    os.rename(downloaded_file, new_path)

    logger.info("PDF successfully renamed to: %s", new_path)
    return True
def process_view_link_and_download_invoice(view_link, combined_order, index):
    logger.info(" Processing View Link %d: %s", index, view_link)

    # Open view link
    driver.get(view_link)

    wait.until(
        EC.presence_of_element_located((By.ID, "abo_status"))
    )

    # ================== 1️ Select Packed ==================
    status_dropdown = Select(driver.find_element(By.ID, "abo_status"))
    status_dropdown.select_by_visible_text("Packed")
    logger.info(" Selected Status: Packed")

    # ================== 2 Select CL Surface ==================
    reason_dropdown = Select(driver.find_element(By.ID, "abo_packed_reason"))
    reason_dropdown.select_by_visible_text("CL Surface")
    logger.info(" Selected Packed Reason: CL Surface")

    # ================== 3️ Click Submit ==================
    submit_btn = driver.find_element(By.NAME, "update_order_status")
    driver.execute_script("arguments[0].click();", submit_btn)
    logger.info(" Submit clicked")

    # ================== 4️ Click OK ==================
    wait.until(
        EC.element_to_be_clickable((By.ID, "popup_ok"))
    ).click()
    logger.info(" Popup OK clicked")

    time.sleep(2)

    # ================== 5️ Click Generate Invoice ==================
    generate_btn = wait.until(
        EC.element_to_be_clickable((By.ID, "btnGenerateInvoice"))
    )
    driver.execute_script("arguments[0].click();", generate_btn)
    logger.info(" Generate Invoice clicked")

    # ================== 6️ Wait & Rename PDF ==================
    new_filename = f"{combined_order}-{index}"  # Don't add .pdf here
    wait_for_pdf_download_and_rename(new_filename)


def process_single_view_link(view_url, combined_order, index):
    logger.info(" Processing Invoice for Order #%d", index)

    # Open in same tab
    driver.get(view_url)

    wait.until(
        EC.presence_of_element_located((By.ID, "abo_status"))
    )

    # 1️ Select Packed
    status_dropdown = Select(driver.find_element(By.ID, "abo_status"))
    status_dropdown.select_by_visible_text("Packed")
    logger.info(" Status set to Packed")

    # 2️ Select CL Surface
    packed_reason = Select(driver.find_element(By.ID, "abo_packed_reason"))
    packed_reason.select_by_visible_text("CL Surface")
    logger.info(" Packed Reason set to CL Surface")

    # 3️ Click Submit
    submit_btn = driver.find_element(By.NAME, "update_order_status")
    driver.execute_script("arguments[0].click();", submit_btn)
    logger.info(" Submit clicked")

    # 4️ Click OK popup
    wait.until(EC.element_to_be_clickable((By.ID, "popup_ok")))
    driver.find_element(By.ID, "popup_ok").click()
    logger.info(" OK popup clicked")

    # 5️ Click Generate Invoice
    wait.until(EC.element_to_be_clickable((By.ID, "btnGenerateInvoice")))
    download_start = time.time()

    driver.find_element(By.ID, "btnGenerateInvoice").click()
    logger.info(" Generate Invoice clicked")

    invoice_name = f"{combined_order}-{index}"
    invoice_success = wait_for_pdf_download_and_rename(invoice_name, download_start)

    if not invoice_success:
        return False

    return True

def send_last_pdfs_to_seller(receiver_email, combined_order, order_count, seller_name, company_name):

    logger.info(" Preparing email...")

    sender_email = "dispatch.agribegri@gmail.com"
    sender_password = "ndkatxukfthopwyn"
    seller_clean = seller_name.strip().lower()

    company_clean = company_name.strip().lower() if company_name else ""
    company_clean = company_clean.replace(", india .", "")
    company_clean = company_clean.replace(", india", "")
    company_clean = company_clean.strip()

    special_sellers = [
        "agribegri trade link pvt. ltd.",
        "noble crop science",
        "rain bio tech"
    ]

    special_email_map = {
        "barrix agro science pvt. ltd.": "info@barrix.in",
        "neptune fairdeal products pvt. ltd": "crop10.order@gmail.com",
        "real trust exim corporation": "mangesh.ingawale@gmail.com"
    }

    #  ATPL + Special Company
    # ================= EMAIL DECISION LOGIC =================

    # ================= EMAIL DECISION LOGIC =================

    final_email = receiver_email  # default


    # ---------- ATPL ----------
    if seller_clean == "atpl":

        if "barrix" in company_clean:
            final_email = "info@barrix.in"
            logger.info("ATPL + Barrix → info@barrix.in")

        elif "neptune" in company_clean:
            final_email = "crop10.order@gmail.com"
            logger.info("ATPL + Neptune → crop10.order@gmail.com")

        elif "real trust" in company_clean:
            final_email = "mangesh.ingawale@gmail.com"
            logger.info("ATPL + Real Trust → mangesh.ingawale@gmail.com")

        else:
            final_email = receiver_email
            logger.info("ATPL + Other Company → seller email")


    # ---------- AGRIBEGRI TRADE LINK ----------
    elif "agribegri trade link" in seller_clean:
        final_email = "shipping.agribegri@gmail.com"
        logger.info("Agribegri Trade Link → shipping")


    # ---------- INTERNAL SELLERS ----------
    elif seller_clean in ["noble crop science", "rain bio tech"]:
        final_email = "shipping.agribegri@gmail.com"
        logger.info("Internal seller → shipping")


    # ---------- OTHER SELLERS ----------
    else:
        final_email = receiver_email
        logger.info("Other seller → seller email")



    receiver_email = final_email

    downloads_folder = r"C:\Websmith\Agribegri\Downloads"

    #  Filter only files belonging to this combined order
    pdf_files = [
        os.path.join(downloads_folder, f)
        for f in os.listdir(downloads_folder)
        if f.lower().endswith(".pdf")
        and f.startswith(combined_order)
    ]

    if not pdf_files:
        logger.warning(" No matching PDFs found for this order!")
        return False

    # Sort properly:
    # Label first (exact match), then -1, -2, -3...
    def sort_key(file_path):
        filename = os.path.basename(file_path)

        if filename == f"{combined_order}.pdf":
            return 0  # Shipping label first

        match = re.search(rf"{re.escape(combined_order)}-(\d+)\.pdf", filename)
        if match:
            return int(match.group(1))

        return 999

    pdf_files.sort(key=sort_key)

    latest_pdfs = pdf_files

    logger.info(" Sending Latest PDFs:")
    for idx, file in enumerate(latest_pdfs, start=1):
        logger.info("%d. %s", idx, os.path.basename(file))

    # ================= EMAIL =================

    msg = EmailMessage()
    msg["From"] = sender_email
    msg["To"] = receiver_email
    msg["Subject"] = f"Shipment Documents - {combined_order}"

    msg.set_content(f"""
Hello {seller_name},

Please find attached shipping label and invoices.

Order Reference: {combined_order}
Total Orders Clubbed: {order_count}

Regards,
Dispatch Team
""")

    for file_path in latest_pdfs:
        with open(file_path, "rb") as f:
            file_data = f.read()

        msg.add_attachment(
            file_data,
            maintype="application",
            subtype="pdf",
            filename=os.path.basename(file_path)
        )

    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
            server.login(sender_email, sender_password)
            server.send_message(msg)

        logger.info(" Email sent successfully to %s", receiver_email)
        return True

    except smtplib.SMTPAuthenticationError:
        logger.error(" Email Failed: Authentication Error. Check app password!")
        return False
    except Exception as e:
        logger.error(" Email Failed: %s", str(e))
        return False


def remark_process_in_new_tab(order_no, combined_order, target_url=None):
    print(f"\n Starting remark process for: {order_no}")

    if target_url:
        # Use the captured direct remark link
        driver.execute_script("window.open(arguments[0], '_blank');", target_url)
        driver.switch_to.window(driver.window_handles[-1])
        print(f" Direct remark page opened: {target_url}")
        
        # Wait for the table to load on this specific page
        try:
            wait.until(EC.presence_of_element_located((By.XPATH, "//table[@id='dyntable']/tbody/tr")))
        except:
             print(" Error: Table did not load on target URL")
    else:
        # ---------------- OPEN MANAGE ORDERS IN NEW TAB ----------------
        driver.execute_script(
            "window.open('https://agribegri.com/admin/manage_orders.php','_blank');"
        )
        driver.switch_to.window(driver.window_handles[-1])
        print(" Manage Orders opened in new tab")

        # ---------------- APPLY FILTERS ----------------
        click_from_date_30_days_ago()
        click_to_date_today()
        select_order_status_packed()
        select_shipping_through_courier()
        click_search_and_wait()

    # ---------------- SEARCH ORDER NUMBER ----------------
    search_input = wait.until(
        EC.element_to_be_clickable((By.XPATH, "//input[@aria-controls='dyntable']"))
    )

    search_input.clear()
    search_input.send_keys(order_no)

    driver.execute_script("""
        arguments[0].dispatchEvent(new Event('input'));
    """, search_input)

    print(" Searching order in filtered table...")
    time.sleep(5)

    # ---------------- WAIT FOR ROW ----------------
    try:
        row_element = wait.until(
            EC.presence_of_element_located((
                By.XPATH,
                f"//table[@id='dyntable']//td[contains(normalize-space(),'{order_no}')]/parent::tr"
            ))
        )
    except TimeoutException:
        print(f" Order {order_no} not found after filtering")

        log_order_status(
            order_no,
            "REMARK NOT ADDED",
            "Order not found in Packed filter"
        )

        driver.close()
        driver.switch_to.window(main_tab)
        return False

    # ---------------- CLICK REMARK BUTTON ----------------
    try:
        remark_btn = row_element.find_element(
            By.XPATH,
            ".//a[contains(@href,'#modal') and contains(@class,'get_id')]"
        )

        driver.execute_script("arguments[0].click();", remark_btn)

        print(" Remark button clicked")

    except Exception:
        print(" Remark button not found")

        log_order_status(
            order_no,
            "REMARK NOT FOUND",
            "Remark icon missing in row"
        )

        driver.close()
        driver.switch_to.window(main_tab)
        return False

    # ---------------- TYPE REMARK ----------------
    remark_box = wait.until(
        EC.element_to_be_clickable((By.ID, "order_remark"))
    )

    # Click automatically in the remark box
    driver.execute_script("arguments[0].click();", remark_box)
    time.sleep(0.5)

    remark_box.clear()
    remark_box.send_keys(combined_order)

    submit_btn = wait.until(
        EC.element_to_be_clickable((By.NAME, "submit_remark"))
    )

    driver.execute_script("arguments[0].click();", submit_btn)

    logger.info(" Remark submitted successfully")

    # ---------------- CLICK OK POPUP ----------------
    try:
        wait.until(EC.element_to_be_clickable((By.ID, "popup_ok"))).click()
        logger.info(" OK popup clicked")
    except:
        pass

    # ---------------- CLOSE TAB ----------------
    driver.close()
    driver.switch_to.window(main_tab)

    logger.info(" Returned to main tab")

    return True


while True:
    try:
        browser_start_time = time.time()
        
        # Initialize Chrome
        chrome_options = webdriver.ChromeOptions()
        prefs = {
            "download.default_directory": download_dir,
            "download.prompt_for_download": False,
            "download.directory_upgrade": True,
            "plugins.always_open_pdf_externally": True,
            "safebrowsing.enabled": True
        }
        chrome_options.add_experimental_option("prefs", prefs)
        service = Service(ChromeDriverManager().install())
        driver = webdriver.Chrome(service=service, options=chrome_options)
        wait = WebDriverWait(driver, 30)

        _USERNAME = sys.argv[1] if len(sys.argv) > 1 else 'Clubbed'
        _PASSWORD = sys.argv[2] if len(sys.argv) > 2 else 'Clubbed@022026'
        _OTP      = sys.argv[3] if len(sys.argv) > 3 else '123456'

        login_to_agribegri(_USERNAME, _PASSWORD, _OTP)
        main_tab = driver.current_window_handle

        step_wait(5, "After Login")

        click_manage_orders_nav()
        step_wait(5, "After clicking Manage Orders")

        click_from_date_30_days_ago()
        step_wait(5, "After setting From Date")

        click_to_date_today()
        step_wait(5, "After setting To Date")

        select_order_status_confirm()
        step_wait(5, "After selecting Order Status = Confirm")

        select_shipping_through_courier()
        step_wait(5, "After selecting Shipping Through Courier")

        click_search_and_wait()
        step_wait(5, "After Search results loaded")


        #  GET TOTAL ORDER COUNT
        total_orders = get_total_orders()

        if total_orders:
            records_per_page = 50
            total_pages = (total_orders // records_per_page) + (1 if total_orders % records_per_page else 0)

            logger.info(" Total Orders: %d", total_orders)
            logger.info(" Total Pages: %d", total_pages)

            #  GO BACK TO FIRST PAGE
            try:
                first_btn = wait.until(
                    EC.element_to_be_clickable((
                        By.XPATH,
                        "//a[contains(@class,'cus_page_act') and text()='First']"
                    ))
                )

                driver.execute_script("arguments[0].click();", first_btn)

                wait.until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, "table tbody tr"))
                )

                logger.info(" Returned to Page 1 successfully")

            except Exception as e:
                logger.warning(" First button not clickable, forcing Page 1 URL")
                driver.get("https://agribegri.com/admin/manage_orders.php?page=1")

                wait.until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, "table tbody tr"))
                )
        else:
            total_pages = 0
            logger.info(" No orders found to process.")

        #  LOAD EXCEL BEFORE PROCESSING
        seller_pickup_map = load_seller_pickup_mapping()
        initialize_log_file()

        #  NOW START PROCESSING FROM PAGE 1
        if total_pages > 0:
            needs_restart = process_all_orders_with_custom_pagination(total_pages, browser_start_time)
            
            if needs_restart:
                logger.info(" Rebooting browser (partial work done)...")
                try:
                    driver.quit()
                except:
                    pass
                time.sleep(5)
                continue # Immediate restart to continue processing the rest
            else:
                logger.info(" All found orders have been processed.")
        else:
            logger.info(" No orders were found in this cycle.")

        #  END OF CYCLE
        try:
            driver.quit()
        except:
            pass
            
        logger.info(" Waiting for 5 minutes before checking again (Cycle Finished)...")
        time.sleep(300)

    except Exception as e:
        logger.error(f" CRITICAL error in main loop: {e}")
        try:
            driver.quit()
        except:
            pass
        logger.info(" Restarting in 1 minute due to error...")
        time.sleep(60)
        continue