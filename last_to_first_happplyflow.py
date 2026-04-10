from selenium import webdriver
import time
from selenium.webdriver.common.by import By
from selenium.common.exceptions import NoSuchElementException, TimeoutException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
import os
import openpyxl
from openpyxl import Workbook, load_workbook
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
import datetime
import math
import sys
import json

BALANCE_FILE = "balance.txt"
ORDERS_COUNT_FILE = "orders_count.txt"
HISTORY_FILE = "processed_orders.json"
COST_PER_ORDER = 0.75

driver = None
browser_start_time = None

def get_balance():
    if not os.path.exists(BALANCE_FILE): return 0.0
    with open(BALANCE_FILE, "r") as f: return float(f.read().strip() or 0)

def update_balance(val):
    with open(BALANCE_FILE, "w") as f: f.write(str(round(val, 2)))

def update_history(entry):
    history = []
    if os.path.exists(HISTORY_FILE):
        with open(HISTORY_FILE, "r") as f:
            try: history = json.load(f)
            except: pass
    entry["timestamp"] = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    history.append(entry)
    with open(HISTORY_FILE, "w") as f: json.dump(history[-1000:], f, indent=4)
    with open(ORDERS_COUNT_FILE, "w") as f: f.write(str(len([e for e in history if e.get("status") == "Processed"])))

# ================== LOGGING SETUP ==================
import logging
import os
import re
LOG_DIR = "logs"
os.makedirs(LOG_DIR, exist_ok=True)

log_file_path = os.path.join(LOG_DIR, "automation.log")

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s | %(levelname)s | %(filename)s:%(lineno)d | %(message)s",
    handlers=[
        logging.FileHandler(log_file_path, mode="a", encoding="utf-8"),
        logging.StreamHandler()  # This prints to console also
    ]
)

logger = logging.getLogger(__name__)
# Initialize EXCEL LOGGING
EXCEL_FILE = "processing_report.xlsx"
WEIGHT_LIMIT_GRAM = 7000   
AMOUNT_LIMIT = 8000

# ================== EMAIL ROUTING FUNCTION ==================

# def get_email_for_seller(seller_name, company_name, seller_email):
#     seller_lower = seller_name.lower().strip()
#     company_lower = company_name.lower().strip()
#     logger.info(f"Email routing started | Seller: '{seller_name}' | Company: '{company_name}'")
#     special_sellers = [
#         "agribegri trade link pvt. ltd.",
#         "noble crop science",
#         "rain bio tech"
#     ]

#     company_email_map = {
#         "barrix agro science pvt. ltd.": "info@barrix.in",
#         "neptune fairdeal products pvt. ltd": "crop10.order@gmail.com",
#         "real trust exim corporation": "mangesh.ingawale@gmail.com"
#     }

#     # 1 Special Seller
#     if seller_lower in special_sellers:
#         logger.info("Matched Special Seller rule → shipping.agribegri@gmail.com")
#         return "shipping.agribegri@gmail.com"

#     elif seller_lower == "atpl":

#         # Clean company name (remove commas, dots, special chars)
#         cleaned_company = re.sub(r'[^a-z0-9 ]', '', company_lower)
#         cleaned_company = re.sub(r'\s+', ' ', cleaned_company).strip()

#         logger.info(f"ATPL Cleaned Company: {cleaned_company}")

#         for key, value in company_email_map.items():
#             cleaned_key = re.sub(r'[^a-z0-9 ]', '', key.lower())
#             cleaned_key = re.sub(r'\s+', ' ', cleaned_key).strip()

#             if cleaned_key in cleaned_company:
#                 logger.info(f"ATPL mapping matched → {value}")
#                 return value

#     # ️ Direct Company Mapping
#     # ️ Direct Company Mapping
#     elif company_lower in company_email_map:
#         logger.info(f"Direct company mapping matched → {company_email_map[company_lower]}")
#         return company_email_map[company_lower]

#     # 4 Default Seller Email DISABLED
#     logger.info("Default seller email disabled — no email for normal sellers")

#     logger.warning("No email routing rule matched → returning None")
#     return None

def get_email_for_seller(seller_name, company_name, seller_email):

    seller_lower = seller_name.lower().strip()
    company_lower = company_name.lower().strip()

    logger.info(f"Email routing started | Seller: '{seller_name}' | Company: '{company_name}'")

    # ================== SPECIAL SELLERS ==================
    special_sellers = [
        "agribegri trade link pvt. ltd.",
        "noble crop science",
        "rain bio tech"
    ]

    if seller_lower in special_sellers:
        logger.info("Matched Special Seller → shipping email")
        return "shipping.agribegri@gmail.com"

    # ================== ATPL MAPPING ==================
    if seller_lower == "atpl":

        if "barrix" in company_lower:
            logger.info("Matched ATPL → Barrix")
            return "info@barrix.in"

        if "neptune" in company_lower:
            logger.info("Matched ATPL → Neptune")
            return "crop10.order@gmail.com"

        if "real trust" in company_lower:
            logger.info("Matched ATPL → Real Trust")
            return "mangesh.ingawale@gmail.com"

    # ================== NORMAL SELLERS ==================
    logger.info("Normal seller → No email will be sent")

    return None

def init_excel():
    if not os.path.exists(EXCEL_FILE):
        logger.info(f"Excel file not found. Creating new report file: {EXCEL_FILE}")
        wb = Workbook()
        ws = wb.active
        ws.title = "Report"
        ws.append(["Timestamp", "Row_Index", "Order_ID", "Status", "Reason"])
        wb.save(EXCEL_FILE)
        logger.info("Excel report file created successfully")
    else:
        logger.info("Excel report file already exists")

init_excel()

def log_to_excel(row_index, order_id, status, reason):
    try:
        # Check if file exists and is not empty or corrupted
        if os.path.exists(EXCEL_FILE) and os.path.getsize(EXCEL_FILE) > 0:
            try:
                wb = load_workbook(EXCEL_FILE)
                ws = wb.active
            except Exception:
                logger.warning("Excel file corrupted or invalid. Recreating...")
                wb = Workbook()
                ws = wb.active
                ws.title = "Report"
                ws.append(["Timestamp", "Row_Index", "Order_ID", "Status", "Reason"])
        else:
            logger.info("Excel file missing or empty. Creating...")
            wb = Workbook()
            ws = wb.active
            ws.title = "Report"
            ws.append(["Timestamp", "Row_Index", "Order_ID", "Status", "Reason"])
        
        ws.append([
            datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            row_index,
            order_id,
            status,
            reason
        ])
        wb.save(EXCEL_FILE)
        logger.info(f"Excel Log Written → Row: {row_index} | Order: {order_id} | Status: {status}")
    except Exception as e:
        logger.exception("Failed to write to Excel report")

def wait_and_rename_pdf(download_dir, new_name, before_files, timeout=60):
    logger.info(f"Waiting for PDF download (Target: {new_name})")

    end_time = time.time() + timeout

    while time.time() < end_time:
        current_files = set(os.listdir(download_dir))
        new_files = current_files - before_files

        # Look for any PDF file (including those with (1), (2), etc.)
        pdf_files = [
            f for f in new_files
            if f.lower().endswith(".pdf") and not f.endswith(".crdownload")
        ]

        if pdf_files:
            # Sort to get the most specific one if multiple (though rare)
            pdf_files.sort(key=len, reverse=True)
            downloaded_pdf = pdf_files[0]

            logger.info(f"Detected downloaded file: {downloaded_pdf}")

            old_path = os.path.join(download_dir, downloaded_pdf)
            safe_name = new_name.replace("/", "_").replace(" ", "_") + ".pdf"
            if not safe_name.lower().endswith(".pdf"):
                safe_name += ".pdf"
            new_path = os.path.join(download_dir, safe_name)

            if os.path.exists(new_path):
                try: os.remove(new_path)
                except: pass

            try:
                os.rename(old_path, new_path)
                logger.info(f"PDF renamed to: {safe_name}")
                return new_path
            except Exception as e:
                logger.warning(f"Initial rename failed: {e}. Retrying in 1s...")
                time.sleep(1)
                os.rename(old_path, new_path)
                return new_path

        time.sleep(1)

    raise Exception("PDF download timeout")
from selenium.webdriver.common.action_chains import ActionChains
import sys
sys.path.append(os.path.dirname(os.path.abspath(__file__)))
import email_helper

# PLACEHOLDERS (User to fill)
SENDER_EMAIL = "dispatch.agribegri@gmail.com"
SENDER_PASSWORD = "ndka txuk ftho pwyn"
ROW_LIMIT = 0  # 0 for ALL rows, or set a number (e.g. 5)
download_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Downloads")
os.makedirs(download_dir, exist_ok=True)

SELLER_PICKUP_MAP = { "abhihsek of be bearings": "OEM India Pvt. Ltd.", "abhijit buch": "Abhijit Buch", "abhilash nair": "Sonikraft", "abhinandan traders": "Abhinandan Traders", "abhishek": "Akshar Farmtech LLP", "abhishek jhawar": "ABHISHEK JHAWAR", "aipm": "AIPM", "airdrops irrigation pvt ltd": "Airdrops Irrigation Pvt Ltd", "aj kisan agrotech": "AJ KISAN AGROTECH", "akash dhanokar": "Apras Polymers & Engineering Co Pvt Ltd", "akhil garg of rs enterprises": "Akhil Garg OF SNM TELECARE", "akshay joshi": "Akshay Joshi", "allbata agriculture biotech pvt ltd": "Albata", "aman agencies": "AMAN AGENCIES", "amit singla": "Amit singla", "amruth organic fertilizers": "Amruth Organic Fertilizers", "amuthalakshmi agro organic,": "Amuthalakshmi Agro Organic,", "anand agro care": "Anand Agro1", "anil packaging": "Anil Packaging", "anjali pitre of urvara": "Urvara Marketing Solutions Pvt Ltd", "annadata organic": "Annadata Organic", "ansh chhabra": "Ansh Chhabra", "ashish tiwari": "RamK Agrotech", "ayushi bansal": "BHUMI AGRO INDUSTRIES", "aziz ahamed": "Sunjaree Fertilizers", "berrysun agro science pvt. ltd.": "Berrysun Agro Science Pvt. Ltd.", "bharat parmar": "Shyam innovations", "bharath r": "Microbi Agrotech Pvt. Ltd.", "bhavani seeds center": "BHAVANI SEEDS CENTER", "bhumi polymers private limited": "BHUMI POLYMERS PRIVATE LIMITED", "biosun agri crop science pvt. ltd": "Biosun Agri Crop Science Pvt. Ltd.", "bm bio energy": "BM BIO ENERGY", "brijesh suresh somani": "Kamal Agrotech", "chetna patidar": "Rashail Agro", "chirag sojitra of bacf": "Chirag Sojitra", "cubic fertichem pvt ltd": "Cubic Fertichem Pvt Ltd", "deepak sharma": "Deepak Sharma", "dhaval gadhiya": "Gujarat Agri-Chem Industries Private Limited", "dr enterprises": "DR ENTERPRISES", "durga beej bhandar": "DURGA BEEJ BHANDAR", "earth innovation": "EARTH INNOVATION", "easykrishi private limited": "Easykrishi Private Limited", "ecotika india": "1 Ecotika India", "essential biosciences": "Essential Biosciences", "excellar production private limited": "EXCELLAR PRODUCTION PRIVATE LIMITED", "exosolar private limited": "Exosolar Private Limited", "farm chem india private limited": "FARM CHEM INDIA PRIVATE LIMITED", "farmson biotech pvt ltd": "FARMSON BIOTECH PVT LTD", "fenton technologies": "Fenton Technologies", "gabani brothers limited": "GABANI BROTHERS LIMITED", "gaiagen technologies private limited": "Gaiagen Technologies Private Limited", "gassin pierre pvt ltd": "Gassin Pierre Pvt Ltd", "gaurav kakkar": "GAURAV KAKKAR", "geolife agritech india pvt. ltd": "Geolife Agritech India Pvt. Ltd", "greatindos": "Greatindos", "green raksha agro seva": "Green Raksha Agro Seva", "green revolution": "Green Revolution", "greeno biotech": "Greeno Biotech", "greenovate agrotech pvt ltd": "1Greenovate Agrotech Pvt Ltd", "greenpeace agro industries": "1Greenpeace Agro", "greenvayu innovations private limited": "GREENVAYU INNOVATIONS PRIVATE LIMITED", "gumtree traps pvt.ltd.": "Gumtree Traps", "gupta agro enterprises": "GUPTA AGRO ENTERPRISES", "gvd electricals": "GVD ELECTRICALS", "harmony ecotech pvt ltd": "Harmony Ecotech Pvt Ltd", "hemendra patidar": "Hemendra Patidar", "hifield-ag chem india pvt ltd.": "Hifield-AG", "hussain lokhandwala": "HM Organics", "infinite biotech co.": "Infinite Biotech Co", "ishan mistry of sk inter": "ISHAN MISTRY Of SK Inter", "ishan sikka": "TUFFPAULIN", "jagannath biotech pvt ltd": "Jagannath Bio Tech Pvt Ltd", "janatha fish meal and oil products": "JANATHA FISH MEAL AND OIL PRODUCTS", "jignesh parmar": "Singhal Industries Pvt. Ltd.", "jinendra magdum": "Jinendra Magdum", "jivit seeds pvt. ltd.": "Jivit Seeds Pvt. Ltd.", "jiya enterprise": "JIYA ENTERPRISE", "kalpana tomar of haridwar.shoppee": "Kalpana Tomar Of Haridwar.Shoppee", "kamna dhawan": "kamna Dhawan", "kanchan kushwaha": "Kanchan Kushwaha", "kandasamy agency": "Kandasamy Agency", "kap associates": "KAP Associates", "kartavya agritech private limited": "KARTAVYA AGRITECH PRIVATE LIMITED", "kashinath c chilshetty": "Kashinath C Chilshetty", "katra fertilizers and chemicals private limited": "Katra", "katyayani organics": "Katyayani Organics", "keshav goyal": "Keshav Goyal", "kisan agrotech": "KISAN AGROTECH", "kodagu agritec private limited": "Kodagu Agritec Private Limited", "krishna seeds farm": "Goldi", "krushna krushi": "KRUSHNA KRUSHI", "kukkar spary centre": "Kukkar Spary Centre", "lta trust": "LTA Trust", "maltose bio innovations private limited": "MALTOSE BIO", "manan patel": "VINSPIRE AGROTECH (I) PVT LTD", "manik": "Manik", "martanbhai patel": "M. N. AGRO INDUSTRIES", "megatex protective fabrics private limited": "MEGATEX PROTECTIVE FABRICS PRIVATE LIMITED", "mettur agro traders": "METTUR AGRO TRADERS", "mipatex india": "Mipatex", "mitrasena": "Mitrasena", "mohan merchandise pvt. ltd": "Mohan Merchandise Private Limited", "monika malakar": "Modish Tractoraurkisan Pvt Ltd", "mrs. mitali hingorani": "Mrs. Mitali Hingorani", "muhammed meerasha": "Muhammed meerasha", "nathsagar bio genetics private limited": "Nathsagar Bio-Genetics Pvt Ltd", "navik organic products": "Navik Organic Products", "navin panchal": "Navin Panchal1", "nihar jain of siesto, chhattisgarh.": "SRT Agro Science Pvt. Ltd.", "nilesh deshpande, maharashtra.": "Urvara Marketing Solutions Pvt Ltd", "nimesh patel": "Nimesh Patel", "ninganagouda biradar": "Ninganagouda Biradar", "octopus crop care": "Octopus Crop Care", "ojas bhattad": "Ojas bhattad", "parth savaliya": "EUREKA SEEDS INDIA PRIVATE LIMITED", "pawan kumar": "Pawan kumar", "pheromone chemicals": "Pheromone Chemicals", "pioneer agro industry": "Pioneer agro Industry", "piyush garg": "Piyush", "piyush kataria": "PIYUSH KATARIA", "prabhat krishi kendra": "Prabhat Krishi Kendra", "pratiksha": "Ajay Bio-Tech(India)Ltd.", "radhe agri center": "Radhe Agri Center", "rahul garg": "Rahul Garg", "rahul jain": "Aquagri Processing Private Ltd", "raj kumar jaiswal": "DHANDA AGRO CHEMICAL INDUSTRIES", "rajib sarkar": "Rajib Sarkar", "rakhi naskar": "Rakhi Naskar", "ramandeep singh": "Ramandeep singh", "ravikumar patel": "Jakson Seeds Private Limited", "ravindra pofalkar": "Ravindra Pofalkar", "risepect enterprise": "Risepect Enterprise", "rk chemicals": "dobariya sunil Of RK chemicals", "romvijay bio tech (p) ltd.": "ROMVIJAY BIOO TECH PVT LTD", "rukcho biotech": "Rukcho Biotech", "sachin bharud": "Global Polyplast", "sagar biotech pvt ltd": "Sagar Biotech Pvt Ltd", "sagar padgilwar": "Pad Corp", "sanchali rawat": "JAIPUR BIO FERTILIZERS", "sandeep kumar raghuwanshi": "Sandeep kumar Raghuwanshi", "sanjay brothers": "Sanjay Brothers", "sanket jagtap": "Sethu Farmer Producer Company Limited", "sarpan seeds": "Sarpan", "sarthak": "SARTHAK", "satish sajjan": "Satish Sajjan", "shanti devi agriculture store matour": "Shanti Devi Agriculture Store Matour", "shine brand seeds": "SHINE BRAND SEEDS", "shiv kailash green energy private limited": "Shiv kailash Green Energy Private Limited", "shree agro agencies": "Shree Agro Agencies", "shree gopalji impex": "SHREE GOPALJI IMPEX", "shree industries": "Shree Industries", "shree sanwariya trading": "SHREE SANWARIYA TRADING", "shriyap mushroom": "1Shriyap Mushroom", "shubham": "Shubham01", "shubham enterprises": "SHUBHAM ENTERPRISES", "sickle innovations private limited": "sickle", "siddharth parikh": "AGREO SOLUTIONS", "siddhi vinayak enterprises": "Siddhi Vinayak Enterprises", "siva reddy": "Hilfiger Chems", "sk agrotech": "SK AGROTECH", "slavs agrotech": "SLAVS AGROTECH1", "sonkul agro industries pvt. ltd.": "Sonkul Agro Industries Pvt. Ltd.", "sri jyotiba fertilizers": "Sri Jyotiba Fertilizers", "sri sai forestry": "New SRI SAI FORESTRY", "sridhar r": "SRIDHAR R", "srinivasan krishnan": "Srinivasan Krishnan", "startek chemicals limited": "1Star Chemicals", "sudhir goyal": "Sudhir Goyal", "sugan chand sunil kumar": "Sugan Chand Sunil Kumar", "suttind seeds": "Suttind Seeds", "thylakoid biotech pvt. ltd.": "Thylakoid Biotech Pvt. Ltd.", "titan agritech limited": "TITAN AGRITECH LIMITED", "torrent crop science": "TORRENT CROP SCIENCE", "tummuru suresh": "tummuru suresh", "turning point natural care": "Turning point natural care", "tushan sharma": "JAI BALAJI MINERALS", "unison engg. industries": "1Unison Engg. Industries", "urja agricare co.": "Urja Agricare Co.", "urja agriculture company": "1Urja Agriculture Company", "utkarsh agrochem pvt. ltd.": "Utkarsh Agrochem", "vahra technology": "VAHRA TECHNOLOGY", "vaibhav singh chhabra of pep solution": "Vaibhav Singh Chhabra of PEP Solution", "vanproz mp": "1Vanproz Agrovate", "vansthali kisan producer company limited": "Vansthali Kisan Producer Company Limited", "vasudha irrigation": "VASUDHA IRRIGATION", "vedant speciality packaging": "Vedant Speciality Packaging", "venus agro chemicals": "VENUS AGRO CHEMICALS", "vetmantra formulations": "Vetmantra Formulations", "vijay makwana": "Kirtiman Agro Genetics Ltd.", "vikalp": "Vikalp Bio", "vikas aggarwal": "Jayesh Enterprises", "villajio technologies private limited": "VILLAJIO TECHNOLOGIES PRIVATE LIMITED", "vinay krishna c of tech source solutions": "Vinay Krishna C", "vishakha adekar": "Vishakha adekar", "vivek malani": "Vivek Malani", "vr international": "VR International", "v-sar enterprise": "V-Sar Enterprise", "yash pardeshi": "Yash Pardeshi", "z. n. global nation": "Z. N. Global Nation", "zeal biologicals": "Zeal Biologicals", "agribegri trade link pvt. ltd.": "Godawon", "noble crop science": "Godawon", "rain bio tech": "Godawon"," rain biotech industries private limited gujarat":"RAIN BIOTECH INDUSTRIES PRIVATEi LIMITED", "atpl": "Real Trust Exim Corporation, India .", }


# ================== BROWSER SETUP ==================

def init_driver():
    global driver, browser_start_time
    logger.info("Chrome browser starting...")

    chrome_options = webdriver.ChromeOptions()

    # Disable Chrome Safe Browsing protections
    chrome_options.add_argument("--safebrowsing-disable-download-protection")
    chrome_options.add_argument("--disable-client-side-phishing-detection")
    chrome_options.add_argument("--disable-features=SafeBrowsing")
    chrome_options.add_argument("--ignore-certificate-errors")
    chrome_options.add_argument("--allow-running-insecure-content")
    chrome_options.add_argument("--disable-web-security")
    chrome_options.add_argument("--disable-site-isolation-trials")
    chrome_options.add_argument("--disable-features=SafeBrowsing,DownloadBubble")
    chrome_options.add_argument("--no-sandbox")
    chrome_options.add_argument("--disable-dev-shm-usage")
    prefs = {
    "download.default_directory": download_dir,
    "download.prompt_for_download": False,
    "download.directory_upgrade": True,
    "plugins.always_open_pdf_externally": True,
    "safebrowsing.enabled": False,
    "safebrowsing.disable_download_protection": True,
    "profile.default_content_settings.popups": 0,
    "profile.content_settings.exceptions.automatic_downloads.*.setting": 1
}

    chrome_options.add_experimental_option("prefs", prefs)

    service = Service(ChromeDriverManager().install())

    driver = webdriver.Chrome(service=service, options=chrome_options)
    driver.execute_cdp_cmd(
    "Page.setDownloadBehavior",
    {"behavior": "allow", "downloadPath": download_dir}
)

    browser_start_time = time.time()
    logger.info("Chrome browser started successfully")

    return driver
# Initial driver setup
driver = init_driver()

def login_to_agribegri(username, password, otp):
    logger.info("Opening Agribegri admin login page")
    driver.get('https://agribegri.com/admin/')
    driver.maximize_window()
    logger.info("Entering username")
    username_field = driver.find_element(By.ID, "username").send_keys(username)
    
    next_button = driver.find_element(By.ID, "btnSubmit")
    next_button.click()
    time.sleep(3)
    logger.info("Entering password")
    password_field = driver.find_element(By.ID, "password").send_keys(password)
    pass_next_button = driver.find_element(By.ID, "btnSubmit")
    pass_next_button.click()
    time.sleep(3)
    logger.info("Entering OTP")
    enter_otp = driver.find_element(By.ID, "otp").send_keys(otp)
    sign_in = driver.find_element(By.ID, "btnSubmit")
    sign_in.click()
    logger.info("Login completed successfully")
    time.sleep(2)
def type_search_number(driver):
    number = "854852AD"
    
    search_box = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.XPATH, "//input[@aria-controls='dyntable']"))
    )
    
    search_box.clear()
    search_box.send_keys(number)

    

def select_shipping_through_courier():
    wait = WebDriverWait(driver, 30)

    # 1 Open Shipping Through dropdown
    shipping_dropdown_btn = wait.until(
        EC.element_to_be_clickable((
            By.XPATH,
            "//button[contains(@class,'multiselect') and .//span[contains(text(),'Select Shipping Through')]]"
        ))
    )
    shipping_dropdown_btn.click()

    logger.info("Shipping Through dropdown opened")

    # 2 Click "Shipping Through Courier" checkbox
    courier_checkbox = wait.until(
        EC.element_to_be_clickable((
            By.XPATH,
            "//label[normalize-space()='Shipping Through Courier']/input"
        ))
    )

    driver.execute_script("arguments[0].click();", courier_checkbox)

    logger.info("Shipping Through Courier selected")

def apply_filter(target_order_id=None):
    logger.info("Opening Manage Orders page")
    driver.get('https://agribegri.com/admin/manage_orders.php')

    # Wait for page to load completely
    time.sleep(3)

    wait = WebDriverWait(driver, 30)

    logger.info("Manage Orders page loaded")

        # ================== SET DATE FILTER (FROM = 30 DAYS AGO, TO = TODAY) ==================

    from datetime import datetime, timedelta

    today_date = datetime.now()
    from_date_value = (today_date - timedelta(days=30)).strftime("%Y-%m-%d")
    to_date_value = today_date.strftime("%Y-%m-%d")
    logger.info(f"Applying Date Filter | From: {from_date_value} | To: {to_date_value}")
    print(f" Setting From Date: {from_date_value}")
    print(f" Setting To Date: {to_date_value}")

    # Set From Date using JS (because input is readonly)
    driver.execute_script("""
        let fromInput = document.getElementById('from_date');
        fromInput.removeAttribute('readonly');
        fromInput.value = arguments[0];
        fromInput.dispatchEvent(new Event('change', { bubbles: true }));
    """, from_date_value)

    time.sleep(0.5)
    
    # Set To Date using JS
    driver.execute_script("""
        let toInput = document.getElementById('to_date');
        toInput.removeAttribute('readonly');
        toInput.value = arguments[0];
        toInput.dispatchEvent(new Event('change', { bubbles: true }));
    """, to_date_value)

    time.sleep(1)

    logger.info("Date filter applied successfully")


    # ================== FILTER BY STATUS: CONFIRM ==================
    logger.info("Selecting Order Status: Confirm")

    # 1. Click Dropdown Button
    status_dropdown_btn = wait.until(EC.element_to_be_clickable(
        (By.XPATH, "//button[@title='Select Order Status']")
    ))
    status_dropdown_btn.click()
    time.sleep(1)

    # 2. Select 'Confirm'
    confirm_option = wait.until(EC.element_to_be_clickable(
        (By.XPATH, "//input[@value='Confirm']")
    ))
    confirm_option.click()
    print(" 'Confirm' status selected")
    time.sleep(0.5)
    logger.info("Order Status 'Confirm' selected")
    # ================== FILTER BY SHIPPING THROUGH COURIER ==================
    select_shipping_through_courier()

    # phone_input = wait.until(
    # EC.presence_of_element_located((By.ID, 'srchby_phone'))
    # )
    # phone_input.clear()
    # phone_input.send_keys('7227029400')


    # 3. Click Search
    search_button = driver.find_element(By.ID, 'srchSubmit')
    search_button.click()

    # wait for table load
    wait.until(
        EC.presence_of_element_located((By.CSS_SELECTOR, "table tbody tr"))
    )
    if target_order_id:
        logger.info(f"Searching for single order: {target_order_id} inside results...")
        search_box = wait.until(EC.presence_of_element_located((By.XPATH, "//div[@id='dyntable_filter']//input")))
        search_box.clear()
        search_box.send_keys(target_order_id)
        time.sleep(5)  # Give time for the table to filter
        try:
            order_found = driver.find_element(By.XPATH, f"//table[@id='dyntable']/tbody/tr[contains(., '{target_order_id}')]")
            logger.info(f"Single order {target_order_id} found.")
        except NoSuchElementException:
            logger.warning(f"Order {target_order_id} NOT found in filtered results. Stopping.")
            return "NOT_FOUND"

    logger.info("Search completed. Orders table loaded successfully")


    time.sleep(1)



def extract_row_data(row, row_index):
    try:
        logger.info(f"Extracting data for Row Index: {row_index}")
        tds = row.find_elements(By.TAG_NAME, "td")
        order_no = tds[1].text.strip() if len(tds) > 1 else "Unknown"
        # If the text spans multiple lines, only take the first line (Order ID)
        order_no = order_no.split("\n")[0].strip()

        remark_element = row.find_element(By.CSS_SELECTOR, "a.get_id")
        remark_id = remark_element.get_attribute("id")
        remark_href = remark_element.get_attribute("href")

        view_element = row.find_element(
            By.XPATH,
            ".//img[@title='View']/ancestor::a"
        )
        view_link = view_element.get_attribute("href")

        truck_element = row.find_element(By.CSS_SELECTOR, "a.get_order_id")
        truck_id = truck_element.get_attribute("id")
        truck_href = truck_element.get_attribute("href")

        row_dict = {
            "row_index": row_index,
            "order_no": order_no,
            "remark_id": remark_id,
            "remark_href": remark_href,
            "view_link": view_link,
            "truck_id": truck_id,
            "truck_href": truck_href
        }

        logger.info(f"Row {row_index} extraction successful")

        for k, v in row_dict.items():
            print(f"{k}: {v}")
        print(" --------------------------\n")

        return row_dict

    except Exception as e:
        logger.exception(f"Failed extracting row data at index {row_index}")
        return None


def click_truck_icons_one_by_one(row_index):
    try:
        logger.info(f"Processing row index: {row_index}")
        wait = WebDriverWait(driver, 30)

        # Re-fetch rows every time to avoid Stale Elements
        rows = wait.until(
            EC.presence_of_all_elements_located(
                (By.XPATH, "//table[@id='dyntable']/tbody/tr")
            )
        )
        logger.info(f"Total filtered rows in table: {len(rows)}")


        # Check bounds
        if row_index >= len(rows):
            logger.info(f"Row index {row_index} out of bounds → End of list")
            return {"status": "END"}

        # Select the target row
        try:
            row = rows[row_index]
            row_text = row.text.lower()

            # ================== EXTRACT ROW DATA FIRST ==================
            row_data = extract_row_data(row, row_index)

            if not row_data:
                logger.error(f"Row data extraction failed at index {row_index}")
                return {
                    "status": "ERROR",
                    "order_id": "Unknown",
                    "reason": "Row data extraction failed"
                }

            order_number = row_data["order_no"]
            logger.info(f"Order detected: {order_number}")
        except IndexError:
            logger.info("IndexError encountered → End of list")
            return {"status": "END"}


        # CHECK REMARK
        if "shipping through transport" in row_text:
            logger.info(f"Order {order_number} skipped → Shipping Through Transport")
            return {"status": "Skipped", "reason": "Shipping Through Transport", "order_id": "N/A"}
        else:
            logger.info(f"Order {order_number} valid for processing")

        # ================== PAYMENT STATUS CHECK ==================

        is_partial_payment = False

        if "partial" in row_text and "payment" in row_text:
            is_partial_payment = True
            logger.info(f"Order {order_number} → Partial payment detected")
        else:
            logger.info(f"Order {order_number} → Full payment detected")



            #   # ================== EARLY WEIGHT / AMOUNT CHECK ==================

            # # ================== EARLY WEIGHT CHECK ==================

            # WEIGHT_LIMIT_GRAM = 13000    # fixed = 13 KG

            # # open order detail page
            # view_icon = row.find_element(By.CSS_SELECTOR, "img[title='View']")
            # driver.execute_script("arguments[0].click();", view_icon)
            # driver.switch_to.window(driver.window_handles[-1])

            # wait = WebDriverWait(driver, 20)
            # print(" Opened order detail page for weight check")

            # # -------- WEIGHT (GRAM per unit) --------
            # weight_elem = wait.until(
            #     EC.presence_of_element_located((By.CSS_SELECTOR, "span.weight_lbl"))
            # )
            # weight_per_unit = float(weight_elem.text.replace(",", "").strip())

            # # -------- QUANTITY --------
            # product_row = wait.until(
            #     EC.presence_of_element_located((
            #         By.XPATH,
            #         "//table[contains(@class,'table')]/tbody/tr"
            #     ))
            # )

            # qty_text = product_row.find_elements(By.TAG_NAME, "td")[7].text
            # quantity = float(qty_text.strip())

            # # -------- CALCULATION --------
            # total_weight = weight_per_unit * quantity

            # print(f" Weight/unit: {weight_per_unit} g")
            # print(f" Quantity: {quantity}")
            # print(f" Total weight: {total_weight} g")

            # # -------- STOP CONDITION (ONLY WEIGHT) --------
            # if total_weight > WEIGHT_LIMIT_GRAM:
            #     print(" STOP ORDER")
            #     print(f" Total weight {total_weight} g exceeds limit {WEIGHT_LIMIT_GRAM} g")

            #     driver.close()
            #     driver.switch_to.window(driver.window_handles[0])
            #     return

            # print(" Weight OK — continuing")

            # driver.close()
            # driver.switch_to.window(driver.window_handles[0])

        # ================== PRE-CHECK WEIGHT & AMOUNT ==================

        view_icon = row.find_element(By.CSS_SELECTOR, "img[title='View']")
        driver.execute_script("arguments[0].click();", view_icon)

        driver.switch_to.window(driver.window_handles[-1])

        wait = WebDriverWait(driver, 20)

        # ---------- READ WEIGHT ----------
        weight_elem = wait.until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "span.weight_lbl"))
        )
        weight_per_unit = float(weight_elem.text.replace(",", "").strip())

        product_row = wait.until(
            EC.presence_of_element_located((
                By.XPATH,
                "//table[contains(@class,'table')]/tbody/tr"
            ))
        )

        quantity = float(product_row.find_elements(By.TAG_NAME, "td")[7].text.strip())

        total_weight = weight_per_unit * quantity

        # ---------- READ AMOUNT ----------
        net_amount_elem = wait.until(
            EC.presence_of_element_located((By.ID, "span_abo_net_amt"))
        )

        net_amount = float(net_amount_elem.text.replace(",", "").strip())

        logger.info(f"Order {order_number} → Pre-check Weight: {total_weight}")
        logger.info(f"Order {order_number} → Pre-check Amount: {net_amount}")

        # ---------- SKIP CONDITION ----------
        if total_weight >= WEIGHT_LIMIT_GRAM or net_amount >= AMOUNT_LIMIT:

            logger.info(
                f"Order {order_number} skipped BEFORE truck → Weight: {total_weight} | Amount: {net_amount}"
            )

            driver.close()
            driver.switch_to.window(driver.window_handles[0])

            return {
                "status": "Skipped",
                "order_id": order_number,
                "reason": "Weight or Amount exceeds allowed limit"
            }

        # Close detail tab if OK
        driver.close()
        driver.switch_to.window(driver.window_handles[0])

        # scroll to row
        driver.execute_script(
            "arguments[0].scrollIntoView({block:'center'});", row
        )
        time.sleep(0.5)

        # click truck icon
        truck_icon = row.find_element(By.CSS_SELECTOR, "a.get_order_id")
        driver.execute_script("arguments[0].click();", truck_icon)

        logger.info(f"Order {order_number} → Truck icon clicked")


        # wait for modal
        wait.until(
            EC.visibility_of_element_located(
            (By.CSS_SELECTOR, 'div[data-remodal-id="surface_modal"]')
        )

        )
        logger.info(f"Order {order_number} → Surface modal opened")


        # click dropdown
        surface_dropdown = wait.until(
            EC.element_to_be_clickable((By.ID, "serviceSurface"))
        )
        
        # 1. Try selecting via Selenium Select class
        logger.info(f"Order {order_number} → Selecting 'Agribegri Surface'")
        try:
            select_surf = Select(surface_dropdown)
            select_surf.select_by_visible_text("Agribegri Surface")
            logger.warning(f"Order {order_number} → Surface selection failed, retrying fallback")
        except Exception as e:
            logger.info(f"Order {order_number} → Surface selected successfully")
        
        # 2. FORCE update value via JS (Backup)
        driver.execute_script("""
            var select = document.getElementById('serviceSurface');
            for(var i=0; i<select.options.length; i++){
                if(select.options[i].text.trim() == 'Agribegri Surface'){
                    select.options[i].selected = true;
                    select.value = select.options[i].value;
                    break;
                }
            }
            select.dispatchEvent(new Event('change', { bubbles: true }));
            select.dispatchEvent(new Event('input', { bubbles: true }));
        """)
        
        time.sleep(1)

        # 3. VERIFY Selection
        selected_option = Select(surface_dropdown).first_selected_option
        print(f" Current Selection: '{selected_option.text}'")

        if "Select One" in selected_option.text:
            print(" Selection failed! Retrying with explicit index...")
            # fallback by index if text fails (assuming it's usually 2nd or 3rd option)
            try:
                Select(surface_dropdown).select_by_index(1) 
                print(" Selected by index 1")
            except:
                pass

        #  force onchange (CRITICAL)
        driver.execute_script("""
            const select = document.getElementById('serviceSurface');
            select.dispatchEvent(new Event('change', { bubbles: true }));
        """)

        print(" Agribegri Surface selected")


        
        #  correct pickup visibility wait
        wait.until(
            lambda d: d.execute_script(
                "return document.getElementById('li_pickup_address').style.display !== 'none';"
            )
        )


        # ================== PICKUP ADDRESS SELECTION START ==================

        # get seller name from table row (7th column)
        seller_raw = row.find_elements(By.TAG_NAME, "td")[7].text
        seller_name = " ".join(seller_raw.split()).lower()


        logger.info(f"Order {order_number} → Seller detected: {seller_name}")

        # decide pickup address (mapping + fallback)
        # decide pickup address (PARTIAL MATCH SAFE VERSION)

        seller_name_clean = seller_name.strip().lower()
        row_text_lower = row.text.lower()
        pickup_text = None

        # ================== CUSTOM PICKUP RULES (USER REQUEST) ==================
        if ("atpl" in seller_name_clean or "aptl" in seller_name_clean) and "barrix agro science" in row_text_lower:
            pickup_text = "Barrix Agro Science Pvt. Ltd."
            logger.info(f"Order {order_number} -> Custom Pickup: Barrix Agro Science Pvt. Ltd.")
        elif "neptune fairdeal" in row_text_lower:
            pickup_text = "Neptune"
            logger.info(f"Order {order_number} -> Custom Pickup: Neptune")
        elif "real trust exim" in row_text_lower:
            pickup_text = "Real Trust Exim Corporation, India"
            logger.info(f"Order {order_number} -> Custom Pickup: Real Trust Exim Corporation, India")

        # Fallback to existing mapping if no custom rule matched
        if not pickup_text:
            for key in sorted(SELLER_PICKUP_MAP.keys(), key=len, reverse=True):
                value = SELLER_PICKUP_MAP[key]
                if key in seller_name_clean:
                    pickup_text = value
                    logger.info(f"Order {order_number} → Pickup mapping matched: {key}")
                    break

        if not pickup_text:
            pickup_text = seller_name_clean
            logger.warning(f"Order {order_number} → No pickup mapping found, using seller name directly")

        print(" Target Pickup Name:", pickup_text)


        # select pickup address from dropdown
        pickup_dropdown = wait.until(
            EC.element_to_be_clickable((By.ID, "servicePickupAddress"))
        )

        options = pickup_dropdown.find_elements(By.TAG_NAME, "option")

        selected_value = None
        matched_text = ""

        # ================== SAFE PICKUP MATCH ==================

        pickup_text_clean = pickup_text.strip().lower()

        selected_value = None
        matched_text = ""

        # ---- EXACT MATCH ONLY ----
        for opt in options:
            opt_text_clean = opt.text.strip().lower()

            if pickup_text_clean == opt_text_clean:
                selected_value = opt.get_attribute("value")
                matched_text = opt.text
                logger.info(
                    f"Order {order_number} → Pickup selected (Exact Match): {matched_text}"
                )
                break

        # ---- OPTIONAL SAFE FALLBACK (Word Boundary Match) ----
        if not selected_value:
            import re
            pattern = r"\b" + re.escape(pickup_text_clean) + r"\b"

            for opt in options:
                opt_text_clean = opt.text.strip().lower()
                if re.search(pattern, opt_text_clean):
                    selected_value = opt.get_attribute("value")
                    matched_text = opt.text
                    logger.info(
                        f"Order {order_number} → Pickup selected (Boundary Match): {matched_text}"
                    )
                    break

        if not selected_value:
            logger.error(
                f"Order {order_number} → Pickup NOT found for: {pickup_text}"
            )
            raise Exception(
                f"Pickup address not found for seller: {seller_name} (Target: {pickup_text})"
            )

        logger.info(
            f"Order {order_number} → Final Pickup Selected: {matched_text}"
        )
        #  SET VALUE DIRECTLY (Using Select Class now for better reliability)
        try:
            # User Select class (Global import)
            
            # 1. Click the dropdown first to ensure focus
            driver.execute_script("arguments[0].click();", pickup_dropdown)
            time.sleep(0.5)

            # 2. Use Select class
            select_elem = Select(pickup_dropdown)
            select_elem.select_by_value(selected_value)
            logger.info(
        f"Order {order_number} → Pickup value set successfully: {matched_text}"
    )
            
            # 3. Explicitly trigger events just in case
            driver.execute_script("""
                const select = arguments[0];
                select.dispatchEvent(new Event('change', { bubbles: true }));
                select.dispatchEvent(new Event('input', { bubbles: true }));
                select.dispatchEvent(new Event('blur', { bubbles: true }));
            """, pickup_dropdown)

        except Exception as e:
            logger.warning(
        f"Order {order_number} → Select class failed, using JS fallback"
    )
            driver.execute_script("""
                const select = document.getElementById('servicePickupAddress');
                select.value = arguments[0];
                select.dispatchEvent(new Event('change', { bubbles: true }));
                select.dispatchEvent(new Event('input', { bubbles: true }));
                select.dispatchEvent(new Event('blur', { bubbles: true }));
            """, selected_value)

        logger.info(f"Order {order_number}  Pickup address selected successfully")
        time.sleep(1)

        # REMOVED "RE-SELECT SURFACE" BLOCK causing reset issues


        # ================== SUBMIT SURFACE FORM ==================

        submit_btn = wait.until(
            EC.element_to_be_clickable((By.NAME, "submit_surface"))
        )

        driver.execute_script("arguments[0].click();", submit_btn)

        logger.info(f"Order {order_number}  Submit button clicked successfully")

        # ================== SUCCESS POPUP HANDLING ==================

        # wait for success popup
        ok_button = wait.until(
            EC.element_to_be_clickable((By.ID, "popup_ok"))
        )

        driver.execute_script("arguments[0].click();", ok_button)

        logger.info(f"Order {order_number} → Surface shipment created successfully")

        # optional: small wait for backend processing
        time.sleep(2)

        # ================== OPEN ORDER DETAIL (ONCE) ==================

        view_icon = row.find_element(By.CSS_SELECTOR, "img[title='View']")
        driver.execute_script("arguments[0].click();", view_icon)
        driver.switch_to.window(driver.window_handles[-1])
        logger.info(f"Order {order_number} → Order detail page opened")
         
        # ================== EXTRACT SELLER + COMPANY ==================

        wait = WebDriverWait(driver, 20)

        # --- From Product Table ---
        product_row = wait.until(
            EC.presence_of_element_located((
                By.XPATH,
                "//table[@id='dyntable']/tbody/tr"
            ))
        )

        cells = product_row.find_elements(By.TAG_NAME, "td")

        company_from_table = cells[2].text.strip()
        seller_from_table = cells[3].text.strip()

        seller_name_view = seller_from_table
        company_name_view = company_from_table
        logger.info(
            f"Order {order_number} → Seller: {seller_name_view} | Company: {company_name_view}"
        )

        # ================== EXTRACT SELLER EMAIL ==================
        try:
            seller_email_elem = driver.find_element(
                By.XPATH,
                "//h5[text()='Seller Details']/following-sibling::table//td[text()='Email :']/following-sibling::td"
            )
            seller_email = seller_email_elem.text.strip()
            logger.info(f"Order {order_number} → Seller Email found: {seller_email}")
        except:
            seller_email = ""
            logger.warning(f"Order {order_number} → Seller Email not found")


        # print("\n -------- EMAIL ROUTING LOG --------")
        # print(f"Seller Name   : {seller_name_view}")
        # print(f"Company Name  : {company_name_view}")
        # print(f"Seller Email  : {seller_email}")
        # print("--------------------------------------")

        # email_to_send = None

        # #  Special Seller Rule
        # if seller_lower in special_sellers:
        #     email_to_send = "shipping.agribegri@gmail.com"
        #     print("Rule Applied  : Special Seller")

        # #  ATPL Company Mapping
        # elif seller_lower == "atpl":
        #     for key, value in company_email_map.items():
        #         if key in company_lower:
        #             email_to_send = value
        #             print("Rule Applied  : ATPL Company Mapping")
        #             break

        # #  Company-Based Email Routing (NEW)
        # elif company_lower in company_email_map:
        #     email_to_send = company_email_map[company_lower]
        #     print("Rule Applied  : Direct Company Mapping")

        # #  Default Seller Email
        # elif seller_email:
        #     email_to_send = seller_email
        #     print("Rule Applied  : Default Seller Email")

        # else:
        #     print("Rule Applied  : No Email (Upload Only)")

        # print(f" FINAL EMAIL WILL GO TO: {email_to_send}")
        # print(" --------------------------------------\n")


        # # -------- EXECUTE --------
        # if email_to_send:
        #     email_helper.send_email_with_attachment(
        #         SENDER_EMAIL,
        #         SENDER_PASSWORD,
        #         email_to_send,
        #         f"Shipping Label for {order_number}",
        #         f"Attached shipping label for {order_number}",
        #         renamed_pdf_path
        #     )
        #     print(" Email sent successfully")
        # else:
        #     print(" PDF uploaded only — No email sent")



        # --- From Seller Details Section ---
        # try:
        #     seller_name_view = driver.find_element(
        #         By.XPATH,
        #         "//h5[text()='Seller Details']/following-sibling::table//td[contains(text(),'Seller Name')]/following-sibling::td"
        #     ).text.strip()
        # except:
        #     seller_name_view = ""

        # try:
        #     company_name_view = driver.find_element(
        #         By.XPATH,
        #         "//h5[text()='Seller Details']/following-sibling::table//td[contains(text(),'Company Name')]/following-sibling::td"
        #     ).text.strip()
        # except:
        #     company_name_view = ""

        print("\n -------- SELLER INFO --------")
        print("Company (Table):", company_from_table)
        print("Seller (Table):", seller_from_table)
        print("Seller Name:", seller_name_view)
        print("Company Name:", company_name_view)
        logger.info(
            f"Order {order_number} → Seller: {seller_name_view} | Company: {company_name_view}"
        )

        print(" -----------------------------\n")
        # ================== SPECIAL SELLER FLAG ==================

        special_sellers = [
            "agribegri trade link pvt. ltd.",
            "noble crop science",
            "rain bio tech"
        ]

        seller_lower = seller_name_view.lower().strip()
        is_special_seller = seller_lower in special_sellers

        if is_special_seller:
            logger.info(f"Order {order_number} → Special seller detected")
        else:
            logger.info(f"Order {order_number} → Normal seller")

        agribegri_order_tab = driver.current_window_handle
        print(" Order detail tab opened & stored")

        # ================== READ DIMENSION (CORRECT PLACE) ==================

        import re   # make sure this is at TOP of file ideally

        dimension_text = WebDriverWait(driver, 20).until(
            EC.visibility_of_element_located((
                By.XPATH, "//span[contains(@class,'dimension_lbl')]"
            ))
        ).text.strip()

        values = re.findall(r"\d+(?:\.\d+)?", dimension_text)

        if len(values) != 3:
            raise Exception(f" Invalid dimension format: {dimension_text}")

        length, breadth, height = values

        logger.info(
            f"Order {order_number} → Dimensions parsed: L={length}, B={breadth}, H={height}"
        )



        # ================== POST-SURFACE WEIGHT CHECK ==================

        # if not is_partial_payment:
        #     print("⚖ Performing weight check (non-partial payment)")

        #     WEIGHT_LIMIT_GRAM = 13000

        #     split_count = math.ceil(total_weight / WEIGHT_LIMIT_GRAM)
        #     per_split_weight = total_weight / split_count

        #     print(f" Total shipments required: {split_count}")
        #     print(f" Per shipment weight: {per_split_weight:.2f} g")

        #     # open order detail page
        #     view_icon = row.find_element(By.CSS_SELECTOR, "img[title='View']")
        #     driver.execute_script("arguments[0].click();", view_icon)
        #     driver.switch_to.window(driver.window_handles[-1])

        #     wait = WebDriverWait(driver, 20)

        #     weight_elem = wait.until(
        #         EC.presence_of_element_located((By.CSS_SELECTOR, "span.weight_lbl"))
        #     )
        #     weight_per_unit = float(weight_elem.text.replace(",", "").strip())

        #     product_row = wait.until(
        #         EC.presence_of_element_located((
        #             By.XPATH,
        #             "//table[contains(@class,'table')]/tbody/tr"
        #         ))
        #     )

        #     quantity = float(product_row.find_elements(By.TAG_NAME, "td")[7].text.strip())

        #     total_weight = weight_per_unit * quantity

        #     print(f" Total weight: {total_weight} g")

        #     print(" Weight within limit")

        #     driver.close()
        #     driver.switch_to.window(driver.window_handles[0])
        # else:
        #     print(" Partial payment order — skipping weight check")



        # ================== CLICK VIEW ICON ==================

        # view_icon = row.find_element(By.CSS_SELECTOR, "img[title='View']")
        # driver.execute_script("arguments[0].click();", view_icon)

        # print(" View icon clicked (new tab opened)")

        # ================== SWITCH TO VIEW TAB ==================

        driver.switch_to.window(driver.window_handles[-1])
        logger.info("Switched to order detail tab")

        agribegri_order_tab = driver.current_window_handle
        logger.info("Agribegri order tab stored")


        # ================== COPY ORDER NUMBER (FIXED) ==================
        order_number_elem = WebDriverWait(driver, 20).until(
            EC.presence_of_element_located(
                (By.XPATH, "//span[contains(text(),'Order Number')]/strong")
            )
        )



        order_number = order_number_elem.text.strip()
        logger.info(f"Order Number copied: {order_number}")

        # ================== READ NET AMOUNT FROM AGRIBEGRI ==================

        driver.switch_to.window(agribegri_order_tab)

        net_amount_elem = WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.ID, "span_abo_net_amt"))
        )

        net_amount = float(net_amount_elem.text.replace(",", "").strip())

        logger.info(f"Net Amount from Agribegri: {net_amount}")

        # calculate per-split amount
        # no split anymore
        per_split_amount = net_amount
        logger.info(f"Order {order_number} → Shipment amount: {per_split_amount}")

        # switch back to Delhivery tab
        driver.switch_to.window(driver.window_handles[-1])


        agribegri_edit_url = driver.current_url
        logger.info(f"Stored Agribegri edit URL: {agribegri_edit_url}")
        

        # ================== EXTRACT SHIPPING ADDRESS DETAILS ==================

        logger.info(f"Order {order_number} → Extracting shipping address details...")

        # Name
        customer_name = wait.until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "span.name_lbl"))
        ).text.strip()

        # Email
        try:
            # Email is in a <td> element within the shipping address table
            email_elem = driver.find_element(
                By.XPATH,
                "//td[text()='Email :']/following-sibling::td"
            )
            customer_email = email_elem.text.strip()
        except:
            customer_email = ""  # Optional field

        # Address
        customer_address = driver.find_element(By.CSS_SELECTOR, "span.address_lbl").text.strip()

        # Landmark (optional - might be empty)
        try:
            customer_landmark = driver.find_element(By.CSS_SELECTOR, "span.landmark_lbl").text.strip()
        except:
            customer_landmark = ""

        # Pincode
        customer_pincode = driver.find_element(By.CSS_SELECTOR, "span.zipcode_lbl").text.strip()

        # Phone
        customer_phone = driver.find_element(By.CSS_SELECTOR, "span.phone_lbl").text.strip()

        # City
        customer_city = driver.find_element(By.CSS_SELECTOR, "span.city_lbl").text.strip()

        # Taluka (optional)
        try:
            customer_taluka = driver.find_element(By.CSS_SELECTOR, "span.taluka_lbl").text.strip()
        except:
            customer_taluka = ""

        # District (optional)
        try:
            customer_district = driver.find_element(By.CSS_SELECTOR, "span.district_lbl").text.strip()
        except:
            customer_district = ""

        # State
        customer_state = driver.find_element(By.CSS_SELECTOR, "span.state_lbl").text.strip()

        print(f" Customer Name: {customer_name}")
        print(f" Customer Email: {customer_email}")
        print(f" Customer Address: {customer_address}")
        print(f" Customer Pincode: {customer_pincode}")
        print(f" Customer Phone: {customer_phone}")
        print(f" Customer City: {customer_city}")
        print(f" Customer Taluka: {customer_taluka}")
        print(f" Customer District: {customer_district}")
        print(f" Customer State: {customer_state}")
        logger.info(f"Order {order_number} → Customer Details: Name={customer_name}, Email={customer_email}, Address={customer_address}, Pincode={customer_pincode}, Phone={customer_phone}, City={customer_city}, Taluka={customer_taluka}, District={customer_district}, State={customer_state}")

        # ================== EXTRACT PRODUCT DESCRIPTION ==================

        logger.info(f"Order {order_number} → Extracting product description...")

        # Product description from the table (column 2 - index 1)
        product_row = wait.until(
            EC.presence_of_element_located((
                By.XPATH,
                "//table[contains(@class,'table')]/tbody/tr"
            ))
        )

        # Product name is in the 2nd column (index 1)
        product_description = product_row.find_elements(By.TAG_NAME, "td")[1].text.strip()

        logger.info(f"Order {order_number} → Product Description: {product_description}")

        # # ================== EXTRACT SELLER EMAIL (New) ==================
        # print(" Extracting Seller Name & Company Name...")

        # try:
        #     seller_name_elem = driver.find_element(
        #         By.XPATH,
        #         "//h5[text()='Seller Details']/following-sibling::table//td[contains(text(),'Seller Name')]/following-sibling::td"
        #     )
        #     seller_name_view = seller_name_elem.text.strip()
        # except:
        #     seller_name_view = ""

        # try:
        #     company_name_elem = driver.find_element(
        #         By.XPATH,
        #         "//h5[text()='Seller Details']/following-sibling::table//td[contains(text(),'Company Name')]/following-sibling::td"
        #     )
        #     company_name_view = company_name_elem.text.strip()
        # except:
        #     company_name_view = ""

        # print(f" Seller Name: {seller_name_view}")
        # print(f" Company Name: {company_name_view}")

        # try:
        #     seller_email_elem = driver.find_element(
        #         By.XPATH,
        #         "//h5[text()='Seller Details']/following-sibling::table//td[text()='Email :']/following-sibling::td"
        #     )
        #     seller_email = seller_email_elem.text.strip()
        #     print(f" Seller Email found: {seller_email}")
        # except:
        #     seller_email = ""
        #     print(" Seller Email not found")

        # ================== OPEN DELHIVERY ==================
        time.sleep(15)
        driver.execute_script("window.open('https://one.delhivery.com/v2/login','_blank')")
        driver.switch_to.window(driver.window_handles[-1])
        logger.info("Delhivery tab opened")

        # ================== DELHIVERY LOGIN (WITH AUTO-SKIP) ==================

        logger.info("Checking Delhivery login status...")

        time.sleep(3)  # let page render

        login_fields = driver.find_elements(By.NAME, "email")

        if login_fields:
            logger.info(" Login page detected — proceeding with login")

            wait = WebDriverWait(driver, 30)

            # Enter Email
            email_input = wait.until(
                EC.visibility_of_element_located((By.NAME, "email"))
            )
            email_input.clear()
            email_input.send_keys("complain@agribegri.com")
            logger.info(" Email entered")

            # Click Continue
            continue_btn = wait.until(
                EC.element_to_be_clickable((By.XPATH, "//button[contains(text(),'Continue')]"))
            )
            driver.execute_script("arguments[0].click();", continue_btn)

            # Enter Password
            password_input = wait.until(
                EC.visibility_of_element_located((By.XPATH, "/html/body/div/div/div[2]/div[2]/div[2]/form/div[2]/div/div/section/input"))
            )
            password_input.send_keys("Agribegri@CL#26")
            logger.info(" Password entered")

            # Click Login
            login_btn = wait.until(
                EC.element_to_be_clickable((By.XPATH, "/html/body/div/div/div[2]/div[2]/div[2]/form/button"))
            )
            driver.execute_script("arguments[0].click();", login_btn)


            logger.info(" Login completed")

            # wait for dashboard
            WebDriverWait(driver, 60).until(
                EC.presence_of_element_located((
                    By.XPATH,
                    "//button[contains(@class,'ap-menu-trigger-root')]"
                ))
            )

        else:
            logger.info(" Already logged in — skipping login")

        # ================== SELECT AGRIBEGRI SURFACE (TOP-RIGHT DROPDOWN) ==================

        # Wait until dashboard is fully loaded
        WebDriverWait(driver, 60).until(
            EC.presence_of_element_located((
                By.XPATH,
                "//button[contains(@class,'ap-menu-trigger-root')]"
            ))
        )

        logger.info("Delhivery dashboard fully loaded")


    # ================== CLICK DOMESTIC / AGRIBEGRI SURFACE DROPDOWN ==================

        domestic_dropdown = WebDriverWait(driver, 30).until(
            EC.element_to_be_clickable((
                By.XPATH,
                "//button[contains(@class,'ap-menu-trigger-root')]"
                "[.//i[contains(@class,'fa-truck')]]"
            ))
        )

        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", domestic_dropdown)
        time.sleep(0.5)

        driver.execute_script("arguments[0].click();", domestic_dropdown)
        logger.info(" Domestic dropdown clicked")



            # ================== SELECT AGRIBEGRI SURFACE ==================

        agribegri_surface = WebDriverWait(driver, 30).until(
            EC.element_to_be_clickable((
                By.XPATH,
                "//button[contains(@class,'ap-menu-item')]"
                "[.//div[text()='AGRIBEGRI SURFACE']]"
            ))
        )

        driver.execute_script("arguments[0].click();", agribegri_surface)
        logger.info("AGRIBEGRI SURFACE selected")

        time.sleep(1.5)
        
        # ================== WEIGHT > 13 KG SPECIAL FLOW ==================

        # ================== CLICK AWB DROPDOWN ==================

        awb_dropdown = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((
                By.XPATH,
                "//button[.//span[text()='AWB']]"
            ))
        )

        driver.execute_script("arguments[0].click();", awb_dropdown)
        logger.info(" AWB dropdown opened")

        order_id_option = WebDriverWait(driver, 20).until(
        EC.element_to_be_clickable((
            By.XPATH,
            "//button[.//span[text()='Order ID']]"
        ))
        )

        driver.execute_script("arguments[0].click();", order_id_option)
        logger.info(" Order ID selected")

        order_search_input = WebDriverWait(driver, 20).until(
        EC.presence_of_element_located((
            By.XPATH,
            "//input[contains(@placeholder,'ORDER ID')]"
        ))
    )

        order_search_input.clear()
        order_search_input.send_keys(order_number)

        logger.info(f" Order ID pasted: {order_number}")

        # ================== CLICK ORDER SEARCH RESULT (WITH RETRY) ==================
        
        time.sleep(3) # Wait for search results to appear

        search_success = False
        for attempt in range(3):
            try:
                logger.info(f" Waiting for search result (Attempt {attempt+1})...")
                search_result = WebDriverWait(driver, 10).until(
                    EC.element_to_be_clickable((
                        By.XPATH,
                        "//div[contains(@class,'ucp__global-search__results')]"
                        "//div[contains(@class,'cursor-pointer')]"
                        f"[.//span[contains(text(), '{order_number}')]]"
                    ))
                )
                driver.execute_script("arguments[0].click();", search_result)
                logger.info(" Order search result clicked")
                search_success = True
                break
            except Exception as e:
                logger.error(f" Attempt {attempt+1} search failed: {e}")
                if attempt < 2:
                    # Retry entering text
                    order_search_input.clear()
                    time.sleep(0.5)
                    order_search_input.send_keys(order_number)
                    time.sleep(2)
        
        if not search_success:
            logger.error(f" Error: Could not find order {order_number} in Delhivery search results after 3 attempts.")
            return {"status": "Error", "reason": "Order not found in Delhivery", "order_id": order_number}

        time.sleep(2)




        # ================== PARTIAL PAYMENT HANDLING IN DELHIVERY ==================

        if is_partial_payment:
            logger.info(" Partial payment flow activated in Delhivery")

            # ---------- SWITCH TO AGRIBEGRI TAB ----------
            driver.switch_to.window(agribegri_order_tab)
            driver.refresh()
            time.sleep(3)

            logger.info(" Forced return to Agribegri edit order page")


            # ---------- COPY COD PAYMENT AMOUNT ----------
            cod_amount_elem = WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((
                    By.XPATH,
                    "//p[contains(@class,'p_disp_advance_amt') and contains(.,'COD Payment Amount')]"
                ))
            )

            import re
            cod_text = cod_amount_elem.text
            cod_amount = re.search(r"([\d]+\.\d+)", cod_text).group(1)

            logger.info(f" Correct COD Amount copied: {cod_amount}")

            # ---------- SWITCH BACK TO DELHIVERY ----------
            driver.switch_to.window(driver.window_handles[-1])

            # ---------- CLICK  EDIT ICON (CORRECT ELEMENT) ----------
            edit_payment_btn = WebDriverWait(driver, 20).until(
                EC.element_to_be_clickable((
                    By.XPATH,
                    "//button[@data-action='edit-payment-mode']"
                ))
            )

            driver.execute_script("arguments[0].click();", edit_payment_btn)
            logger.info(" Payment Details edit icon clicked (correct one)")

            time.sleep(0.8)

            # ---------- FORCE-OPEN PAYMENT MODE DROPDOWN ----------
            payment_dropdown = WebDriverWait(driver, 20).until(
                EC.element_to_be_clickable((
                    By.XPATH,
                    "//button[contains(@class,'ap-menu-trigger-root')]"
                ))
            )

            driver.execute_script("arguments[0].click();", payment_dropdown)
            time.sleep(0.5)

            cod_option = WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((
                    By.XPATH,
                    "//div[contains(@class,'ucp__order-creation__select-payment__dropdown-item--cod')]"
                ))
            )

            driver.execute_script("arguments[0].click();", cod_option)
            logger.info(" Cash On Delivery selected")
            time.sleep(0.5)


            # ---------- ENTER COLLECTABLE AMOUNT ----------
            collectable_input = WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "input.input[type='number']"))
            )
            collectable_input.clear()
            collectable_input.send_keys(cod_amount)
            logger.info(f" Collectable amount entered: {cod_amount}")

            # ---------- CLICK  SAVE ----------
            save_btn = WebDriverWait(driver, 20).until(
                EC.element_to_be_clickable((By.XPATH, "//button[@title='Save']"))
            )
            driver.execute_script("arguments[0].click();", save_btn)
            logger.info("COD payment saved")

            time.sleep(2)

        # ================== CLICK PRINT SHIPPING LABEL ==================

        print_label_btn = WebDriverWait(driver, 30).until(
            EC.element_to_be_clickable((
                By.XPATH,
                "//button[normalize-space()='Print Shipping Label']"
            ))
        )

        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", print_label_btn)
        time.sleep(0.5)

        # Capture files BEFORE clicking download
        # Clear any existing Shipping_Label.pdf to avoid conflict
        for f in ["Shipping_Label.pdf", "Shipping_Label (1).pdf", "Shipping_Label (2).pdf"]:
            p = os.path.join(download_dir, f)
            if os.path.exists(p):
                try: os.remove(p)
                except: pass

        before_files = set(os.listdir(download_dir))
        logger.info(f"Files before label download: {len(before_files)}")

        driver.execute_script("arguments[0].click();", print_label_btn)
        time.sleep(2)
        logger.info(" Print Shipping Label button clicked")

        # call the global function
        try:
            renamed_pdf_path = wait_and_rename_pdf(download_dir, order_number, before_files)

        except Exception as e:
            logger.error(f"Label download failed for order {order_number}")

            log_to_excel(
                row_index,
                order_number,
                "MANUAL REQUIRED",
                "Label PDF not downloaded"
            )

            return {
                "status": "Error",
                "order_id": order_number,
                "reason": "Label PDF not downloaded"
            }

        # ================== VERIFY LABEL PDF EXISTS ==================

        if not os.path.exists(renamed_pdf_path):
            logger.error(f"Label PDF missing for order {order_number} → skipping order")

            log_to_excel(
                row_index,
                order_number,
                "MANUAL REQUIRED",
                "Label PDF missing"
            )

            return {
                "status": "Error",
                "order_id": order_number,
                "reason": "Label PDF missing"
            }
        seller_lower = seller_name_view.lower().strip()
        company_lower = company_name_view.lower().strip()

        special_sellers = [
            "agribegri trade link pvt. ltd.",
            "noble crop science",
            "rain bio tech"
        ]

        company_email_map = {
            "barrix agro science pvt. ltd.": "info@barrix.in",
            "neptune fairdeal products pvt. ltd": "crop10.order@gmail.com",
            "real trust exim corporation": "mangesh.ingawale@gmail.com"
        }

        # # ================== NORMAL FLOW EMAIL (ONLY SPECIAL SELLERS) ==================

        # if is_special_seller:

        #     email_to_send = get_email_for_seller(
        #         seller_name_view,
        #         company_name_view,
        #         seller_email
        #     )

        #     if email_to_send:
        #         email_helper.send_email_with_attachment(
        #             SENDER_EMAIL,
        #             SENDER_PASSWORD,
        #             email_to_send,
        #             f"Shipping Label for {order_number}",
        #             f"Attached shipping label for {order_number}",
        #             renamed_pdf_path
        #         )
        #         print(" Special Seller - Email Sent")
        #     else:
        #         print(" Special Seller but no email rule matched")

        # else:
        #     print(" Not a Special Seller - No Email Sent")


        # # -------- NO SPLIT EMAIL LOGIC --------
        # if seller_lower in special_sellers:
        #     send_email_flag = True
        #     email_to_send = "shipping.agribegri@gmail.com"

        # elif seller_lower == "atpl":
        #     for key, value in company_email_map.items():
        #         if key in company_lower:
        #             send_email_flag = True
        #             email_to_send = value
        #             break

        # # -------- EXECUTE --------
        # if send_email_flag:
        #     print(f" Sending NO-SPLIT email to {email_to_send}")

        #     email_helper.send_email_with_attachment(
        #         SENDER_EMAIL,
        #         SENDER_PASSWORD,
        #         email_to_send,
        #         f"Shipping Label for {order_number}",
        #         f"Attached shipping label for {order_number}",
        #         renamed_pdf_path
        #     )
        # else:
        #     print(" Other seller — will upload PDF only (no email)")


        # ================== SWITCH BACK TO AGRIBEGRI ADMIN TAB ==================
        driver.switch_to.window(agribegri_order_tab)
        driver.refresh()
        time.sleep(2)

        logger.info(" Switched back to Agribegri edit order page (final)")

        # ================== SET ORDER STATUS TO PACKED ==================

        # wait for Order Status dropdown
        order_status_select = WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.ID, "abo_status"))
        )

        # use Select for <select> tag
        select = Select(order_status_select)

        # select "Packed"
        select.select_by_visible_text("Packed")

        logger.info(" Order status set to Packed")

        time.sleep(1)

        # ================== SET PACKED REASON TO CL SURFACE ==================

        packed_reason_select = WebDriverWait(driver, 20).until(
            EC.presence_of_element_located((By.ID, "abo_packed_reason"))
        )

        packed_reason = Select(packed_reason_select)

        packed_reason.select_by_visible_text("CL Surface")

        logger.info(" Packed reason set to CL Surface")

        time.sleep(1)


        # ================== SET PAYMENT STATUS IF PARTIAL PAYMENT ==================

        if is_partial_payment:
            logger.info(" Setting payment status: Bank Partial Payment Recieved")

            payment_status_select = WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((By.ID, "abo_payment_type"))
            )

            payment_select = Select(payment_status_select)

            payment_select.select_by_visible_text("Bank Partial Payment Recieved")

            logger.info(" Payment status set to Bank Partial Payment Recieved")

            time.sleep(1)
        else:
            logger.info(" Skipping payment status change (not partial payment)")

        time.sleep(1)
        # ================== CLICK SUBMIT BUTTON ==================

        submit_btn = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((
                By.NAME, "update_order_status"
            ))
        )

        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", submit_btn)
        time.sleep(0.5)

        driver.execute_script("arguments[0].click();", submit_btn)
        logger.info(" Submit button clicked successfully")

        time.sleep(2)
        # ================== CLICK OK ON SUCCESS POPUP ==================

        ok_btn = WebDriverWait(driver, 20).until(
            EC.element_to_be_clickable((By.ID, "popup_ok"))
        )

        driver.execute_script("arguments[0].click();", ok_btn)
        logger.info(" Success popup OK clicked")

        time.sleep(1.5)

        # ================== UPLOAD LABEL PDF ==================
        # label_pdf_path = os.path.join(download_dir, f"{order_number}.pdf")
        # ================== UPLOAD LABEL PDF ==================
        label_pdf_path = renamed_pdf_path

        seller_lower = seller_name_view.lower().strip()

        if is_special_seller or seller_lower == "atpl":
            logger.info(f" Special seller '{seller_name_view}' detected → Skipping label upload to Agribegri.")
        else:
            logger.info(f" Normal seller detected → Proceeding with label upload.")

            label_input = WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((By.NAME, "label_file"))
            )

            if not os.path.exists(label_pdf_path):
                raise Exception(f" Label PDF not found for upload: {label_pdf_path}")

            label_input.send_keys(label_pdf_path)
            logger.info(f" Label PDF uploaded: {label_pdf_path}")

            time.sleep(1)
            
            submit_manifest_btn = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.NAME, "update_manifest_file"))
            )
            driver.execute_script("arguments[0].click();", submit_manifest_btn)
            logger.info(" Manifest upload submit clicked")

            ok_popup_btn = WebDriverWait(driver, 10).until(
                EC.element_to_be_clickable((By.ID, "popup_ok"))
            )
            driver.execute_script("arguments[0].click();", ok_popup_btn)
            logger.info(" OK button clicked on upload popup")
        time.sleep(1)

        # ================== COMBINED EMAIL (LABEL + INVOICE) ==================
        logger.info("Generating Invoice and sending combined email...")

        # ================== EMAIL ROUTING ==================

        email_to_send = get_email_for_seller(
            seller_name_view,
            company_name_view,
            seller_email
        )

        if email_to_send:

            logger.info("Email required → Generating invoice")

            # Clear any existing downloads.htm or Invoice.pdf to avoid conflict
            for f in ["downloads.htm", "Invoice.pdf", "invoice.pdf", "Shipping_Label.pdf"]:
                p = os.path.join(download_dir, f)
                if os.path.exists(p):
                    try: os.remove(p)
                    except: pass

            generate_btn = WebDriverWait(driver, 20).until(
                EC.element_to_be_clickable((By.ID, "btnGenerateInvoice"))
            )

            before_files = set(os.listdir(download_dir))

            driver.execute_script("arguments[0].click();", generate_btn)
            time.sleep(4)

            driver.switch_to.window(driver.window_handles[-1])

            try:
                invoice_pdf_path = wait_and_rename_pdf(
                    download_dir,
                    f"{order_number}_invoice",
                    before_files,
                    timeout=60
                )
            except Exception:
                logger.error(f"Invoice download failed for order {order_number}")

                log_to_excel(
                    row_index,
                    order_number,
                    "MANUAL REQUIRED",
                    "Invoice PDF not downloaded"
                )

                return {
                    "status": "Error",
                    "order_id": order_number,
                    "reason": "Invoice PDF not downloaded"
                }

            driver.close()
            driver.switch_to.window(agribegri_order_tab)

            attachments = [
                label_pdf_path,
                invoice_pdf_path
            ]

            email_helper.send_email_multiple_attachments(
                SENDER_EMAIL,
                SENDER_PASSWORD,
                email_to_send,
                f"Shipping Documents for {order_number}",
                f"Attached shipping label and invoice for {order_number}",
                attachments
            )

            logger.info(f"Combined Email Sent to {email_to_send}")

        else:
            logger.info("No email rule → Skipping invoice download")
        return {
            "status": "Processed",
            "order_id": order_number,
            "seller": seller_name_view,
            "amount": net_amount,
            "customer": customer_name
        }


        # # ================== SPECIAL SELLER NO-SPLIT EMAIL ==================
        # if is_special_seller:

        #     print(" Generating Invoice for Special Seller (No Split)...")

        #     # Click Generate Invoice
        #     generate_btn = WebDriverWait(driver, 20).until(
        #         EC.element_to_be_clickable((By.ID, "btnGenerateInvoice"))
        #     )
        #     driver.execute_script("arguments[0].click();", generate_btn)
        #     time.sleep(4)

        #     # Switch to invoice tab
        #     driver.switch_to.window(driver.window_handles[-1])

        #     invoice_pdf_path = wait_and_rename_pdf(
        #         download_dir,
        #         f"{order_number}_invoice",
        #         timeout=60
        #     )

        #     driver.close()
        #     driver.switch_to.window(agribegri_order_tab)

        #     email_to_send = get_email_for_seller(
        #         seller_name_view,
        #         company_name_view,
        #         seller_email
        #     )

        #     attachments = [
        #         label_pdf_path,
        #         invoice_pdf_path
        #     ]

        #     email_helper.send_email_multiple_attachments(
        #         SENDER_EMAIL,
        #         SENDER_PASSWORD,
        #         email_to_send,
        #         f"Shipping Documents for {order_number}",
        #         f"Attached shipping label and invoice for {order_number}",
        #         attachments
        #     )

        #     print(" Special Seller No-Split Combined Email Sent")


    
    except Exception as e:
        logger.error(f"Error inside click_truck_icons_one_by_one: {e}")
        import traceback
        traceback.print_exc()
        return {
            "status": "Error",
            "order_id": order_number if 'order_number' in locals() else "Unknown",
            "reason": str(e)
        }



def main_workflow(username, password, otp, target_order_id=None):
    global driver
    
    total_processed_global = 0
    orders_since_restart = 0

    while True:  # Outer loop for restarting from Row 1
        login_to_agribegri(username, password, otp)
        filter_result = apply_filter(target_order_id)
        if filter_result == "NOT_FOUND":
            return

        logger.info("Filtering done. Checking total orders...")

        # ================== GET TOTAL COUNT ==================
        try:
            total_count_elem = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, "//*[@id='seller_frm']/div/span[1]"))
            )
            total_text = total_count_elem.text.strip()
            # Extract number from string like "Total No of Orders : 75 "
            import re
            match = re.search(r"(\d+)", total_text)
            if match:
                total_orders = int(match.group(1))
            else:
                total_orders = 50 # Fallback default
                logger.warning("Could not parse total count, defaulting to 50")
                
            logger.info(f" Total Orders Found: {total_orders}")

        except Exception as e:
            logger.error(f" Error reading total count: {e}. Defaulting to single page processing.")
            total_orders = 50

        
        import math
        ORDERS_PER_PAGE = 50
        total_pages = math.ceil(total_orders / ORDERS_PER_PAGE)
        logger.info(f" Total Pages to process: {total_pages}")

        if total_pages < 2:
            print("only 50 order are exist")
            logger.info("only 50 order are exist")
            return

        restart_needed = False

        # ================== GO TO LAST PAGE ==================
        if total_pages > 1:
            try:
                logger.info(f"Navigating to the Last page (Page {total_pages})...")
                last_btn = WebDriverWait(driver, 15).until(
                    EC.element_to_be_clickable((
                        By.XPATH, 
                        "//a[contains(@class, 'cus_page_act') and contains(text(), 'Last')]"
                    ))
                )
                driver.execute_script("arguments[0].scrollIntoView({block:'center'});", last_btn)
                time.sleep(0.5)
                driver.execute_script("arguments[0].click();", last_btn)
                logger.info("Last button clicked. Waiting for page to load...")
                time.sleep(5)
            except Exception as e:
                print("only 50 order are exist")
                logger.info("only 50 order are exist")
                return

        # ================== PAGE LOOP (REVERSE) ==================
        for page in range(total_pages, 1, -1):
            if restart_needed:
                break
            logger.info(f"\\n --- STARTING PAGE {page} / {total_pages} (Reverse Order) ---")
            
            current_index = 0

            while True:
                # Check Global Limit (if set)
                if ROW_LIMIT > 0 and total_processed_global >= ROW_LIMIT:
                    logger.info(f" Global Limit of {ROW_LIMIT} rows reached.")
                    return

                if get_balance() < COST_PER_ORDER:
                    logger.info("🛑 Low Balance! Stopping automation.")
                    return

                logger.info(f"\\n --- Processing Row Index: {current_index} (Page {page}) ---")
                
                try:
                    # Call the function for the current index
                    result = click_truck_icons_one_by_one(current_index)

                    # Check for End of List on this page
                    if isinstance(result, dict) and result.get("status") == "END":
                        logger.info(f"End of rows on Page {page}.")
                        break
                    
                    # Retrieve Order ID safely
                    order_id_log = result.get("order_id", "Unknown") if isinstance(result, dict) else "N/A"

                    # Check Success
                    if isinstance(result, dict) and result.get("status") == "Processed":
                        logger.info(f" Row {current_index} processed successfully.")
                        log_to_excel(current_index, order_id_log, "SUCCESS", "Processed")
                        total_processed_global += 1
                        orders_since_restart += 1
                        update_balance(get_balance() - COST_PER_ORDER)
                        update_history(result)

                        if orders_since_restart >= 40:
                            logger.info(" 40 orders completed. Restarting Chrome browser as requested...")
                            try:
                                driver.quit()
                                logger.info(" Chrome browser closed.")
                            except:
                                pass
                            time.sleep(2)
                            driver = init_driver()
                            orders_since_restart = 0
                            restart_needed = True
                            break # Break row loop

                    # Check Skipped
                    elif isinstance(result, dict) and result.get("status") == "Skipped":
                        reason = result.get("reason", "Skipped")
                        logger.info(f" Row {current_index} skipped: {reason}")
                        log_to_excel(current_index, "N/A", "SKIPPED", reason)
                        update_history(result)
                    
                    elif isinstance(result, dict) and result.get("status") == "Error":
                        reason = result.get("reason", "Unknown Error")
                        order_id = result.get("order_id", "N/A")
                        logger.error(f" Row {current_index} failed: {reason} (Order: {order_id})")
                        log_to_excel(current_index, order_id, "ERROR", reason)
                        result["error"] = reason
                        update_history(result)
                    
                    else:
                        # Fallback for unexpected return
                        logger.warning(f" Row {current_index} skipped/unknown status.")
                        log_to_excel(current_index, "N/A", "UNKNOWN", "Unexpected return value")
                        
                    current_index += 1

                except Exception as e:
                    logger.error(f" Error processing row {current_index}: {e}")
                    import traceback
                    traceback.print_exc()
                    log_to_excel(current_index, "Error", "ERROR", str(e))
                    current_index += 1 # Skip on error
                    
                # ================== CLEANUP TABS ==================
                try:
                    while len(driver.window_handles) > 1:
                        driver.switch_to.window(driver.window_handles[-1])
                        logger.info(" Closing extra tab...")
                        driver.close()
                    
                    if len(driver.window_handles) > 0:
                        driver.switch_to.window(driver.window_handles[0])
                except Exception as cleanup_err:
                    logger.warning(f" Cleanup warning: {cleanup_err}")

            # ================== PREVIOUS PAGE LOGIC ==================
            if not restart_needed:
                if page > 2:
                    logger.info(f" Page {page} done. Moving to Page {page - 1} (Previous)...")
                    
                    try:
                        # Find PREVIOUS button
                        prev_btn = WebDriverWait(driver, 10).until(
                            EC.element_to_be_clickable((
                                By.XPATH, 
                                "//a[contains(@class, 'cus_page_act') and contains(text(), 'Previous')]"
                            ))
                        )
                        
                        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", prev_btn)
                        time.sleep(0.5)
                        driver.execute_script("arguments[0].click();", prev_btn)
                        logger.info(" Previous button clicked")
                        
                        # Wait for table to reload
                        logger.info(" Waiting for previous page to load...")
                        time.sleep(5) 
                        
                    except Exception as e:
                        logger.error(f" Could not click Previous button: {e}")
                        break
                else:
                    logger.info(" All pages processed.")
            else:
                break # Break page loop to restart from Row 1

        if not restart_needed:
            break # Exit while True loop if we finished everything without needing a restart

    logger.info(f"\n Workflow completed. Total rows processed: {total_processed_global}")



if __name__ == "__main__":
    if len(sys.argv) < 4:
        print("Usage: python m_happy_flow_final.py <username> <password> <otp> [target_order_id]")
        sys.exit(1)
        
    username = sys.argv[1]
    password = sys.argv[2]
    otp = sys.argv[3]
    target_order_id = sys.argv[4] if len(sys.argv) > 4 else None

    try:
        main_workflow(username, password, otp, target_order_id)
        logger.info("Workflow finished normally.")
    except Exception as e:
        logger.error(f"\n CRITICAL ERROR: {e}")
        import traceback
        traceback.print_exc()
    finally:
        logger.info("\n Script finished. Closing browser in 5 seconds...")
        time.sleep(5)
        try:
            driver.quit()
            logger.info("Chrome browser closed.")
        except:
            pass
